from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, make_response
import jwt as pyjwt
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3
import os
import json
from datetime import datetime, date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

import io
from PIL import Image as PILImage

from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-secret-key')
JWT_SECRET = os.environ.get('JWT_SECRET', 'change-this-jwt-secret')
JWT_EXPIRY_HOURS = 8

# ─── AUTH HELPERS ─────────────────────────────────────────────────────────────

# Default minimal permissions
MINIMAL_PERMS = {
    'can_view_orders': 0, 'can_edit_orders': 0, 'can_delete_orders': 0, 'can_approve_orders': 0,
    'can_view_quotes': 0, 'can_edit_quotes': 0, 'can_delete_quotes': 0,
    'can_view_reports': 0, 'can_view_vehicle_log': 0,
    'can_view_pool_vehicles': 0,
    'can_view_worktime': 0, 'can_edit_worktime': 0, 'can_confirm_worktime': 0, 'can_reopen_worktime': 0, 'can_copy_worktime': 0,
    'can_view_invoices': 0, 'can_edit_invoices': 0, 'can_liquidate_invoices': 0, 'can_edit_invoices_liquidated': 0,
    'can_view_loans': 0, 'can_edit_loans': 0, 'can_lock_loans': 0,
}
# Full admin permissions
ADMIN_PERMS = {k: 1 for k in MINIMAL_PERMS}

def get_user_permissions(user_id, is_admin):
    if is_admin:
        return ADMIN_PERMS.copy()
    conn = get_db()
    row = conn.execute('''SELECT p.* FROM profiles p
                          JOIN users u ON u.profile_id = p.id
                          WHERE u.id = ?''', (user_id,)).fetchone()
    # Provjeri ima li korisnik dodijeljeno vozilo 24/7
    urow = conn.execute("SELECT employee_id FROM users WHERE id=?", (user_id,)).fetchone()
    has_assigned_vehicle = False
    if urow and urow['employee_id']:
        v = conn.execute(
            "SELECT id FROM vehicles WHERE vehicle_type='assigned' AND assigned_employee_id=? LIMIT 1",
            (urow['employee_id'],)
        ).fetchone()
        has_assigned_vehicle = v is not None
    conn.close()
    if not row:
        perms = MINIMAL_PERMS.copy()
    else:
        perms = {}
        row_keys = row.keys()
        for k in MINIMAL_PERMS:
            perms[k] = row[k] if k in row_keys else 0
    # Ako ima dodijeljeno vozilo, automatski dobiva can_view_vehicle_log
    if has_assigned_vehicle:
        perms['can_view_vehicle_log'] = 1
    return perms

def create_token(user_id, username, is_admin, profile_id=None, display_name=None):
    perms = get_user_permissions(user_id, is_admin)
    payload = {
        'user_id': user_id,
        'username': username,
        'display_name': display_name or username,
        'is_admin': is_admin,
        'profile_id': profile_id,
        'exp': datetime.now().timestamp() + JWT_EXPIRY_HOURS * 3600,
        'iat': datetime.now().timestamp(),
        'auth_provider': 'local',
        **perms,
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm='HS256')

def decode_token(token):
    try:
        return pyjwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except pyjwt.ExpiredSignatureError:
        # Token je istekao — logiraj prekid sessije
        try:
            import pyjwt as _jwt
            data = _jwt.decode(token, JWT_SECRET, algorithms=['HS256'], options={"verify_exp": False})
            conn = get_db()
            conn.execute(
                """INSERT INTO audit_log (user_id, username, action, module, detail, created_at)
                   VALUES (?,?,?,?,?,?)""",
                (data.get('user_id'), data.get('username'), 'session_expired', 'auth',
                 f'Sessija istekla (JWT expiry {JWT_EXPIRY_HOURS}h)', datetime.now().isoformat())
            )
            conn.commit()
            conn.close()
        except: pass
        return None
    except:
        return None

def get_current_user():
    token = request.cookies.get('auth_token')
    if not token:
        return None
    return decode_token(token)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def require_perm(perm):
    """Decorator that checks a specific permission. Admins always pass."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect('/login')
            if user.get('is_admin') or user.get(perm):
                return f(*args, **kwargs)
            return render_template('403.html', user=user), 403
        return decorated
    return decorator

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect('/login')
        if not user.get('is_admin'):
            return render_template('403.html', user=user), 403
        return f(*args, **kwargs)
    return decorated

# Cache za settings koji se rijetko mijenjaju — osvježava se svaki put kad
# admin spremi postavke (via /api/settings), inače se čuva u memoriji.
_settings_cache = {}

def _get_cached_settings():
    """Vrati company_name i company_logo iz cachea. Puni cache ako je prazan."""
    if not _settings_cache:
        try:
            conn = get_db()
            row = conn.execute("SELECT value FROM settings WHERE key='company_name'").fetchone()
            logo_row = conn.execute("SELECT value FROM settings WHERE key='company_logo'").fetchone()
            conn.close()
            _settings_cache['company_name'] = row['value'] if row and row['value'] else 'MicroBusiness App'
            _settings_cache['company_logo'] = logo_row['value'] if logo_row and logo_row['value'] else 'logo.png'
        except:
            _settings_cache['company_name'] = 'MicroBusiness App'
            _settings_cache['company_logo'] = 'logo.png'
    return _settings_cache['company_name'], _settings_cache['company_logo']

def _invalidate_settings_cache():
    """Obriši cache — poziva se nakon spremanja postavki."""
    _settings_cache.clear()

@app.context_processor
def inject_current_user():
    user = get_current_user()
    company_name, company_logo = _get_cached_settings()
    return {'current_user': user or {}, 'company_name': company_name, 'company_logo': company_logo}

@app.template_filter('fmt_date')
def fmt_date(value):
    """Convert YYYY-MM-DD to DD.MM.YYYY."""
    if not value:
        return ''
    try:
        s = str(value)[:10]
        if len(s) == 10 and s[4] == '-':
            y, m, d = s.split('-')
            return f"{d}.{m}.{y}."
        return s
    except:
        return str(value)

@app.template_filter('fmt_datetime')
def fmt_datetime(value):
    """Convert YYYY-MM-DDTHH:MM to DD.MM.YYYY. HH:MM"""
    if not value:
        return ''
    try:
        s = str(value)
        date_part = s[:10]
        time_part = s[11:16] if len(s) > 10 else ''
        if len(date_part) == 10 and date_part[4] == '-':
            y, m, d = date_part.split('-')
            formatted = f"{d}.{m}.{y}."
            if time_part:
                formatted += f" {time_part}"
            return formatted
        return s
    except:
        return str(value)

DB_PATH = os.path.join(os.path.dirname(__file__), 'putni_nalog.db')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def get_vehicles_for_user(conn, user):
    """
    Vraća vozila ovisno o pravima korisnika:
    - Admin: sva vozila
    - Ima dodijeljeno vozilo 24/7: samo svoje vozilo
    - can_view_pool_vehicles: samo pool automobili
    - can_view_vehicle_log (pregled): sva vozila (samo za čitanje)
    """
    all_vehicles = [dict(v) for v in conn.execute("SELECT * FROM vehicles ORDER BY name").fetchall()]

    if not user:
        return all_vehicles

    if user.get('is_admin'):
        return all_vehicles

    user_employee_id = None
    if user.get('user_id'):
        urow = conn.execute("SELECT employee_id FROM users WHERE id=?", (user['user_id'],)).fetchone()
        if urow:
            user_employee_id = urow['employee_id']

    # Ima dodijeljeno vozilo 24/7 — vidi samo svoje
    if user_employee_id:
        assigned = [v for v in all_vehicles
                    if v.get('vehicle_type') == 'assigned'
                    and v.get('assigned_employee_id') == user_employee_id]
        if assigned:
            return assigned

    # can_view_pool_vehicles — samo pool auta (bez obzira na ostala prava)
    if user.get('can_view_pool_vehicles'):
        return [v for v in all_vehicles
                if not v.get('vehicle_type') or v.get('vehicle_type') == 'pool']

    # can_view_vehicle_log (voditelj/pregled) — sva vozila sortirana
    def sort_key(v):
        if v.get('vehicle_type') == 'assigned' and v.get('assigned_employee_id') == user_employee_id and user_employee_id:
            return 0
        if not v.get('vehicle_type') or v.get('vehicle_type') == 'pool':
            return 1
        return 2

    all_vehicles.sort(key=sort_key)
    return all_vehicles

def get_default_vehicle_for_user(conn, user):
    """Vraća vozilo dodijeljeno korisniku 24/7, ili None."""
    if not user or not user.get('user_id'):
        return None
    urow = conn.execute("SELECT employee_id FROM users WHERE id=?", (user['user_id'],)).fetchone()
    if not urow or not urow['employee_id']:
        return None
    veh = conn.execute(
        "SELECT * FROM vehicles WHERE vehicle_type='assigned' AND assigned_employee_id=? LIMIT 1",
        (urow['employee_id'],)
    ).fetchone()
    return dict(veh) if veh else None

def user_has_vehicle_log_access(user, conn):
    """
    Pristup modulu Službeni automobil (vidi listu):
    - Admin, ILI
    - can_view_vehicle_log (pregled), ILI
    - can_view_pool_vehicles (evidencija pool), ILI
    - Ima dodijeljen auto 24/7
    """
    if not user:
        return False
    if user.get('is_admin') or user.get('can_view_vehicle_log') or user.get('can_view_pool_vehicles'):
        return True
    assigned = get_default_vehicle_for_user(conn, user)
    return assigned is not None

def user_can_edit_vehicle_log(user, conn, log_id=None):
    """
    Može kreirati/uređivati evidenciju:
    - Admin: uvijek
    - Ima dodijeljeno vozilo 24/7: samo za svoju evidenciju
    - can_view_pool_vehicles: samo za evidencije pool automobila
    """
    if not user:
        return False
    if user.get('is_admin'):
        return True

    # Provjeri dodijeljeno vozilo
    assigned = get_default_vehicle_for_user(conn, user)
    if assigned:
        if log_id:
            # Može editirati samo evidenciju svog vozila
            log = conn.execute("SELECT vehicle_id FROM vehicle_log WHERE id=?", (log_id,)).fetchone()
            return log and log['vehicle_id'] == assigned['id']
        return True  # Nova evidencija — smije kreirati

    # can_view_pool_vehicles — samo pool auta
    if user.get('can_view_pool_vehicles'):
        if log_id:
            log = conn.execute(
                "SELECT v.vehicle_type FROM vehicle_log vl "
                "JOIN vehicles v ON v.id = vl.vehicle_id WHERE vl.id=?", (log_id,)
            ).fetchone()
            return log and (not log['vehicle_type'] or log['vehicle_type'] == 'pool')
        return True  # Nova evidencija — smije kreirati (vozilo će biti pool jer mu se nude samo pool)

    return False

def rows_to_dicts(rows):
    return [dict(r) for r in rows] if rows else []

def row_to_dict(row):
    return dict(row) if row else None

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.executescript('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            position TEXT,
            is_direktor INTEGER DEFAULT 0,
            is_blagajnik INTEGER DEFAULT 0,
            is_knjizio INTEGER DEFAULT 0,
            is_validator INTEGER DEFAULT 0,
            is_default INTEGER DEFAULT 0,
            signature_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS vehicles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            reg_plate TEXT,
            is_default INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS cost_centers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            name TEXT NOT NULL,
            is_default INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            is_default INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS destinations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            is_default INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS expense_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            name_en TEXT,
            is_default INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS report_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS travel_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            auto_id TEXT UNIQUE NOT NULL,
            status TEXT DEFAULT 'draft',
            issue_date TEXT,
            employee_id INTEGER,
            destination TEXT,
            purpose TEXT,
            client_info TEXT,
            expected_duration TEXT,
            departure_date TEXT,
            vehicle_id INTEGER,
            start_km INTEGER,
            end_km INTEGER,
            trip_start_datetime TEXT,
            trip_end_datetime TEXT,
            trip_duration_days INTEGER DEFAULT 0,
            trip_duration_hours INTEGER DEFAULT 0,
            trip_duration_minutes INTEGER DEFAULT 0,
            daily_allowance_count REAL DEFAULT 0,
            daily_allowance_rate REAL DEFAULT 0,
            daily_allowance_total REAL DEFAULT 0,
            advance_payment REAL DEFAULT 0,
            total_expenses REAL DEFAULT 0,
            total_amount REAL DEFAULT 0,
            payout_amount REAL DEFAULT 0,
            report_text TEXT,
            place_of_report TEXT,
            approved_by_id INTEGER,
            validator_id INTEGER,
            cost_center_id INTEGER,
            pdf_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            travel_order_id INTEGER NOT NULL,
            category_id INTEGER,
            description TEXT,
            paid_privately INTEGER DEFAULT 0,
            amount REAL DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (travel_order_id) REFERENCES travel_orders(id) ON DELETE CASCADE
        );
    ''')

    # Insert default data if tables are empty
    c.execute("SELECT COUNT(*) FROM employees")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO employees (name, position, is_direktor, is_blagajnik, is_default) VALUES (?, ?, 1, 1, 0)",
                  ("", ""))
        c.execute("INSERT INTO employees (name, position, is_default) VALUES (?, ?, 1)",
                  ("", ""))

    c.execute("SELECT COUNT(*) FROM vehicles")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO vehicles (name, reg_plate, is_default) VALUES (?, ?, 1)",
                  ("", ""))

    c.execute("SELECT COUNT(*) FROM expense_categories")
    if c.fetchone()[0] == 0:
        for cat in [("Gorivo", "Fuel"), ("Autocesta", "Toll"), ("Parking", "Parking"),
                    ("Smještaj", "Accommodation"), ("Ostalo", "Other")]:
            c.execute("INSERT INTO expense_categories (name, name_en) VALUES (?, ?)", cat)

    c.execute("SELECT COUNT(*) FROM destinations")
    if c.fetchone()[0] == 0:
        for dest in ["Zagreb", "Split", "Rijeka", "Osijek"]:
            c.execute("INSERT INTO destinations (name) VALUES (?)", (dest,))

    c.execute("SELECT value FROM settings WHERE key='daily_allowance_rate'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('daily_allowance_rate', '30')")
    c.execute("SELECT value FROM settings WHERE key='company_name'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_name', '')")
    c.execute("SELECT value FROM settings WHERE key='company_address'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_address', '')")
    c.execute("SELECT value FROM settings WHERE key='company_oib'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_oib', '')")
    c.execute("SELECT value FROM settings WHERE key='last_order_number'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('last_order_number', '12')")
    c.execute("SELECT value FROM settings WHERE key='last_order_year'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('last_order_year', '2026')")
    c.execute("SELECT value FROM settings WHERE key='default_place_of_report'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('default_place_of_report', 'Zagreb')")

    # Migration: car_logs table
    c.executescript('''
        CREATE TABLE IF NOT EXISTS car_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month INTEGER NOT NULL,
            year INTEGER NOT NULL,
            start_km REAL DEFAULT 0,
            end_km REAL DEFAULT 0,
            total_km REAL DEFAULT 0,
            official_km REAL DEFAULT 0,
            private_km REAL DEFAULT 0,
            pn_list TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')

    # Migration: add clients extra fields
    try:
        c.execute("ALTER TABLE clients ADD COLUMN address TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE clients ADD COLUMN oib TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE clients ADD COLUMN is_client INTEGER DEFAULT 0")
        # Existing clients are real clients
        c.execute("UPDATE clients SET is_client=1")
    except: pass


    # Migration: create quotes table
    c.executescript('''
        CREATE TABLE IF NOT EXISTS quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            auto_id TEXT UNIQUE NOT NULL,
            status TEXT DEFAULT 'draft',
            issue_date TEXT,
            valid_days INTEGER DEFAULT 7,
            place_of_issue TEXT DEFAULT 'Zagreb',
            client_id INTEGER,
            prepared_by_id INTEGER,
            notes TEXT,
            pdv_rate REAL DEFAULT 25,
            total_net REAL DEFAULT 0,
            total_pdv REAL DEFAULT 0,
            total_gross REAL DEFAULT 0,
            pdf_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS quote_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id INTEGER NOT NULL,
            sort_order INTEGER DEFAULT 0,
            description TEXT,
            unit TEXT DEFAULT 'Kom',
            quantity REAL DEFAULT 1,
            unit_price REAL DEFAULT 0,
            total REAL DEFAULT 0,
            FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE
        );
    ''')

    # Default settings for quotes
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_view_loans INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_edit_loans INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_lock_loans INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE loans ADD COLUMN is_locked INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE loans ADD COLUMN schedule_json TEXT")
    except: pass
    try:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS work_fund (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                fond INTEGER DEFAULT 0,
                radni INTEGER DEFAULT 0,
                neradni INTEGER DEFAULT 0,
                obracunskih INTEGER DEFAULT 0,
                radnih_dana INTEGER DEFAULT 0,
                neradnih_dana INTEGER DEFAULT 0,
                UNIQUE(year, month)
            );
        """)
    except: pass
    try:
        # Uvezi postojeće WORK_FUND_2026 podatke ako tablica još nema 2026
        existing = c.execute("SELECT COUNT(*) FROM work_fund WHERE year=2026").fetchone()[0]
        if existing == 0:
            wf2026 = [
                (2026,1,176,160,16,22,20,2),(2026,2,160,160,0,20,20,0),(2026,3,176,176,0,22,22,0),
                (2026,4,176,168,8,22,21,1),(2026,5,168,160,8,21,20,1),(2026,6,176,160,16,22,20,2),
                (2026,7,184,184,0,23,23,0),(2026,8,168,160,8,21,20,1),(2026,9,176,176,0,22,22,0),
                (2026,10,176,176,0,22,22,0),(2026,11,168,160,8,21,20,1),(2026,12,184,176,8,23,22,1),
            ]
            c.executemany("INSERT OR IGNORE INTO work_fund (year,month,fond,radni,neradni,obracunskih,radnih_dana,neradnih_dana) VALUES (?,?,?,?,?,?,?,?)", wf2026)
    except: pass
    try:
        c.execute("ALTER TABLE loan_payments ADD COLUMN category TEXT DEFAULT 'glavnica'")
    except: pass
    try:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS vehicle_log_days (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                start_km REAL DEFAULT 0,
                end_km REAL DEFAULT 0,
                official_km REAL DEFAULT 0,
                private_km REAL DEFAULT 0,
                total_km REAL DEFAULT 0,
                comment TEXT DEFAULT '',
                is_pn INTEGER DEFAULT 0,
                trips_json TEXT DEFAULT NULL,
                UNIQUE(log_id, date)
            );
        """)
    except: pass
    # Loans module tables
    try:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS loans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                total_amount REAL NOT NULL,
                loan_date TEXT,
                interest_rate REAL DEFAULT 0,
                repayment_start TEXT,
                repayment_end TEXT,
                notes TEXT,
                schedule_json TEXT,
                is_locked INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS loan_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                loan_id INTEGER NOT NULL REFERENCES loans(id) ON DELETE CASCADE,
                payment_type TEXT NOT NULL CHECK(payment_type IN ('one_time','recurring','conversion')),
                amount REAL NOT NULL,
                payment_date TEXT,
                recurring_day INTEGER,
                recurring_start TEXT,
                recurring_end TEXT,
                description TEXT,
                category TEXT DEFAULT 'glavnica',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
    except: pass

    # Invoice storage path default
    try:
        r = c.execute("SELECT value FROM settings WHERE key='invoice_storage_path'").fetchone()
        if not r:
            import os as _os
            _def = _os.path.join(_os.path.dirname(__file__), 'uploads', 'racuni')
            c.execute("INSERT INTO settings (key, value) VALUES ('invoice_storage_path', ?)", (_def,))
    except: pass

    c.execute("SELECT value FROM settings WHERE key='last_quote_number'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('last_quote_number', '0')")
    c.execute("SELECT value FROM settings WHERE key='last_quote_year'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('last_quote_year', '2024')")
    c.execute("SELECT value FROM settings WHERE key='company_phone'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_phone', '')")
    c.execute("SELECT value FROM settings WHERE key='company_email'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_email', '')")
    c.execute("SELECT value FROM settings WHERE key='company_iban'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_iban', '')")
    c.execute("SELECT value FROM settings WHERE key='company_bank'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_bank', '')")
    c.execute("SELECT value FROM settings WHERE key='company_mbs'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_mbs', '')")
    c.execute("SELECT value FROM settings WHERE key='company_capital'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_capital', '')")
    c.execute("SELECT value FROM settings WHERE key='company_director'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('company_director', '')")
    c.execute("SELECT value FROM settings WHERE key='quote_pdv_rate'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('quote_pdv_rate', '25')")
    c.execute("SELECT value FROM settings WHERE key='quote_default_valid_days'")
    if not c.fetchone():
        c.execute("INSERT INTO settings (key, value) VALUES ('quote_default_valid_days', '7')")
    c.execute("SELECT value FROM settings WHERE key='quote_default_notes'")
    if not c.fetchone():
        default_notes = '- Ovo nije fiskalizirani racun\n- Fakturiranje po izvršenoj usluzi'
        c.execute("INSERT INTO settings (key, value) VALUES ('quote_default_notes', ?)", (default_notes,))

    # Migration: add comment to quotes
    try:
        c.execute("ALTER TABLE quotes ADD COLUMN comment TEXT")
    except: pass

    # Migration: add notes column to quotes if missing
    try:
        c.execute("ALTER TABLE quotes ADD COLUMN comment TEXT")
    except: pass

    # Migration: add trips_json to vehicle_log_days
    try:
        c.execute("ALTER TABLE vehicle_log_days ADD COLUMN trips_json TEXT DEFAULT NULL")
    except: pass

    # Migration: create users table
    c.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            employee_id INTEGER,
            display_name TEXT,
            email TEXT,
            auth_provider TEXT DEFAULT 'local',
            ad_object_id TEXT,
            last_login TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    # Create default admin user if no users exist
    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        from werkzeug.security import generate_password_hash
        c.execute('''INSERT INTO users (username, password_hash, is_admin, display_name, email, auth_provider)
                     VALUES (?, ?, 1, ?, ?, 'local')''',
                  ('admin', generate_password_hash('changeme123', method='pbkdf2:sha256'), 'Administrator', ''))

    # Migration: create profiles table
    c.executescript('''
        CREATE TABLE IF NOT EXISTS profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            can_view_orders INTEGER DEFAULT 0,
            can_edit_orders INTEGER DEFAULT 0,
            can_delete_orders INTEGER DEFAULT 0,
            can_approve_orders INTEGER DEFAULT 0,
            can_view_quotes INTEGER DEFAULT 0,
            can_edit_quotes INTEGER DEFAULT 0,
            can_delete_quotes INTEGER DEFAULT 0,
            can_view_reports INTEGER DEFAULT 0,
            can_view_vehicle_log INTEGER DEFAULT 0,
            can_edit_vehicle_log INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    # Add profile_id to users if missing
    try:
        c.execute("ALTER TABLE users ADD COLUMN profile_id INTEGER")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_view_worktime INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_edit_worktime INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_confirm_worktime INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_reopen_worktime INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE worktime_reports ADD COLUMN submitted_at TIMESTAMP")
    except: pass
    try:
        c.execute("ALTER TABLE worktime_reports ADD COLUMN submitted_by INTEGER")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_copy_worktime INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_view_invoices INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_edit_invoices INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_liquidate_invoices INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_edit_invoices_liquidated INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE vehicle_log ADD COLUMN is_approved INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE vehicle_log ADD COLUMN approved_at TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE vehicle_log ADD COLUMN approved_by_id INTEGER")
    except: pass
    try:
        c.execute("ALTER TABLE vehicle_log ADD COLUMN employee_signature_path TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE vehicles ADD COLUMN vehicle_type TEXT DEFAULT 'pool'")
    except: pass
    try:
        c.execute("ALTER TABLE vehicles ADD COLUMN assigned_employee_id INTEGER DEFAULT NULL")
    except: pass
    try:
        c.execute("ALTER TABLE vehicles ADD COLUMN home_address TEXT DEFAULT NULL")
    except: pass
    try:
        c.execute("ALTER TABLE vehicles ADD COLUMN home_city TEXT DEFAULT NULL")
    except: pass
    try:
        c.execute("ALTER TABLE profiles ADD COLUMN can_view_pool_vehicles INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE invoices ADD COLUMN is_deleted INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE invoices ADD COLUMN deleted_at TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE invoices ADD COLUMN created_by INTEGER")
    except: pass
    # Suppliers tablica
    try:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                oib TEXT,
                address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
    except: pass

    # Invoice tables
    c.executescript('''
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT,
            partner_name TEXT,
            partner_oib TEXT,
            invoice_date TEXT,
            due_date TEXT,
            amount_total REAL,
            currency TEXT DEFAULT 'EUR',
            original_filename TEXT,
            stored_filename TEXT,
            stored_path TEXT,
            is_paid INTEGER DEFAULT 0,
            paid_at TEXT,
            paid_card_last4 TEXT,
            is_liquidated INTEGER DEFAULT 0,
            liquidated_at TEXT,
            liquidated_by INTEGER,
            liquidated_pdf_path TEXT,
            notes TEXT,
            ocr_raw TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS bank_cards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_name TEXT NOT NULL,
            last4 TEXT NOT NULL,
            card_type TEXT,
            is_active INTEGER DEFAULT 1,
            notes TEXT
        );
    ''')
    # Create default profiles if none exist
    if c.execute("SELECT COUNT(*) FROM profiles").fetchone()[0] == 0:
        c.execute('''INSERT INTO profiles (name, description,
            can_view_orders, can_edit_orders, can_delete_orders, can_approve_orders,
            can_view_quotes, can_edit_quotes, can_delete_quotes,
            can_view_reports, can_view_vehicle_log, can_edit_vehicle_log)
            VALUES (?, ?, 1, 1, 0, 0, 1, 1, 0, 0, 0, 0)''',
            ('Zaposlenik', 'Kreiranje i pregled putnih naloga i ponuda'))
        c.execute('''INSERT INTO profiles (name, description,
            can_view_orders, can_edit_orders, can_delete_orders, can_approve_orders,
            can_view_quotes, can_edit_quotes, can_delete_quotes,
            can_view_reports, can_view_vehicle_log, can_edit_vehicle_log)
            VALUES (?, ?, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1)''',
            ('Voditelj', 'Puni pristup svim modulima osim postavki'))

    # Migration: add must_change_password to users
    try:
        c.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER DEFAULT 0")
    except: pass

    # Migration: create worktime_reports table
    c.executescript('''
        CREATE TABLE IF NOT EXISTS worktime_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            status TEXT DEFAULT 'draft',
            submitted_at TIMESTAMP,
            submitted_by INTEGER,
            confirmed_at TIMESTAMP,
            confirmed_by INTEGER,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(employee_id, year, month)
        );
        CREATE TABLE IF NOT EXISTS worktime_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL,
            day INTEGER NOT NULL,
            row_num INTEGER NOT NULL,
            hours REAL DEFAULT 0,
            FOREIGN KEY (report_id) REFERENCES worktime_reports(id) ON DELETE CASCADE
        );
    ''')

    # Migration: create vehicle_log table
    c.executescript('''
        CREATE TABLE IF NOT EXISTS vehicle_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_id INTEGER,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            start_km REAL DEFAULT 0,
            end_km REAL DEFAULT 0,
            total_km REAL DEFAULT 0,
            official_km REAL DEFAULT 0,
            private_km REAL DEFAULT 0,
            notes TEXT,
            csv_filename TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(vehicle_id, year, month)
        );
    ''')

    # Migration: add pdf_path column if missing
    try:
        c.execute("ALTER TABLE travel_orders ADD COLUMN pdf_path TEXT")
    except:
        pass
    # Migration: add is_place_of_report to destinations
    try:
        c.execute("ALTER TABLE destinations ADD COLUMN is_place_of_report INTEGER DEFAULT 0")
        # Set Zagreb as default place of report
        c.execute("UPDATE destinations SET is_place_of_report=1 WHERE name='Zagreb'")
    except:
        pass
    # Migration: add is_deleted column if missing
    try:
        c.execute("ALTER TABLE travel_orders ADD COLUMN is_deleted INTEGER DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE travel_orders ADD COLUMN deleted_at TIMESTAMP")
    except:
        pass
    try:
        c.execute("ALTER TABLE travel_orders ADD COLUMN is_paid INTEGER DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE travel_orders ADD COLUMN paid_at DATE")
    except:
        pass

    # Audit log tablica
    try:
        c.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                action TEXT NOT NULL,
                module TEXT,
                entity TEXT,
                entity_id INTEGER,
                detail TEXT,
                ip_address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
    except: pass
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_log(user_id)")
    except: pass
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_log(created_at)")
    except: pass

    conn.commit()
    conn.close()

def audit(action, module=None, entity=None, entity_id=None, detail=None):
    """Logiraj akciju trenutnog korisnika."""
    try:
        from flask import request as _req
        user = get_current_user()
        user_id = user.get('user_id') if user else None
        username = user.get('username') if user else 'nepoznat'
        try:
            ip = _req.headers.get('X-Forwarded-For', _req.remote_addr or '').split(',')[0].strip()
        except:
            ip = None
        conn = get_db()
        conn.execute(
            """INSERT INTO audit_log (user_id, username, action, module, entity, entity_id, detail, ip_address, created_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (user_id, username, action, module, entity, entity_id, detail, ip, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
    except Exception as e:
        pass  # Audit nikad ne smije srušiti aplikaciju

def get_next_auto_id():
    conn = get_db()
    c = conn.cursor()
    current_year = datetime.now().year

    # Get ALL used auto_ids (including deleted) to avoid UNIQUE constraint
    rows = c.execute(
        "SELECT auto_id FROM travel_orders WHERE auto_id LIKE ?",
        (f"{current_year}-%",)
    ).fetchall()
    used_nums = set()
    for row in rows:
        try:
            aid = row['auto_id']
            if aid.endswith('-I'):
                continue  # soft-deleted, broj je slobodan za ponovnu upotrebu
            num = int(aid.split('-')[1])
            used_nums.add(num)
        except: pass

    # Find lowest free number not used by any active (non-deleted) order
    candidate = 1
    while candidate in used_nums:
        candidate += 1

    next_num = candidate
    auto_id = f"{current_year}-{next_num}"
    conn.close()
    return auto_id, current_year, next_num

def _calc_loan_repaid(loan, payments):
    """Izračunaj ukupno otplaćeno do danas."""
    today = date.today().isoformat()
    total = 0.0
    for p in payments:
        if isinstance(p, dict):
            ptype = p.get('payment_type')
            amount = float(p.get('amount') or 0)
            pdate = p.get('payment_date') or ''
            rstart = p.get('recurring_start') or ''
            rend = p.get('recurring_end') or today
            rday = p.get('recurring_day') or 1
        else:
            ptype = p['payment_type']
            amount = float(p['amount'] or 0)
            pdate = p['payment_date'] or ''
            rstart = p['recurring_start'] or ''
            rend = p['recurring_end'] or today
            rday = p['recurring_day'] or 1

        if ptype in ('one_time', 'conversion'):
            if pdate and pdate <= today:
                total += amount
        elif ptype == 'recurring':
            if rstart:
                import calendar
                try:
                    sy, sm = int(rstart[:4]), int(rstart[5:7])
                    ey, em = int(min(rend, today)[:4]), int(min(rend, today)[5:7])
                    d = date(sy, sm, 1)
                    end_d = date(ey, em, 1)
                    while d <= end_d:
                        total += amount
                        m = d.month + 1
                        y = d.year + (m > 12)
                        m = m if m <= 12 else 1
                        d = date(y, m, 1)
                except: pass
    return round(total, 2)


def calculate_dnevnice(start_dt_str, end_dt_str, rate):
    try:
        start_dt = datetime.fromisoformat(start_dt_str)
        end_dt = datetime.fromisoformat(end_dt_str)
        delta = end_dt - start_dt
        total_minutes = int(delta.total_seconds() / 60)
        total_hours = total_minutes // 60
        remaining_minutes = total_minutes % 60
        total_days = total_hours // 24
        remaining_hours = total_hours % 24

        # Calculate dnevnice
        dnevnice = total_days  # full days
        if remaining_hours < 8:
            pass  # no extra dnevnica
        elif remaining_hours <= 12:
            dnevnice += 0.5
        else:
            dnevnice += 1.0

        return {
            'days': total_days,
            'hours': remaining_hours,
            'minutes': remaining_minutes,
            'dnevnice': dnevnice,
            'total': round(dnevnice * rate, 2)
        }
    except:
        return {'days': 0, 'hours': 0, 'minutes': 0, 'dnevnice': 0, 'total': 0}

# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        user = get_current_user()
        if user:
            return redirect('/')
        return render_template('login.html')

    data = request.json if request.is_json else request.form
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()
    conn.close()

    if not user or not check_password_hash(user['password_hash'], password):
        # Logiraj neuspješan pokušaj
        try:
            ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
            conn2 = get_db()
            conn2.execute(
                """INSERT INTO audit_log (username, action, module, detail, ip_address, created_at)
                   VALUES (?,?,?,?,?,?)""",
                (username, 'login_failed', 'auth', f'Neuspješan pokušaj prijave', ip, datetime.now().isoformat())
            )
            conn2.commit()
            conn2.close()
        except: pass
        if request.is_json:
            return jsonify({'error': 'Pogrešno korisničko ime ili lozinka'}), 401
        return render_template('login.html', error='Pogrešno korisničko ime ili lozinka')

    # Update last login
    conn = get_db()
    conn.execute("UPDATE users SET last_login=? WHERE id=?", (datetime.now().isoformat(), user['id']))
    conn.commit()
    conn.close()

    must_change = bool(user['must_change_password']) if 'must_change_password' in user.keys() else False
    token = create_token(user['id'], user['username'], bool(user['is_admin']),
                         user['profile_id'] if 'profile_id' in user.keys() else None,
                         user['display_name'] if 'display_name' in user.keys() else None)
    redirect_url = '/change-password' if must_change else '/'
    resp = make_response(jsonify({'success': True, 'redirect': redirect_url, 'must_change_password': must_change}))
    resp.set_cookie('auth_token', token, httponly=True, samesite='Lax',
                   max_age=JWT_EXPIRY_HOURS * 3600)
    audit('login', module='auth', detail=f'Prijava s IP: {request.headers.get("X-Forwarded-For", request.remote_addr or "")}')
    return resp

@app.route('/change-password', methods=['GET'])
@login_required
def change_password_page():
    return render_template('change_password.html')

@app.route('/logout')
def logout():
    audit('logout', module='auth', detail='Korisnik se odjavio')
    resp = make_response(redirect('/login'))
    resp.delete_cookie('auth_token')
    return resp

@app.route('/api/auth/me')
@login_required
def auth_me():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Not authenticated'}), 401
    return jsonify({
        'user_id': user['user_id'],
        'username': user['username'],
        'is_admin': user['is_admin'],
        'auth_provider': user.get('auth_provider', 'local')
    })

@app.route('/')
@login_required
def index():
    user = get_current_user()
    if not user:
        return redirect('/login')

    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    now_month = datetime.now().month
    now_year = datetime.now().year

    # Dohvati display_name iz baze (token može biti star i ne sadržavati ga)
    if user.get('user_id'):
        urow = conn.execute("SELECT display_name FROM users WHERE id=?", (user['user_id'],)).fetchone()
        if urow and urow['display_name']:
            user = dict(user)
            user['display_name'] = urow['display_name']

    # Ulazni računi
    inv_total = conn.execute("SELECT COUNT(*) FROM invoices WHERE is_deleted=0 OR is_deleted IS NULL").fetchone()[0]
    inv_pending = conn.execute("SELECT COUNT(*) FROM invoices WHERE is_liquidated=0 AND (is_deleted=0 OR is_deleted IS NULL)").fetchone()[0]
    inv_overdue = conn.execute("""SELECT partner_name, due_date, amount_total FROM invoices
        WHERE is_liquidated=0 AND due_date!='' AND due_date<=? AND (is_deleted=0 OR is_deleted IS NULL)
        ORDER BY due_date LIMIT 3""", (today,)).fetchall()
    inv_upcoming = conn.execute("""SELECT partner_name, due_date, amount_total FROM invoices
        WHERE is_liquidated=0 AND due_date>? AND (is_deleted=0 OR is_deleted IS NULL)
        ORDER BY due_date LIMIT 3""", (today,)).fetchall()

    # Ponude
    quotes_open = conn.execute("SELECT COUNT(*) FROM quotes WHERE status NOT IN ('accepted','rejected','cancelled')").fetchone()[0]
    quotes_accepted = conn.execute("SELECT COUNT(*) FROM quotes WHERE status='accepted'").fetchone()[0]
    quotes_recent = conn.execute("""SELECT q.auto_id, q.status, c.name as client_name, q.total_gross
        FROM quotes q LEFT JOIN clients c ON c.id=q.client_id
        ORDER BY q.updated_at DESC LIMIT 3""").fetchall()

    # Putni nalozi
    orders_total = conn.execute("SELECT COUNT(*) FROM travel_orders").fetchone()[0]
    orders_pending = conn.execute("SELECT COUNT(*) FROM travel_orders WHERE status='submitted'").fetchone()[0]

    # Radno vrijeme
    wt_reports = conn.execute("""SELECT wr.status, e.name as employee_name
        FROM worktime_reports wr LEFT JOIN employees e ON e.id=wr.employee_id
        WHERE wr.month=? AND wr.year=?
        ORDER BY e.name""", (now_month, now_year)).fetchall()

    # Pozajmice
    import json as _json
    loans_raw = conn.execute("SELECT * FROM loans ORDER BY total_amount DESC").fetchall()
    loans_data = []
    for loan in loans_raw:
        ld = row_to_dict(loan)
        schedule = []
        try: schedule = _json.loads(ld.get('schedule_json') or '[]')
        except: pass
        if schedule:
            repaid = sum(float(s.get('amount',0)) for s in schedule if s.get('paid') and s.get('type')!='conversion')
            converted = sum(float(s.get('amount',0)) for s in schedule if s.get('paid') and s.get('type')=='conversion')
        else:
            payments = conn.execute("SELECT * FROM loan_payments WHERE loan_id=?", (loan['id'],)).fetchall()
            repaid = _calc_loan_repaid(ld, payments)
            converted = sum(float(p['amount'] or 0) for p in payments
                           if p['payment_type']=='conversion' and (p['payment_date'] or '') <= today)
        ld['repaid'] = round(repaid + converted, 2)
        ld['remaining'] = round(loan['total_amount'] - ld['repaid'], 2)
        ld['pct'] = round(ld['repaid'] / loan['total_amount'] * 100, 1) if loan['total_amount'] else 0
        loans_data.append(ld)

    total_loans_remaining = sum(l['remaining'] for l in loans_data)

    # Audit log — nedavna aktivnost
    activity = conn.execute("""SELECT al.action, al.module, al.detail, al.created_at, u.username
        FROM audit_log al LEFT JOIN users u ON u.id=al.user_id
        ORDER BY al.created_at DESC LIMIT 5""").fetchall()

    # Sljedeći blagdan
    next_holiday = None
    for date_str in sorted(ALL_HOLIDAYS.keys()):
        if date_str > today:
            from datetime import date as _date
            hd = _date(int(date_str[:4]), int(date_str[5:7]), int(date_str[8:10]))
            td = _date.today()
            days_left = (hd - td).days
            next_holiday = {'date': date_str, 'name': ALL_HOLIDAYS[date_str], 'days_left': days_left,
                           'day': date_str[8:10], 'month_short': ['Sij','Velj','Ožu','Tra','Svi','Lip','Srp','Kol','Ruj','Lis','Stu','Pro'][int(date_str[5:7])-1]}
            break

    conn.close()

    return render_template('dashboard.html',
        user=user, active='dashboard',
        inv_total=inv_total, inv_pending=inv_pending,
        inv_overdue=rows_to_dicts(inv_overdue), inv_upcoming=rows_to_dicts(inv_upcoming),
        quotes_open=quotes_open, quotes_accepted=quotes_accepted,
        quotes_recent=rows_to_dicts(quotes_recent),
        orders_total=orders_total, orders_pending=orders_pending,
        wt_reports=rows_to_dicts(wt_reports),
        loans=loans_data, total_loans_remaining=total_loans_remaining,
        activity=rows_to_dicts(activity),
        next_holiday=next_holiday,
        now_month=now_month, now_year=now_year,
        months=MONTHS_HR)

@app.route('/orders')
@require_perm('can_view_orders')
def orders_list():
    audit('view', module='putni_nalozi', entity='list')
    conn = get_db()
    orders = conn.execute('''
        SELECT to2.*, e.name as employee_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        WHERE to2.is_deleted = 0 OR to2.is_deleted IS NULL
        ORDER BY to2.id DESC
    ''').fetchall()
    employees_all = conn.execute("SELECT id, name FROM employees ORDER BY name").fetchall()
    clients_all = conn.execute("SELECT id, name, oib, address FROM clients ORDER BY name").fetchall()
    templates_all = conn.execute("SELECT name, content FROM report_templates ORDER BY name").fetchall()
    default_place_row = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    default_dest_row = conn.execute("SELECT name FROM destinations WHERE is_default=1 LIMIT 1").fetchone()
    directors = conn.execute("SELECT id FROM employees WHERE is_direktor=1 LIMIT 1").fetchall()
    validators = conn.execute("SELECT id FROM employees WHERE is_validator=1 LIMIT 1").fetchall()
    default_veh = get_default_vehicle_for_user(conn, get_current_user())
    if not default_veh:
        default_veh = conn.execute("SELECT * FROM vehicles WHERE vehicle_type='pool' OR vehicle_type IS NULL LIMIT 1").fetchone()
        default_veh = dict(default_veh) if default_veh else None
    conn.close()
    return render_template('orders.html', orders=orders, show_trash=False,
                          employees_all=rows_to_dicts(employees_all),
                          clients_all=rows_to_dicts(clients_all),
                          templates_all=rows_to_dicts(templates_all),
                          default_place_setting=default_place_row['name'] if default_place_row else 'Zagreb',
                          default_destination_name=default_dest_row['name'] if default_dest_row else '',
                          directors=rows_to_dicts(directors),
                          validators=rows_to_dicts(validators),
                          default_vehicle_id=default_veh['id'] if default_veh else None)

@app.route('/orders/trash')
@require_perm('can_view_orders')
def orders_trash():
    conn = get_db()
    orders = conn.execute('''
        SELECT to2.*, e.name as employee_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        WHERE to2.is_deleted = 1
        ORDER BY to2.deleted_at DESC
    ''').fetchall()
    employees_all = conn.execute("SELECT id, name FROM employees ORDER BY name").fetchall()
    clients_all = conn.execute("SELECT id, name, oib, address FROM clients ORDER BY name").fetchall()
    templates_all = conn.execute("SELECT name, content FROM report_templates ORDER BY name").fetchall()
    default_place_row = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    default_dest_row = conn.execute("SELECT name FROM destinations WHERE is_default=1 LIMIT 1").fetchone()
    directors = conn.execute("SELECT id FROM employees WHERE is_direktor=1 LIMIT 1").fetchall()
    validators = conn.execute("SELECT id FROM employees WHERE is_validator=1 LIMIT 1").fetchall()
    default_veh = get_default_vehicle_for_user(conn, get_current_user())
    if not default_veh:
        default_veh = conn.execute("SELECT * FROM vehicles WHERE vehicle_type='pool' OR vehicle_type IS NULL LIMIT 1").fetchone()
        default_veh = dict(default_veh) if default_veh else None
    conn.close()
    return render_template('orders.html', orders=orders, show_trash=True, active='trash',
                          employees_all=rows_to_dicts(employees_all),
                          clients_all=rows_to_dicts(clients_all),
                          templates_all=rows_to_dicts(templates_all),
                          default_place_setting=default_place_row['name'] if default_place_row else 'Zagreb',
                          default_destination_name=default_dest_row['name'] if default_dest_row else '',
                          directors=rows_to_dicts(directors),
                          validators=rows_to_dicts(validators),
                          default_vehicle_id=default_veh['id'] if default_veh else None)

@app.route('/orders/new')
@require_perm('can_edit_orders')
def new_order():
    conn = get_db()
    user = get_current_user()
    auto_id, year, num = get_next_auto_id()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    vehicles = get_vehicles_for_user(conn, user)
    destinations = conn.execute("SELECT * FROM destinations ORDER BY name").fetchall()
    categories = conn.execute("SELECT * FROM expense_categories ORDER BY name").fetchall()
    templates = conn.execute("SELECT * FROM report_templates ORDER BY name").fetchall()
    rate_row = conn.execute("SELECT value FROM settings WHERE key='daily_allowance_rate'").fetchone()
    rate = float(rate_row['value']) if rate_row else 30.0

    default_employee = conn.execute("SELECT * FROM employees WHERE is_default=1 LIMIT 1").fetchone()
    default_vehicle = get_default_vehicle_for_user(conn, user)
    default_destination = conn.execute("SELECT * FROM destinations WHERE is_default=1 LIMIT 1").fetchone()

    validators = conn.execute("SELECT * FROM employees WHERE is_validator=1").fetchall()
    directors = conn.execute("SELECT * FROM employees WHERE is_direktor=1").fetchall()
    clients = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()

    default_place = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    conn.close()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('form.html',
        auto_id=auto_id, today=today,
        employees=rows_to_dicts(employees),
        vehicles=vehicles,
        destinations=rows_to_dicts(destinations),
        categories=rows_to_dicts(categories),
        templates=rows_to_dicts(templates),
        clients=rows_to_dicts(clients),
        rate=rate,
        default_employee=row_to_dict(default_employee),
        default_vehicle=default_vehicle,
        default_destination=row_to_dict(default_destination),
        default_place=default_place['name'] if default_place else 'Zagreb',
        validators=rows_to_dicts(validators),
        directors=rows_to_dicts(directors),
        order=None, expenses=[])

@app.route('/orders/<int:order_id>/edit')
@require_perm('can_view_orders')
def edit_order(order_id):
    conn = get_db()
    user = get_current_user()
    order = conn.execute("SELECT * FROM travel_orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for('orders_list'))
    expenses = conn.execute("SELECT * FROM expenses WHERE travel_order_id=? ORDER BY sort_order", (order_id,)).fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    vehicles = get_vehicles_for_user(conn, user)
    destinations = conn.execute("SELECT * FROM destinations ORDER BY name").fetchall()
    categories = conn.execute("SELECT * FROM expense_categories ORDER BY name").fetchall()
    templates = conn.execute("SELECT * FROM report_templates ORDER BY name").fetchall()
    rate_row = conn.execute("SELECT value FROM settings WHERE key='daily_allowance_rate'").fetchone()
    rate = float(rate_row['value']) if rate_row else 30.0
    validators = conn.execute("SELECT * FROM employees WHERE is_validator=1").fetchall()
    directors = conn.execute("SELECT * FROM employees WHERE is_direktor=1").fetchall()
    clients_list = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
    default_place = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    conn.close()
    today = date.today().strftime('%Y-%m-%d')
    is_deleted = bool(order['is_deleted'])
    return render_template('form.html',
        auto_id=order['auto_id'], today=today,
        employees=rows_to_dicts(employees),
        vehicles=rows_to_dicts(vehicles),
        destinations=rows_to_dicts(destinations),
        categories=rows_to_dicts(categories),
        templates=rows_to_dicts(templates),
        clients=rows_to_dicts(clients_list),
        rate=rate,
        default_employee=None, default_vehicle=None, default_destination=None,
        default_place=default_place['name'] if default_place else 'Zagreb',
        validators=rows_to_dicts(validators),
        directors=rows_to_dicts(directors),
        order=dict(order), expenses=rows_to_dicts(expenses),
        is_deleted=is_deleted)

@app.route('/api/orders', methods=['POST'])
@login_required
def save_order():
    data = request.json
    conn = get_db()
    c = conn.cursor()

    auto_id = data.get('auto_id')
    order_id = data.get('id')
    new_status = data.get('status', 'draft')

    # Check if existing order is locked (submitted or approved)
    if order_id:
        existing = conn.execute("SELECT status FROM travel_orders WHERE id=?", (order_id,)).fetchone()
        if existing:
            old_status = existing['status']
            # Only allow status changes from locked states, not data edits
            locked_statuses = ['submitted', 'approved']
            if old_status in locked_statuses:
                # Only allow: approved->rejected, submitted->rejected, submitted->approved
                allowed_transitions = {
                    'submitted': ['approved', 'rejected'],
                    'approved': ['rejected', 'knjizeno'],
                    'knjizeno': ['rejected'],
                }
                if new_status not in allowed_transitions.get(old_status, []):
                    conn.close()
                    return jsonify({'error': f'Cannot edit order with status: {old_status}'}), 403
                # Status-only change - just update status
                conn.execute("UPDATE travel_orders SET status=?, updated_at=? WHERE id=?",
                           (new_status, datetime.now().isoformat(), order_id))
                # If approving - generate and save PDF
                if new_status == 'approved':
                    conn.commit()
                    conn.close()
                    return _generate_and_save_pdf(order_id)
                conn.commit()
                conn.close()
                return jsonify({'success': True, 'id': order_id, 'auto_id': auto_id, 'status': new_status})

    # Calculate dnevnice
    rate_row = conn.execute("SELECT value FROM settings WHERE key='daily_allowance_rate'").fetchone()
    rate = float(rate_row['value']) if rate_row else 30.0
    calc = calculate_dnevnice(data.get('trip_start_datetime',''), data.get('trip_end_datetime',''), rate)

    # Calculate totals
    expenses = data.get('expenses', [])
    private_total = sum(float(e.get('amount', 0)) for e in expenses if e.get('paid_privately'))
    total_expenses = private_total
    daily_total = calc['total']
    total_amount = total_expenses + daily_total
    advance = float(data.get('advance_payment', 0))
    payout = total_amount - advance

    fields = {
        'auto_id': auto_id,
        'status': data.get('status', 'draft'),
        'issue_date': data.get('issue_date'),
        'employee_id': data.get('employee_id') or None,
        'destination': data.get('destination'),
        'purpose': data.get('purpose'),
        'client_info': data.get('client_info'),
        'expected_duration': data.get('expected_duration'),
        'departure_date': data.get('departure_date'),
        'vehicle_id': data.get('vehicle_id') or None,
        'start_km': data.get('start_km') or None,
        'end_km': data.get('end_km') or None,
        'trip_start_datetime': data.get('trip_start_datetime'),
        'trip_end_datetime': data.get('trip_end_datetime'),
        'trip_duration_days': calc['days'],
        'trip_duration_hours': calc['hours'],
        'trip_duration_minutes': calc['minutes'],
        'daily_allowance_count': calc['dnevnice'],
        'daily_allowance_rate': rate,
        'daily_allowance_total': daily_total,
        'advance_payment': advance,
        'total_expenses': total_expenses,
        'total_amount': total_amount,
        'payout_amount': payout,
        'report_text': data.get('report_text'),
        'place_of_report': data.get('place_of_report'),
        'approved_by_id': data.get('approved_by_id') or None,
        'validator_id': data.get('validator_id') or None,
        'updated_at': datetime.now().isoformat()
    }

    if order_id:
        sets = ', '.join(f"{k}=?" for k in fields if k != 'auto_id')
        vals = [fields[k] for k in fields if k != 'auto_id'] + [order_id]
        c.execute(f"UPDATE travel_orders SET {sets} WHERE id=?", vals)
        c.execute("DELETE FROM expenses WHERE travel_order_id=?", (order_id,))
    else:
        # New order - generate auto_id using same connection (avoid database locked)
        current_year = datetime.now().year
        used_rows = conn.execute(
            "SELECT auto_id FROM travel_orders WHERE auto_id LIKE ?",
            (f"{current_year}-%",)
        ).fetchall()
        used_nums = set()
        for row in used_rows:
            try:
                if row['auto_id'].endswith('-I'):
                    continue  # soft-deleted, broj slobodan
                used_nums.add(int(row['auto_id'].split('-')[1]))
            except: pass
        candidate = 1
        while candidate in used_nums:
            candidate += 1
        auto_id = f"{current_year}-{candidate}"
        fields['auto_id'] = auto_id
        conn.execute("UPDATE settings SET value=? WHERE key='last_order_number'", (str(candidate),))
        conn.execute("UPDATE settings SET value=? WHERE key='last_order_year'", (str(current_year),))
        fields['created_at'] = datetime.now().isoformat()
        cols = ', '.join(fields.keys())
        placeholders = ', '.join('?' for _ in fields)
        c.execute(f"INSERT INTO travel_orders ({cols}) VALUES ({placeholders})", list(fields.values()))
        order_id = c.lastrowid

    # Save expenses
    for i, exp in enumerate(expenses):
        c.execute('''INSERT INTO expenses (travel_order_id, category_id, description, paid_privately, amount, sort_order)
                     VALUES (?, ?, ?, ?, ?, ?)''',
                  (order_id, exp.get('category_id') or None, exp.get('description'),
                   1 if exp.get('paid_privately') else 0,
                   float(exp.get('amount', 0)) if exp.get('paid_privately') else 0,
                   i))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': order_id, 'auto_id': auto_id})

@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
@require_perm('can_delete_orders')
def delete_order(order_id):
    audit('delete', module='putni_nalozi', entity='travel_order', entity_id=order_id)
    conn = get_db()
    # Dodaj oznaku -I na auto_id da se sačuva slijed brojeva
    order = conn.execute("SELECT auto_id FROM travel_orders WHERE id=?", (order_id,)).fetchone()
    if order:
        current_id = order['auto_id'] or ''
        # Dodaj -I samo ako već nema oznaku
        if not current_id.endswith('-I'):
            new_id = current_id + '-I'
            conn.execute("UPDATE travel_orders SET auto_id=? WHERE id=?", (new_id, order_id))
    conn.execute("UPDATE travel_orders SET is_deleted=1, deleted_at=? WHERE id=?",
                (datetime.now().isoformat(), order_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/reports/popis-naloga')
@require_perm('can_view_reports')
def report_popis_naloga():
    conn = get_db()
    year_filter = request.args.get('year', str(datetime.now().year))
    employee_filter = request.args.get('employee', '')
    
    query = '''
        SELECT to2.*, e.name as employee_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        WHERE (to2.is_deleted = 0 OR to2.is_deleted IS NULL)
    '''
    params = []
    if year_filter:
        query += " AND to2.auto_id LIKE ?"
        params.append(f"{year_filter}-%")
    if employee_filter:
        query += " AND to2.employee_id = ?"
        params.append(employee_filter)
    query += " ORDER BY CAST(substr(to2.auto_id, 1, 4) AS INTEGER) DESC, CAST(substr(to2.auto_id, 6) AS INTEGER) DESC"
    
    orders = conn.execute(query, params).fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    
    # Get available years
    years = conn.execute(
        "SELECT DISTINCT substr(auto_id, 1, 4) as yr FROM travel_orders WHERE is_deleted=0 OR is_deleted IS NULL ORDER BY yr DESC"
    ).fetchall()
    
    # Totals
    total_amount = sum(o['total_amount'] or 0 for o in orders)
    total_paid = sum(o['total_amount'] or 0 for o in orders if o['is_paid'])
    total_unpaid = total_amount - total_paid
    
    conn.close()
    return render_template('report_popis.html',
        orders=rows_to_dicts(orders),
        employees=rows_to_dicts(employees),
        years=[r['yr'] for r in years],
        year_filter=year_filter,
        employee_filter=employee_filter,
        total_amount=total_amount,
        total_paid=total_paid,
        total_unpaid=total_unpaid)


@app.route('/reports/popis-naloga/export')
@require_perm('can_view_reports')
def report_popis_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    conn = get_db()
    year_filter = request.args.get('year', str(datetime.now().year))
    employee_filter = request.args.get('employee', '')

    query = '''
        SELECT to2.*, e.name as employee_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        WHERE (to2.is_deleted = 0 OR to2.is_deleted IS NULL)
    '''
    params = []
    if year_filter:
        query += " AND to2.auto_id LIKE ?"
        params.append(f"{year_filter}-%")
    if employee_filter:
        query += " AND to2.employee_id = ?"
        params.append(employee_filter)
    query += " ORDER BY CAST(substr(to2.auto_id, 1, 4) AS INTEGER) DESC, CAST(substr(to2.auto_id, 6) AS INTEGER) DESC"

    orders = conn.execute(query, params).fetchall()
    settings = {row['key']: row['value'] for row in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Popis PN {year_filter}"

    # Colors
    NAVY = "1A3A5C"
    LIGHT_BLUE = "E8F0F7"
    WHITE = "FFFFFF"
    GREEN = "27AE60"
    ORANGE = "E67E22"
    GRAY = "666666"
    BORDER_COLOR = "AAC4DB"

    thin = Side(style='thin', color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, text, bold=True, bg=NAVY, fg=WHITE, align='center'):
        cell.value = text
        cell.font = Font(name='Arial', bold=bold, color=fg, size=10)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = border

    # ── Title block ──
    company = settings.get('company_name', '')
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"POPIS PUTNIH NALOGA — {year_filter}"
    title_cell.font = Font(name='Arial', bold=True, size=14, color=NAVY)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill('solid', start_color=LIGHT_BLUE)
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    sub_cell.value = f"{company} | Generirano: {datetime.now().strftime('%d.%m.%Y.')}"
    sub_cell.font = Font(name='Arial', size=9, color=GRAY)
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.append([])  # empty row 3

    # ── Header row ──
    headers = ['Br. PN', 'Datum', 'Destinacija', 'Klijent / Partner', 'Djelatnik', 'Iznos (€)', 'Isplaćeno', 'Datum isplate']
    for col, h in enumerate(headers, 1):
        hdr_cell(ws.cell(row=4, column=col), h)
    ws.row_dimensions[4].height = 22

    # ── Data rows ──
    data_start = 5
    for i, o in enumerate(orders):
        r = data_start + i
        bg = WHITE if i % 2 == 0 else LIGHT_BLUE
        thin_data = Side(style='thin', color=BORDER_COLOR)
        brd = Border(left=thin_data, right=thin_data, top=thin_data, bottom=thin_data)

        def dc(col, val, bold=False, align='left', num_fmt=None, fg='000000'):
            c = ws.cell(row=r, column=col)
            c.value = val
            c.font = Font(name='Arial', bold=bold, color=fg, size=9)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal=align, vertical='center')
            c.border = brd
            if num_fmt: c.number_format = num_fmt

        dc(1, f"PN {o['auto_id']}", bold=True, fg=NAVY)
        dc(2, o['issue_date'] or '', align='center')
        dc(3, o['destination'] or '')
        dc(4, o['client_info'] or '')
        dc(5, o['employee_name'] or '')
        dc(6, float(o['total_amount'] or 0), align='right', num_fmt='#,##0.00 €', bold=True)
        paid_val = 'DA' if o['is_paid'] else 'NE'
        paid_color = GREEN if o['is_paid'] else ORANGE
        dc(7, paid_val, align='center', fg=paid_color, bold=True)
        dc(8, o['paid_at'] or '', align='center')
        ws.row_dimensions[r].height = 18

    # ── Totals row ──
    total_row = data_start + len(orders)
    ws.merge_cells(f'A{total_row}:E{total_row}')
    tot_label = ws.cell(row=total_row, column=1)
    tot_label.value = f"UKUPNO ({len(orders)} naloga)"
    tot_label.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    tot_label.fill = PatternFill('solid', start_color=NAVY)
    tot_label.alignment = Alignment(horizontal='right', vertical='center')
    tot_label.border = border

    tot_val = ws.cell(row=total_row, column=6)
    tot_val.value = f"=SUM(F{data_start}:F{total_row-1})"
    tot_val.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    tot_val.fill = PatternFill('solid', start_color=NAVY)
    tot_val.alignment = Alignment(horizontal='right', vertical='center')
    tot_val.number_format = '#,##0.00 €'
    tot_val.border = border

    for col in [7, 8]:
        c = ws.cell(row=total_row, column=col)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.border = border

    ws.row_dimensions[total_row].height = 22

    # ── Column widths ──
    col_widths = [14, 13, 18, 28, 22, 14, 12, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ──
    ws.freeze_panes = 'A5'

    # ── Auto filter ──
    ws.auto_filter.ref = f"A4:H{total_row-1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Popis_PN_{year_filter}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)


@app.route('/api/destinations/<int:dest_id>/place-of-report', methods=['POST'])
@admin_required
def set_place_of_report(dest_id):
    data = request.json
    conn = get_db()
    # Only one destination can be place of report at a time
    if data.get('is_place_of_report'):
        conn.execute("UPDATE destinations SET is_place_of_report=0")
    conn.execute("UPDATE destinations SET is_place_of_report=? WHERE id=?",
                (1 if data.get('is_place_of_report') else 0, dest_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/next-auto-id')
@require_perm('can_edit_orders')
def next_auto_id():
    auto_id, year, num = get_next_auto_id()
    return jsonify({'auto_id': auto_id})


@app.route('/api/orders/bulk-status', methods=['POST'])
@require_perm('can_approve_orders')
def bulk_status():
    data = request.json
    ids = data.get('ids', [])
    new_status = data.get('status')
    audit('status_change', module='putni_nalozi', detail=f'Status → {new_status} za nalog(e): {ids}')
    if not ids or not new_status:
        return jsonify({'error': 'Missing ids or status'}), 400
    allowed = ['knjizeno', 'approved', 'rejected']
    if new_status not in allowed:
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db()
    updated = 0
    for order_id in ids:
        order = conn.execute("SELECT status FROM travel_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            continue
        # Only allow knjizeno from approved
        if new_status == 'knjizeno' and order['status'] == 'approved':
            conn.execute("UPDATE travel_orders SET status='knjizeno', updated_at=? WHERE id=?",
                        (datetime.now().isoformat(), order_id))
            updated += 1
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/orders/<int:order_id>/duplicate', methods=['POST'])
@require_perm('can_edit_orders')
def duplicate_order(order_id):
    audit('duplicate', module='putni_nalozi', entity='travel_order', entity_id=order_id)
    conn = get_db()
    order = conn.execute("SELECT * FROM travel_orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    expenses = conn.execute("SELECT * FROM expenses WHERE travel_order_id=? ORDER BY sort_order", (order_id,)).fetchall()

    # Generate new auto_id
    current_year = __import__('datetime').datetime.now().year
    used_rows = conn.execute("SELECT auto_id FROM travel_orders WHERE auto_id LIKE ?", (f"{current_year}-%",)).fetchall()
    used = set()
    for r in used_rows:
        try:
            if r['auto_id'].endswith('-I'):
                continue  # soft-deleted, broj slobodan
            used.add(int(r['auto_id'].split('-')[1]))
        except: pass
    candidate = 1
    while candidate in used: candidate += 1
    new_auto_id = f"{current_year}-{candidate}"

    now = datetime.now().isoformat()
    today = datetime.now().strftime('%Y-%m-%d')

    c = conn.cursor()
    c.execute('''INSERT INTO travel_orders
        (auto_id, status, issue_date, employee_id, destination, purpose, client_info,
         expected_duration, departure_date, vehicle_id, start_km, end_km,
         trip_start_datetime, trip_end_datetime, trip_duration_days, trip_duration_hours,
         trip_duration_minutes, daily_allowance_count, daily_allowance_rate, daily_allowance_total,
         advance_payment, total_expenses, total_amount, payout_amount,
         report_text, place_of_report, approved_by_id, validator_id, created_at, updated_at)
        VALUES (?,'draft',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (new_auto_id, today,
         order['employee_id'], order['destination'], order['purpose'], order['client_info'],
         order['expected_duration'], order['departure_date'], order['vehicle_id'],
         order['start_km'], order['end_km'], order['trip_start_datetime'], order['trip_end_datetime'],
         order['trip_duration_days'], order['trip_duration_hours'], order['trip_duration_minutes'],
         order['daily_allowance_count'], order['daily_allowance_rate'], order['daily_allowance_total'],
         order['advance_payment'], order['total_expenses'], order['total_amount'], order['payout_amount'],
         order['report_text'], order['place_of_report'], order['approved_by_id'], order['validator_id'],
         now, now))
    new_id = c.lastrowid

    # Copy expenses
    for exp in expenses:
        c.execute('''INSERT INTO expenses (travel_order_id, category_id, description, paid_privately, amount, sort_order)
                     VALUES (?,?,?,?,?,?)''',
                  (new_id, exp['category_id'], exp['description'], exp['paid_privately'], exp['amount'], exp['sort_order']))

    conn.execute("UPDATE settings SET value=? WHERE key='last_order_number'", (str(candidate),))
    conn.execute("UPDATE settings SET value=? WHERE key='last_order_year'", (str(current_year),))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': new_id, 'auto_id': new_auto_id})


@app.route('/api/orders/<int:order_id>/destroy', methods=['DELETE'])
@require_perm('can_delete_orders')
def destroy_order(order_id):
    audit('destroy', module='putni_nalozi', entity='travel_order', entity_id=order_id, detail='Trajno brisanje')
    conn = get_db()
    conn.execute("DELETE FROM travel_orders WHERE id=?", (order_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/orders/<int:order_id>/payment', methods=['POST'])
@require_perm('can_view_orders')
def mark_payment(order_id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE travel_orders SET is_paid=?, paid_at=? WHERE id=?",
                (1 if data.get('is_paid') else 0, data.get('paid_at'), order_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/calculate_dnevnice', methods=['POST'])
@require_perm('can_edit_orders')
def api_calculate_dnevnice():
    data = request.json
    conn = get_db()
    rate_row = conn.execute("SELECT value FROM settings WHERE key='daily_allowance_rate'").fetchone()
    rate = float(rate_row['value']) if rate_row else 30.0
    conn.close()
    result = calculate_dnevnice(data.get('start', ''), data.get('end', ''), rate)
    return jsonify(result)

def _generate_and_save_pdf(order_id):
    conn = get_db()
    order = conn.execute("SELECT * FROM travel_orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Not found'}), 404
    expenses = conn.execute('''SELECT e.*, ec.name as cat_name FROM expenses e
                               LEFT JOIN expense_categories ec ON e.category_id = ec.id
                               WHERE e.travel_order_id=? ORDER BY e.sort_order''', (order_id,)).fetchall()
    employee = conn.execute("SELECT * FROM employees WHERE id=?", (order['employee_id'],)).fetchone() if order['employee_id'] else None
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (order['vehicle_id'],)).fetchone() if order['vehicle_id'] else None
    approved_by = conn.execute("SELECT * FROM employees WHERE id=?", (order['approved_by_id'],)).fetchone() if order['approved_by_id'] else None
    validator = conn.execute("SELECT * FROM employees WHERE id=?", (order['validator_id'],)).fetchone() if order['validator_id'] else None
    blagajnik = conn.execute("SELECT * FROM employees WHERE is_blagajnik=1 LIMIT 1").fetchone()
    knjizio = conn.execute("SELECT * FROM employees WHERE is_knjizio=1 LIMIT 1").fetchone()
    settings = {row['key']: row['value'] for row in conn.execute("SELECT * FROM settings").fetchall()}
    pdf_buffer = create_pdf(dict(order), list(expenses),
                           row_to_dict(employee), row_to_dict(vehicle),
                           row_to_dict(approved_by), row_to_dict(validator),
                           row_to_dict(blagajnik), row_to_dict(knjizio), settings)
    pdf_filename = f"PN_{order['auto_id']}.pdf"
    pdf_folder = os.path.join(os.path.dirname(__file__), 'pdfs')
    os.makedirs(pdf_folder, exist_ok=True)
    pdf_path_full = os.path.join(pdf_folder, pdf_filename)
    with open(pdf_path_full, 'wb') as fh:
        fh.write(pdf_buffer.read())
    conn.execute("UPDATE travel_orders SET pdf_path=? WHERE id=?", (pdf_filename, order_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': order_id, 'auto_id': order['auto_id'],
                   'status': 'approved', 'pdf': pdf_filename})


@app.route('/orders/<int:order_id>/pdf')
@require_perm('can_view_orders')
def generate_pdf(order_id):
    audit('export_pdf', module='putni_nalozi', entity='travel_order', entity_id=order_id)
    conn = get_db()
    order = conn.execute("SELECT * FROM travel_orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return "Not found", 404
    if order['status'] not in ['approved', 'knjizeno']:
        conn.close()
        return "PDF nije dostupan. Nalog mora biti odobren.", 404
    # Try saved file first
    if order['pdf_path']:
        pdf_path_full = os.path.join(os.path.dirname(__file__), 'pdfs', order['pdf_path'])
        if os.path.exists(pdf_path_full):
            conn.close()
            return send_file(pdf_path_full, mimetype='application/pdf',
                           download_name=order['pdf_path'], as_attachment=False)
    # Fallback: generate on the fly
    expenses = conn.execute('''SELECT e.*, ec.name as cat_name FROM expenses e
                               LEFT JOIN expense_categories ec ON e.category_id = ec.id
                               WHERE e.travel_order_id=? ORDER BY e.sort_order''', (order_id,)).fetchall()
    employee = conn.execute("SELECT * FROM employees WHERE id=?", (order['employee_id'],)).fetchone() if order['employee_id'] else None
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (order['vehicle_id'],)).fetchone() if order['vehicle_id'] else None
    approved_by = conn.execute("SELECT * FROM employees WHERE id=?", (order['approved_by_id'],)).fetchone() if order['approved_by_id'] else None
    validator = conn.execute("SELECT * FROM employees WHERE id=?", (order['validator_id'],)).fetchone() if order['validator_id'] else None
    blagajnik = conn.execute("SELECT * FROM employees WHERE is_blagajnik=1 LIMIT 1").fetchone()
    knjizio = conn.execute("SELECT * FROM employees WHERE is_knjizio=1 LIMIT 1").fetchone()
    settings = {row['key']: row['value'] for row in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    pdf_buffer = create_pdf(dict(order), list(expenses),
                           row_to_dict(employee), row_to_dict(vehicle),
                           row_to_dict(approved_by), row_to_dict(validator),
                           row_to_dict(blagajnik), row_to_dict(knjizio), settings)
    filename = f"PN_{order['auto_id']}.pdf"
    return send_file(pdf_buffer, mimetype='application/pdf', download_name=filename, as_attachment=False)

def create_pdf(order, expenses, employee, vehicle, approved_by, validator, blagajnik, knjizio, settings):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             topMargin=15*mm, bottomMargin=15*mm,
                             leftMargin=20*mm, rightMargin=20*mm,
                             title=f"PN {order['auto_id']}",
                             author=settings.get("company_name", ""))

    styles = getSampleStyleSheet()
    W = A4[0] - 40*mm  # content width

    def style(name='Normal', size=9, bold=False, align=TA_LEFT, color=colors.black, leading=None):
        return ParagraphStyle(name, parent=styles['Normal'],
                               fontSize=size, fontName=BOLD_FONT if bold else NORMAL_FONT,
                               alignment=align, textColor=color,
                               leading=leading or (size * 1.3))

    def safe(text):
        """Return text as-is - DejaVu font supports Croatian characters"""
        if not text:
            return ''
        return str(text)

    def fmt_date(val):
        """Format YYYY-MM-DD to DD.MM.YYYY."""
        if not val:
            return ''
        try:
            s = str(val)[:10]
            if len(s) == 10 and s[4] == '-':
                y, m, d = s.split('-')
                return f"{d}.{m}.{y}."
        except:
            pass
        return str(val)

    def fmt_num(val):
        """Display number without decimal if whole, with decimal if not"""
        try:
            f = float(val)
            return str(int(f)) if f == int(f) else str(f)
        except:
            return str(val)

    BLUE = colors.HexColor('#1a3a5c')
    LIGHT_BLUE = colors.HexColor('#e8f0f7')
    BORDER = colors.HexColor('#aac4db')
    GRAY = colors.HexColor('#666666')

    story = []

    # ── REGISTER FONTS ────────────────────────────────────────────────────────
    try:
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        import os
        # Try Arial Unicode (available on macOS)
        arial_path = '/Library/Fonts/Arial Unicode.ttf'
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont('ArialUnicode', arial_path))
            NORMAL_FONT = 'ArialUnicode'
            BOLD_FONT = 'ArialUnicode'  # Arial Unicode has no separate bold
        else:
            NORMAL_FONT = 'Helvetica'
            BOLD_FONT = 'Helvetica-Bold'
    except:
        NORMAL_FONT = 'Helvetica'
        BOLD_FONT = 'Helvetica-Bold'

    # ── HEADER ────────────────────────────────────────────────────────────────
    conn_logo1 = get_db()
    _logo_row1 = conn_logo1.execute("SELECT value FROM settings WHERE key='company_logo'").fetchone()
    conn_logo1.close()
    _logo_file1 = _logo_row1['value'] if _logo_row1 and _logo_row1['value'] else 'logo.png'
    logo_path = os.path.join(os.path.dirname(__file__), 'static', _logo_file1)
    if not os.path.exists(logo_path):
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')
    header_data = [[]]
    if os.path.exists(logo_path):
        img = Image(logo_path, width=41.2*mm, height=12.7*mm)
        img.hAlign = 'LEFT'
        doc_block = [
            Paragraph(f"<b>PUTNI NALOG</b>", style('pt', 14, True, TA_RIGHT, BLUE)),
            Paragraph(f"PN {order['auto_id']}", style('pn', 11, False, TA_RIGHT, BLUE)),
            Paragraph(f"Datum izdavanja: {fmt_date(order['issue_date']) if order.get('issue_date') else ''}", style('pd', 8, False, TA_RIGHT, GRAY)),
        ]
        header_table = Table([[img, doc_block]], colWidths=[90*mm, W-90*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ]))
        story.append(header_table)
    else:
        story.append(Paragraph("<b>PUTNI NALOG</b>", style('t', 16, True, TA_CENTER, BLUE)))
        story.append(Paragraph(f"PN {order['auto_id']} | {order['issue_date']}", style('s', 9, False, TA_CENTER, GRAY)))

    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=4*mm))

    # ── SECTION HELPER ────────────────────────────────────────────────────────
    def section_title(title):
        return Table([[Paragraph(title, style('sh', 9, True, color=colors.white))]],
                     colWidths=[W],
                     style=TableStyle([
                         ('BACKGROUND', (0,0), (-1,-1), BLUE),
                         ('TOPPADDING', (0,0), (-1,-1), 4),
                         ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                         ('LEFTPADDING', (0,0), (-1,-1), 6),
                     ]))

    LABEL_W = 45*mm  # consistent label column width across all rows

    def info_row(label, value, col_w=None):
        cw = col_w or [LABEL_W, W-LABEL_W]
        t = Table([[Paragraph(label, style('lbl', 8, False, color=GRAY)),
                    Paragraph(str(value) if value else '—', style('val', 8, True))]],
                  colWidths=cw)
        t.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 0.3, BORDER),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
        ]))
        return t

    def two_col(pairs):
        half = W / 2
        val_w = half - LABEL_W  # value column = half width minus label
        rows = []
        for i in range(0, len(pairs), 2):
            row = [Paragraph(pairs[i][0], style('lbl2', 8, False, color=GRAY)),
                   Paragraph(str(pairs[i][1]) if pairs[i][1] else '—', style('val2', 8, True))]
            if i+1 < len(pairs):
                row += [Paragraph(pairs[i+1][0], style('lbl3', 8, False, color=GRAY)),
                        Paragraph(str(pairs[i+1][1]) if pairs[i+1][1] else '—', style('val3', 8, True))]
            else:
                row += [Paragraph('', style('e1', 8)), Paragraph('', style('e2', 8))]
            rows.append(row)
        t = Table(rows, colWidths=[LABEL_W, val_w, LABEL_W, val_w])
        t.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 0.3, BORDER),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
        ]))
        return t

    # ── PART 1: PUTNI NALOG ───────────────────────────────────────────────────
    story.append(section_title("NALOG ZA SLUŽBENO PUTOVANJE"))
    story.append(Spacer(1, 2*mm))

    emp_name = safe(employee['name']) if employee else ''
    emp_pos = safe(employee['position']) if employee else ''
    dep_date = fmt_date(order.get('departure_date') or order.get('issue_date') or '')

    story.append(info_row("Zaposlenik:", emp_name))
    story.append(info_row("Radno mjesto:", emp_pos))
    story.append(two_col([
        ("Destinacija:", order['destination'] or ''),
        ("Datum polaska:", dep_date),
    ]))
    story.append(info_row("Svrha putovanja:", safe(order['purpose'] or '')))
    story.append(info_row("Klijent / Partner:", safe(order['client_info'] or '')))
    story.append(two_col([
        ("Predviđeno trajanje:", f"{order['expected_duration']} dan(a)" if order['expected_duration'] else ''),
        ("Akontacija:", f"{order['advance_payment']:.2f} €" if order['advance_payment'] else '0,00 €'),
    ]))

    story.append(Spacer(1, 3*mm))
    story.append(section_title("PRIJEVOZ"))
    story.append(Spacer(1, 2*mm))

    veh_name = safe(f"{vehicle['name']} {vehicle['reg_plate']}") if vehicle else ''
    story.append(info_row("Vozilo:", veh_name))
    story.append(two_col([
        ("Početna kilometraža:", f"{order['start_km']} km" if order['start_km'] else ''),
        ("Završna kilometraža:", f"{order['end_km']} km" if order['end_km'] else ''),
    ]))

    story.append(Spacer(1, 1*mm))

    # Director signature block for part 1
    dir_name = safe(approved_by['name']) if approved_by else ''
    dir_sig_path = os.path.join(UPLOAD_FOLDER, approved_by['signature_path']) if approved_by and approved_by['signature_path'] else None

    # Signature block: all aligned to right side of page
    sig_col_w = 65*mm
    empty_col_w = W - sig_col_w

    # Row 1: empty left | signature image right (centered within its column)
    sig_img_cell = ''
    if dir_sig_path and os.path.exists(dir_sig_path):
        try:
            sig_img_cell = Image(dir_sig_path, width=45*mm, height=18*mm)
        except: pass

    # Row 0: empty | sig image (above line)
    # Row 1: empty | name (below line)
    # Row 2: empty | label
    sig_table = Table([
        [Paragraph('', style('e_sl', 8)), sig_img_cell or Paragraph('', style('e_sr', 8))],
        [Paragraph('', style('e_nl', 8)), Paragraph(dir_name, style('ds2', 8, True, TA_CENTER))],
        [Paragraph('', style('e_ll', 8)), Paragraph('(potpis direktora)', style('dsl2', 7, False, TA_CENTER, GRAY))],
    ], colWidths=[empty_col_w, sig_col_w])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ('VALIGN', (1,0), (1,0), 'BOTTOM'),
        # Line between sig and name
        ('LINEBELOW', (1,0), (1,0), 0.8, BLUE),
        ('TOPPADDING', (0,0), (-1,0), 0),
        ('BOTTOMPADDING', (0,0), (-1,0), 0),
        ('TOPPADDING', (0,1), (-1,1), 2),
        ('BOTTOMPADDING', (0,1), (-1,1), 1),
        ('TOPPADDING', (0,2), (-1,2), 1),
        ('BOTTOMPADDING', (0,2), (-1,2), 2),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(sig_table)
    story.append(Spacer(1, 5*mm))

    # ── PART 2: OBRAČUN ───────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1.5, color=BLUE, spaceAfter=2*mm))
    story.append(section_title("OBRAČUN PUTNOG NALOGA"))
    story.append(Spacer(1, 2*mm))

    # Trip duration
    dur_str = f"{order['trip_duration_days']}d {order['trip_duration_hours']}h {order['trip_duration_minutes']}min"
    def fmt_dt(dt_str):
        if not dt_str:
            return ''
        s = str(dt_str)
        # Convert YYYY-MM-DD or YYYY-MM-DDTHH:MM to DD.MM.YYYY. [HH:MM]
        try:
            date_part = s[:10]
            time_part = s[11:16] if len(s) > 10 else ''
            if len(date_part) == 10 and date_part[4] == '-':
                y, m, d = date_part.split('-')
                result = f"{d}.{m}.{y}."
                if time_part:
                    result += f" {time_part}"
                return result
        except:
            pass
        return s.replace('T', ' ')

    story.append(two_col([
        ("Polazak:", fmt_dt(order['trip_start_datetime'])),
        ("Povratak:", fmt_dt(order['trip_end_datetime'])),
    ]))
    story.append(two_col([
        ("Ukupno na putu:", dur_str),
        ("Broj dnevnica:", str(order['daily_allowance_count'])),
    ]))

    story.append(Spacer(1, 3*mm))

    # Dnevnice table
    story.append(section_title("DNEVNICE"))
    story.append(Spacer(1, 1*mm))
    dn_data = [
        [Paragraph('Vrsta troškova', style('th', 8, True, color=colors.white)),
         Paragraph('Ukupno dnevnica', style('th2', 8, True, TA_CENTER, colors.white)),
         Paragraph('1 dnevnica (€)', style('th3', 8, True, TA_CENTER, colors.white)),
         Paragraph('Ukupno (€)', style('th4', 8, True, TA_RIGHT, colors.white))],
        [Paragraph('Dnevnica', style('dr', 8)),
         Paragraph(fmt_num(order['daily_allowance_count']), style('dc', 8, False, TA_CENTER)),
         Paragraph(f"{order['daily_allowance_rate']:.2f}", style('drc', 8, False, TA_CENTER)),
         Paragraph(f"{order['daily_allowance_total']:.2f}", style('drt', 8, False, TA_RIGHT))],
        [Paragraph('Ukupno:', style('duk', 8, True)),
         Paragraph('', style('e3', 8)),
         Paragraph('', style('e4', 8)),
         Paragraph(f"{order['daily_allowance_total']:.2f}", style('dukt', 8, True, TA_RIGHT))],
    ]
    dn_table = Table(dn_data, colWidths=[W*0.4, W*0.2, W*0.2, W*0.2])
    dn_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), BLUE),
        ('BACKGROUND', (0,2), (-1,2), LIGHT_BLUE),
        ('LINEBELOW', (0,1), (-1,1), 0.3, BORDER),
        ('LINEABOVE', (0,2), (-1,2), 0.5, BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('BOX', (0,0), (-1,-1), 0.5, BORDER),
    ]))
    story.append(dn_table)
    story.append(Spacer(1, 3*mm))

    # Expenses table
    story.append(section_title("OSTALI TROŠKOVI"))
    story.append(Spacer(1, 1*mm))

    exp_header = [
        Paragraph('Kategorija', style('eh', 8, True, color=colors.white)),
        Paragraph('Opis', style('eh2', 8, True, color=colors.white)),
        Paragraph('Privatno', style('eh3', 8, True, TA_CENTER, colors.white)),
        Paragraph('Iznos (€)', style('eh4', 8, True, TA_RIGHT, colors.white)),
    ]
    exp_rows = [exp_header]
    for exp in expenses:
        amt = f"{exp['amount']:.2f}" if exp['paid_privately'] and exp['amount'] else ''
        exp_rows.append([
            Paragraph(safe(exp['cat_name'] or ''), style('er1', 8)),
            Paragraph(safe(exp['description'] or ''), style('er2', 8)),
            Paragraph('✓' if exp['paid_privately'] else '', style('er3', 8, False, TA_CENTER)),
            Paragraph(amt, style('er4', 8, False, TA_RIGHT)),
        ])

    # Total row
    exp_rows.append([
        Paragraph('Ukupno:', style('etuk', 8, True)),
        Paragraph('', style('e5', 8)),
        Paragraph('', style('e6', 8)),
        Paragraph(f"{order['total_expenses']:.2f}", style('etukt', 8, True, TA_RIGHT)),
    ])

    exp_table = Table(exp_rows, colWidths=[W*0.22, W*0.50, W*0.12, W*0.16])
    exp_styles = [
        ('BACKGROUND', (0,0), (-1,0), BLUE),
        ('BACKGROUND', (0,-1), (-1,-1), LIGHT_BLUE),
        ('LINEABOVE', (0,-1), (-1,-1), 0.5, BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('BOX', (0,0), (-1,-1), 0.5, BORDER),
    ]
    for i in range(1, len(exp_rows)-1):
        if i % 2 == 0:
            exp_styles.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#f5f8fb')))
        exp_styles.append(('LINEBELOW', (0,i), (-1,i), 0.2, BORDER))
    exp_table.setStyle(TableStyle(exp_styles))
    story.append(exp_table)
    story.append(Spacer(1, 3*mm))

    # Summary totals
    story.append(section_title("REKAPITULACIJA"))
    story.append(Spacer(1, 1*mm))
    total_data = [
        [Paragraph('Sveukupni troškovi (dnevnice + troškovi):', style('tt', 9, True)), Paragraph(f"{order['total_amount']:.2f} €", style('ttv', 9, True, TA_RIGHT))],
        [Paragraph('Akontacija (ako je isplaćena):', style('at', 9)), Paragraph(f"{order['advance_payment']:.2f} €", style('atv', 9, False, TA_RIGHT))],
        [Paragraph('ZA ISPLATU:', style('pt2', 10, True, color=BLUE)), Paragraph(f"{order['payout_amount']:.2f} €", style('ptv', 10, True, TA_RIGHT, BLUE))],
    ]
    total_table = Table(total_data, colWidths=[W*0.75, W*0.25])
    total_table.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,1), 0.3, BORDER),
        ('LINEABOVE', (0,2), (-1,2), 1.5, BLUE),
        ('BACKGROUND', (0,2), (-1,2), LIGHT_BLUE),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('BOX', (0,0), (-1,-1), 0.5, BORDER),
    ]))
    story.append(total_table)
    story.append(Spacer(1, 4*mm))

    # Report - always on new page
    if order['report_text']:
        story.append(PageBreak())
        story.append(section_title("IZVJEŠĆE S PUTA"))
        story.append(Spacer(1, 1*mm))
        story.append(Paragraph(safe(order['report_text']).replace('\n', '<br/>'), style('rt', 8)))
        story.append(Spacer(1, 2*mm))

    # ── SIGNATURES ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 2*mm))
    place_date = f"{order['place_of_report'] or ''}, {fmt_date(order.get('issue_date', ''))}"
    story.append(Paragraph(f"Mjesto i datum: {place_date}", style('pd2', 8, False, color=GRAY)))
    story.append(Spacer(1, 3*mm))

    def sig_cell(title, person):
        """Returns (above_line_items, below_line_items) for a signature column"""
        above = [Paragraph(title, style(f's_{title}', 7, False, TA_CENTER, GRAY))]
        # Scanned signature or empty space
        if person and person['signature_path']:
            sig_path = os.path.join(UPLOAD_FOLDER, person['signature_path'])
            if os.path.exists(sig_path):
                try:
                    img = Image(sig_path, width=45*mm, height=18*mm)
                    img.hAlign = 'CENTER'
                    above.append(img)
                except:
                    above.append(Spacer(1, 18*mm))
            else:
                above.append(Spacer(1, 18*mm))
        else:
            above.append(Spacer(1, 18*mm))
        below = [Paragraph(safe(person['name']) if person else '_' * 20, style(f'sn_{title}', 8, True, TA_CENTER))]
        return above, below

    col_w = W / 3
    emp_above, emp_below = sig_cell("Podnositelj računa:", employee)
    val_above, val_below = sig_cell("Validator:", validator)
    dir_above, dir_below = sig_cell("Direktor:", approved_by)

    # Single table: row 0 = title+sig (above line), row 1 = name (below line)
    sig_table = Table(
        [[emp_above, val_above, dir_above],
         [emp_below, val_below, dir_below]],
        colWidths=[col_w, col_w, col_w]
    )
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,0), 'BOTTOM'),
        ('VALIGN', (0,1), (-1,1), 'TOP'),
        # Line between row 0 and row 1
        ('LINEBELOW', (0,0), (-1,0), 0.8, BLUE),
        # Tight spacing
        ('TOPPADDING', (0,0), (-1,0), 2),
        ('BOTTOMPADDING', (0,0), (-1,0), 0),
        ('TOPPADDING', (0,1), (-1,1), 2),
        ('BOTTOMPADDING', (0,1), (-1,1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(sig_table)

    story.append(Spacer(1, 5*mm))

    # Blagajnik / Knjižio stacked vertically
    blag_name = safe(blagajnik['name']) if blagajnik else ''
    knj_name = safe(knjizio['name']) if knjizio else ''
    story.append(Spacer(1, 3*mm))
    bottom_data = [
        [Paragraph(f"Blagajnik/likvidator: <b>{blag_name}</b>", style('bl', 8)), Paragraph('', style('e_bl', 8))],
        [Paragraph(f"Knjižio: {'_'*20 if not knj_name else knj_name}", style('kn', 8)), Paragraph('', style('e_kn', 8))],
    ]
    bottom_table = Table(bottom_data, colWidths=[W*0.6, W*0.4])
    bottom_table.setStyle(TableStyle([
        ('LINEABOVE', (0,0), (-1,0), 0.5, BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(bottom_table)

    # Footer note
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "Uz putni nalog obavezno je priložiti: račun za gorivo, cestarine, putne karte, smještaj, itd.",
        style('fn', 7, False, color=GRAY)))

    doc.build(story)
    buffer.seek(0)
    return buffer

# ─── SETTINGS API ─────────────────────────────────────────────────────────────

@app.route('/api/settings/logo', methods=['POST'])
@admin_required
def upload_company_logo():
    if 'logo' not in request.files:
        return jsonify({'error': 'Nema datoteke'}), 400
    f = request.files['logo']
    if not f.filename:
        return jsonify({'error': 'Nema naziva datoteke'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ['.png', '.jpg', '.jpeg', '.svg']:
        return jsonify({'error': 'Dozvoljeni formati: PNG, JPG, SVG'}), 400
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'company_logo' + ext)
    # Obriši stari logo ako postoji
    for old_ext in ['.png', '.jpg', '.jpeg', '.svg']:
        old_path = os.path.join(os.path.dirname(__file__), 'static', 'company_logo' + old_ext)
        if os.path.exists(old_path):
            os.remove(old_path)
    f.save(logo_path)
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('company_logo', ?)",
                 ('company_logo' + ext,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'logo': 'company_logo' + ext})

@app.route('/api/settings/logo', methods=['DELETE'])
@admin_required
def delete_company_logo():
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key='company_logo'").fetchone()
    if row and row['value']:
        logo_path = os.path.join(os.path.dirname(__file__), 'static', row['value'])
        if os.path.exists(logo_path):
            os.remove(logo_path)
    conn.execute("DELETE FROM settings WHERE key='company_logo'")
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ─── QUOTES ───────────────────────────────────────────────────────────────────

def get_next_quote_id():
    conn = get_db()
    current_year = datetime.now().year
    rows = conn.execute("SELECT auto_id FROM quotes WHERE auto_id LIKE ?", (f"{current_year}-%",)).fetchall()
    used = set()
    for r in rows:
        try: used.add(int(r['auto_id'].split('-')[1]))
        except: pass
    candidate = 1
    while candidate in used:
        candidate += 1
    conn.close()
    return f"{current_year}-{candidate}", current_year, candidate

@app.route('/quotes')
@require_perm('can_view_quotes')
def quotes_list():
    audit('view', module='ponude', entity='list')
    conn = get_db()
    quotes = conn.execute('''
        SELECT q.*, c.name as client_name
        FROM quotes q LEFT JOIN clients c ON q.client_id = c.id
        WHERE 1=1 ORDER BY q.id DESC
    ''').fetchall()
    conn.close()
    return render_template('quotes_list.html', quotes=rows_to_dicts(quotes))

@app.route('/quotes/new')
@require_perm('can_edit_quotes')
def new_quote():
    conn = get_db()
    auto_id, _, _ = get_next_quote_id()
    clients = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    default_place = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    conn.close()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('quotes_form.html', auto_id=auto_id, today=today,
                          clients=rows_to_dicts(clients),
                          employees=rows_to_dicts(employees),
                          settings=settings, quote=None, items=[],
                          default_place=default_place['name'] if default_place else 'Zagreb')

@app.route('/quotes/<int:quote_id>/edit')
@require_perm('can_view_quotes')
def edit_quote(quote_id):
    conn = get_db()
    quote = conn.execute("SELECT * FROM quotes WHERE id=?", (quote_id,)).fetchone()
    if not quote: return redirect(url_for('quotes_list'))
    items = conn.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY sort_order", (quote_id,)).fetchall()
    clients = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    default_place = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    conn.close()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('quotes_form.html', auto_id=quote['auto_id'], today=today,
                          clients=rows_to_dicts(clients),
                          employees=rows_to_dicts(employees),
                          settings=settings, quote=dict(quote), items=rows_to_dicts(items),
                          default_place=default_place['name'] if default_place else 'Zagreb')

@app.route('/api/quotes', methods=['POST'])
@login_required
def save_quote():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    quote_id = data.get('id')
    new_status = data.get('status', 'draft')

    # Status-only change for locked quotes
    if quote_id:
        existing = conn.execute("SELECT status FROM quotes WHERE id=?", (quote_id,)).fetchone()
        if existing and existing['status'] in ['sent', 'accepted', 'rejected']:
            allowed = {'sent': ['accepted', 'rejected', 'draft'],
                      'accepted': ['draft'], 'rejected': ['draft']}
            if new_status in allowed.get(existing['status'], []):
                conn.execute("UPDATE quotes SET status=?, updated_at=? WHERE id=?",
                           (new_status, datetime.now().isoformat(), quote_id))
                # Kad se ponuda prihvati — automatski označi klijenta
                if new_status == 'accepted':
                    quote_row = conn.execute("SELECT client_id FROM quotes WHERE id=?", (quote_id,)).fetchone()
                    if quote_row and quote_row['client_id']:
                        conn.execute("UPDATE clients SET is_client=1 WHERE id=?", (quote_row['client_id'],))
                conn.commit()
                conn.close()
                return jsonify({'success': True, 'id': quote_id, 'status': new_status})

    items = data.get('items', [])
    total_net = sum(float(i.get('total', 0)) for i in items)
    pdv_rate = float(data.get('pdv_rate', 25))
    total_pdv = round(total_net * pdv_rate / 100, 2)
    total_gross = round(total_net + total_pdv, 2)

    fields = {
        'status': new_status,
        'issue_date': data.get('issue_date'),
        'valid_days': data.get('valid_days', 7),
        'place_of_issue': data.get('place_of_issue', 'Zagreb'),
        'client_id': data.get('client_id') or None,
        'prepared_by_id': data.get('prepared_by_id') or None,
        'notes': data.get('notes', ''),
        'comment': data.get('comment', ''),
        'comment': data.get('comment', ''),
        'pdv_rate': pdv_rate,
        'total_net': total_net,
        'total_pdv': total_pdv,
        'total_gross': total_gross,
        'updated_at': datetime.now().isoformat(),
    }

    if quote_id:
        sets = ', '.join(f"{k}=?" for k in fields)
        c.execute(f"UPDATE quotes SET {sets} WHERE id=?", list(fields.values()) + [quote_id])
        c.execute("DELETE FROM quote_items WHERE quote_id=?", (quote_id,))
    else:
        # Generate auto_id
        current_year = datetime.now().year
        used_rows = conn.execute("SELECT auto_id FROM quotes WHERE auto_id LIKE ?", (f"{current_year}-%",)).fetchall()
        used = set()
        for r in used_rows:
            try: used.add(int(r['auto_id'].split('-')[1]))
            except: pass
        candidate = 1
        while candidate in used: candidate += 1
        auto_id = f"{current_year}-{candidate}"
        fields['auto_id'] = auto_id
        fields['created_at'] = datetime.now().isoformat()
        cols = ', '.join(fields.keys())
        placeholders = ', '.join('?' for _ in fields)
        c.execute(f"INSERT INTO quotes ({cols}) VALUES ({placeholders})", list(fields.values()))
        quote_id = c.lastrowid
        conn.execute("UPDATE settings SET value=? WHERE key='last_quote_number'", (str(candidate),))
        conn.execute("UPDATE settings SET value=? WHERE key='last_quote_year'", (str(current_year),))

    for i, item in enumerate(items):
        qty = float(item.get('quantity', 1))
        price = float(item.get('unit_price', 0))
        c.execute('''INSERT INTO quote_items (quote_id, sort_order, description, unit, quantity, unit_price, total)
                     VALUES (?, ?, ?, ?, ?, ?, ?)''',
                  (quote_id, i, item.get('description', ''), item.get('unit', 'Kom'), qty, price, round(qty * price, 2)))

    conn.commit()
    conn.close()

    # If transitioning to 'sent', generate and save PDF
    if new_status == 'sent':
        return _generate_and_save_quote_pdf(quote_id)

    return jsonify({'success': True, 'id': quote_id})

@app.route('/quotes/<int:quote_id>/copy')
@require_perm('can_edit_quotes')
def copy_quote(quote_id):
    audit('copy', module='ponude', entity='quote', entity_id=quote_id)
    conn = get_db()
    quote = conn.execute("SELECT * FROM quotes WHERE id=?", (quote_id,)).fetchone()
    if not quote:
        conn.close()
        return redirect(url_for('quotes_list'))
    items = conn.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY sort_order", (quote_id,)).fetchall()
    clients = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    default_place = conn.execute("SELECT name FROM destinations WHERE is_place_of_report=1 LIMIT 1").fetchone()
    # Get next quote auto_id
    auto_id, _, _ = get_next_quote_id()
    conn.close()
    today = date.today().strftime('%Y-%m-%d')
    # Pass quote as template but with new auto_id and draft status
    quote_copy = dict(quote)
    quote_copy['id'] = None
    quote_copy['status'] = 'draft'
    quote_copy['pdf_path'] = None
    return render_template('quotes_form.html', auto_id=auto_id, today=today,
                          clients=rows_to_dicts(clients),
                          employees=rows_to_dicts(employees),
                          settings=settings, quote=quote_copy, items=rows_to_dicts(items),
                          default_place=default_place['name'] if default_place else 'Zagreb',
                          is_copy=True)


@app.route('/api/quotes/<int:quote_id>/comment', methods=['POST'])
@require_perm('can_view_quotes')
def save_quote_comment(quote_id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE quotes SET comment=?, updated_at=? WHERE id=?",
                (data.get('comment', ''), datetime.now().isoformat(), quote_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})




@app.route('/api/quotes/<int:quote_id>', methods=['DELETE'])
@require_perm('can_delete_quotes')
def delete_quote(quote_id):
    audit('delete', module='ponude', entity='quote', entity_id=quote_id)
    conn = get_db()
    conn.execute("DELETE FROM quotes WHERE id=?", (quote_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

def _generate_and_save_quote_pdf(quote_id):
    conn = get_db()
    quote = conn.execute("SELECT * FROM quotes WHERE id=?", (quote_id,)).fetchone()
    if not quote:
        return jsonify({'error': 'Not found'}), 404
    items = conn.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY sort_order", (quote_id,)).fetchall()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (quote['client_id'],)).fetchone() if quote['client_id'] else None
    prepared_by = conn.execute("SELECT * FROM employees WHERE id=?", (quote['prepared_by_id'],)).fetchone() if quote['prepared_by_id'] else None
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    buf = create_quote_pdf(dict(quote), rows_to_dicts(items), row_to_dict(client), row_to_dict(prepared_by), settings)
    pdf_filename = f"Ponuda_{quote['auto_id']}.pdf"
    pdf_folder = os.path.join(os.path.dirname(__file__), 'pdfs')
    os.makedirs(pdf_folder, exist_ok=True)
    pdf_path_full = os.path.join(pdf_folder, pdf_filename)
    with open(pdf_path_full, 'wb') as fh:
        fh.write(buf.read())
    conn.execute("UPDATE quotes SET pdf_path=? WHERE id=?", (pdf_filename, quote_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': quote_id, 'status': 'sent', 'pdf': pdf_filename})


@app.route('/quotes/<int:quote_id>/pdf')
@require_perm('can_view_quotes')
def quote_pdf(quote_id):
    conn = get_db()
    quote = conn.execute("SELECT * FROM quotes WHERE id=?", (quote_id,)).fetchone()
    if not quote: return "Not found", 404
    # Serve saved PDF if exists (frozen at time of sending)
    if quote['pdf_path']:
        pdf_path_full = os.path.join(os.path.dirname(__file__), 'pdfs', quote['pdf_path'])
        if os.path.exists(pdf_path_full):
            conn.close()
            return send_file(pdf_path_full, mimetype='application/pdf',
                           download_name=quote['pdf_path'], as_attachment=False)
    # Fallback: generate on the fly (for drafts or missing file)
    items = conn.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY sort_order", (quote_id,)).fetchall()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (quote['client_id'],)).fetchone() if quote['client_id'] else None
    prepared_by = conn.execute("SELECT * FROM employees WHERE id=?", (quote['prepared_by_id'],)).fetchone() if quote['prepared_by_id'] else None
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    buf = create_quote_pdf(dict(quote), rows_to_dicts(items), row_to_dict(client), row_to_dict(prepared_by), settings)
    filename = f"Ponuda_{quote['auto_id']}.pdf"
    return send_file(buf, mimetype='application/pdf', download_name=filename, as_attachment=False)


def create_quote_pdf(quote, items, client, prepared_by, settings):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=14*mm, bottomMargin=32*mm,
                            leftMargin=20*mm, rightMargin=20*mm,
                            title=f"Ponuda br. {quote['auto_id']}")
    W = A4[0] - 40*mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    styles = getSampleStyleSheet()

    try:
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        arial_path = '/Library/Fonts/Arial Unicode.ttf'
        if os.path.exists(arial_path):
            try: pdfmetrics.registerFont(TTFont('ArialQ', arial_path))
            except: pass
            NF, BF = 'ArialQ', 'ArialQ'
        else:
            NF, BF = 'Helvetica', 'Helvetica-Bold'
    except:
        NF, BF = 'Helvetica', 'Helvetica-Bold'

    NAVY  = colors.HexColor('#1a3a5c')
    GRAY  = colors.HexColor('#444444')
    LGRAY = colors.HexColor('#888888')
    BORDER= colors.HexColor('#dddddd')
    LIGHT_BLUE = colors.HexColor('#eef4fb')
    BLACK = colors.black

    def st(name='n', size=9, bold=False, align=TA_LEFT, color=BLACK, leading_mult=1.4):
        return ParagraphStyle(name, parent=styles['Normal'],
                              fontSize=size, fontName=BF if bold else NF,
                              alignment=align, textColor=color,
                              leading=size*leading_mult,
                              leftIndent=0, firstLineIndent=0,
                              spaceBefore=0, spaceAfter=0)

    def safe(t): return str(t).strip() if t else ''

    story = []
    conn_logo2 = get_db()
    _logo_row2 = conn_logo2.execute("SELECT value FROM settings WHERE key='company_logo'").fetchone()
    conn_logo2.close()
    _logo_file2 = _logo_row2['value'] if _logo_row2 and _logo_row2['value'] else 'logo.png'
    logo_path = os.path.join(os.path.dirname(__file__), 'static', _logo_file2)
    if not os.path.exists(logo_path):
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')

    # ── HEADER: logo lijevo, firma desno ──────────────────────────────────
    comp_name = settings.get('company_name', '')
    comp_oib  = settings.get('company_oib', '')
    comp_addr = settings.get('company_address', '')
    comp_phone= settings.get('company_phone', '')
    comp_email= settings.get('company_email', '')

    company_block = [
        Paragraph(f'<b>{comp_name}</b>', st('cb', 11, True, TA_RIGHT, NAVY)),
        Paragraph(f'OIB: {comp_oib}', st('co', 8, align=TA_RIGHT, color=GRAY)),
        Paragraph(safe(comp_addr), st('ca', 8, align=TA_RIGHT, color=GRAY)),
    ]
    if comp_phone:
        company_block.append(Paragraph(f'<font color="#888888">✆</font> {comp_phone}', st('cp', 8, align=TA_RIGHT, color=GRAY)))
    if comp_email:
        company_block.append(Paragraph(f'<font color="#1a6b9a">✉</font> {comp_email}', st('ce', 8, align=TA_RIGHT, color=colors.HexColor('#1a6b9a'))))

    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=49*mm, height=15.2*mm)
        logo_img.hAlign = 'LEFT'
        hdr_tbl = Table([[logo_img, company_block]], colWidths=[W*0.45, W*0.55])
    else:
        hdr_tbl = Table([[Paragraph(comp_name, st('cn',14,True,color=NAVY)), company_block]],
                        colWidths=[W*0.45, W*0.55])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('ALIGN',        (1,0), (1,0),   'RIGHT'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 0),
        ('BOTTOMPADDING',(0,0), (-1,-1), 0),
    ]))
    story.append(hdr_tbl)
    # HR i naslov u tablici da budu poravnati s ostatkom sadržaja
    story.append(Table(
        [[HRFlowable(width='100%', thickness=1, color=NAVY)]],
        colWidths=[W],
        style=TableStyle([
            ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
            ('TOPPADDING',(0,0),(-1,-1),4*mm),('BOTTOMPADDING',(0,0),(-1,-1),6*mm),
        ])))

    # ── NASLOV PONUDE ─────────────────────────────────────────────────────
    story.append(Table(
        [[Paragraph(f'PONUDA BR. {quote["auto_id"]}', st('pt', 22, True, TA_LEFT, NAVY))]],
        colWidths=[W],
        style=TableStyle([
            ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
            ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),6*mm),
        ])))

    # ── PRIMATELJ + DETALJI PONUDE ────────────────────────────────────────
    cl_name = safe(client.get('name','')) if client else ''
    cl_addr = safe(client.get('address','')) if client else ''
    cl_oib  = safe(client.get('oib','')) if client else ''
    pb_name = safe(prepared_by.get('name','')) if prepared_by else ''

    # Primatelj — lijevo
    # Primatelj podaci kao lista
    cl_addr_lines = [l.strip() for l in cl_addr.split(',') if l.strip()] if cl_addr else []

    # Detalji podaci
    issue_date = safe(quote.get('issue_date',''))
    place      = safe(quote.get('place_of_issue','Osijek'))
    det_rows_data = [('Datum izdavanja:', issue_date), ('Mjesto izdavanja:', place),
                     ('Ponudu izradio:', pb_name)]


    # Prim podaci u isti broj redaka kao detalji (+ header red)
    prim_rows = []
    if cl_name:
        prim_rows.append(Paragraph(f'<b>{cl_name}</b>', st('pn2', 10, True, color=NAVY)))
    prim_rows += [Paragraph(l, st(f'pa{i}', 9, color=GRAY)) for i, l in enumerate(cl_addr_lines)]
    if cl_oib:
        prim_rows.append(Paragraph(f'OIB: {cl_oib}', st('po', 9, color=GRAY)))

    # Ukupan broj redaka = max(len(prim_rows), len(det_rows_data)) + 1 (header red)
    n_data = max(len(prim_rows), len(det_rows_data))

    # Gradi unified tablicu: 4 kolone [prim_content | spacer | det_lbl | det_val]
    lbl_w = 28*mm
    val_w = 40*mm
    spc_w = 4*mm
    prim_w = W - spc_w - lbl_w - val_w

    # Red 0: labele PRIMATELJ / DETALJI PONUDE
    unified_rows = [[
        Paragraph('PRIMATELJ', st('plbl', 7, True, color=LGRAY)),
        '', '',
        Paragraph('DETALJI PONUDE', st('dlbl', 7, True, color=LGRAY)),
        '',
    ]]

    # Redovi 1..n_data: primatelj red i | det red i
    for i in range(n_data):
        p_cell = prim_rows[i] if i < len(prim_rows) else ''
        if i < len(det_rows_data):
            lbl, val = det_rows_data[i]
            l_cell = Paragraph(lbl, st(f'dl{i}', 9, True, color=NAVY)) if val else ''
            v_cell = Paragraph(val, st(f'dv{i}', 9, color=GRAY)) if val else ''
        else:
            l_cell, v_cell = '', ''
        unified_rows.append([p_cell, '', '', l_cell, v_cell])

    unified = Table(unified_rows, colWidths=[prim_w, spc_w, 0, lbl_w, val_w])
    u_styles = [
        ('SPAN',          (3,0), (4,0)),   # DETALJI PONUDE header span
        ('SPAN',          (0,0), (2,0)),   # PRIMATELJ header span
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING',    (0,0), (-1,0),  1),
        ('BOTTOMPADDING', (0,0), (-1,0),  2),
        ('LEFTPADDING',   (0,0), (0,-1),  0),
    ]
    unified.setStyle(TableStyle(u_styles))
    story.append(unified)
    story.append(Spacer(1, 6*mm))

    # ── TABLICA STAVKI ────────────────────────────────────────────────────
    CW = [10*mm, W - 10*mm - 18*mm - 18*mm - 28*mm, 18*mm, 18*mm, 28*mm]
    hdr_row = [
        Paragraph('RB',       st('h0', 8, True, TA_CENTER, colors.white)),
        Paragraph('Opis usluge', st('h1', 8, True, TA_LEFT,   colors.white)),
        Paragraph('JM',       st('h2', 8, True, TA_CENTER, colors.white)),
        Paragraph('Kol.',     st('h3', 8, True, TA_CENTER, colors.white)),
        Paragraph('Iznos (EUR)', st('h4', 8, True, TA_RIGHT, colors.white)),
    ]
    rows = [hdr_row]
    for i, it in enumerate(items):
        qty = it['quantity']
        qty_str = str(int(qty)) if qty == int(qty) else str(qty)
        rows.append([
            Paragraph(str(i+1), st(f'r{i}0', 9, align=TA_CENTER)),
            Paragraph(safe(it.get('description','')).replace('\n','<br/>'), st(f'r{i}1', 9)),
            Paragraph(safe(it.get('unit','Kom')), st(f'r{i}2', 9, align=TA_CENTER)),
            Paragraph(qty_str, st(f'r{i}3', 9, align=TA_CENTER)),
            Paragraph(f"{it['total']:,.2f}".replace(',','.'), st(f'r{i}4', 9, align=TA_RIGHT)),
        ])

    tbl = Table(rows, colWidths=CW)
    ts = [
        ('BACKGROUND',   (0,0),  (-1,0),  NAVY),
        ('TOPPADDING',   (0,0),  (-1,-1), 5),
        ('BOTTOMPADDING',(0,0),  (-1,-1), 5),
        ('LEFTPADDING',  (0,0),  (-1,-1), 5),
        ('RIGHTPADDING', (0,0),  (-1,-1), 5),
        ('VALIGN',       (0,0),  (-1,-1), 'MIDDLE'),
        ('LINEBELOW',    (0,0),  (-1,-1), 0.3, BORDER),
        ('BOX',          (0,0),  (-1,-1), 0.5, BORDER),
    ]
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 2*mm))

    # ── TOTALS ────────────────────────────────────────────────────────────
    pdv_rate  = float(quote.get('pdv_rate', 25))
    total_net = float(quote.get('total_net', 0))
    total_pdv = float(quote.get('total_pdv', 0))
    total_gross= float(quote.get('total_gross', 0))
    def fmt(v): return f"{v:,.2f} EUR".replace(',','.')

    tot_rows = [
        [Paragraph('IZNOS:',   st('tl0',9,True,TA_RIGHT,LGRAY)), Paragraph(fmt(total_net),   st('tv0',9,align=TA_RIGHT,color=GRAY))],
        [Paragraph('OSNOVICA:',st('tl1',9,True,TA_RIGHT,LGRAY)), Paragraph(fmt(total_net),   st('tv1',9,align=TA_RIGHT,color=GRAY))],
        [Paragraph(f'PDV ({int(pdv_rate)}%):',st('tl2',9,True,TA_RIGHT,LGRAY)), Paragraph(fmt(total_pdv), st('tv2',9,align=TA_RIGHT,color=GRAY))],
        [Paragraph('UKUPNO',   st('tl3',12,True,TA_RIGHT,NAVY)), Paragraph(fmt(total_gross), st('tv3',12,True,TA_RIGHT,NAVY))],
    ]
    tot_tbl = Table(tot_rows, colWidths=[W*0.6+25*mm, W*0.4-25*mm])
    tot_tbl.setStyle(TableStyle([
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('LINEABOVE',     (0,3), (-1,3),  1, NAVY),
        ('TOPPADDING',    (0,3), (-1,3),  6),
    ]))
    story.append(tot_tbl)
    story.append(Spacer(1, 6*mm))

    # ── NAPOMENE ──────────────────────────────────────────────────────────
    notes = safe(quote.get('notes',''))
    if notes:
        note_lines = [l.strip() for l in notes.split('\n') if l.strip()]
        note_content = [Paragraph('NAPOMENE', st('ntit', 9, True, color=NAVY))]
        note_content.append(Spacer(1, 3))
        for line in note_lines:
            note_content.append(Paragraph(f'• {line}', st(f'nl{line[:6]}', 9, color=GRAY, leading_mult=1.5)))
        note_box = Table([[note_content]], colWidths=[W])
        note_box.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), LIGHT_BLUE),
            ('BOX',           (0,0), (-1,-1), 1, colors.HexColor('#c5d8ee')),
            ('LEFTPADDING',   (0,0), (-1,-1), 10),
            ('RIGHTPADDING',  (0,0), (-1,-1), 10),
            ('TOPPADDING',    (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(note_box)
        story.append(Spacer(1, 4*mm))

    # ── FOOTER (canvas, uvijek na dnu) ────────────────────────────────────
    comp_mbs     = settings.get('company_mbs','')
    comp_iban    = settings.get('company_iban','')
    comp_bank    = settings.get('company_bank','Zagrebačka banka')
    comp_capital = settings.get('company_capital','')
    comp_director= settings.get('company_director','')

    def draw_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        page_w = A4[0]
        lm, rm = 20*mm, 20*mm
        fw = page_w - lm - rm
        line_y = 28*mm

        # Linija
        canvas_obj.setStrokeColor(BORDER)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(lm, line_y, page_w - rm, line_y)

        col_w = fw / 3
        # 3 kolone
        footer_cols = [
            ('REGISTRACIJA', [
                f'{comp_name}',
                f'Trgovački sud u Zagrebu',
                f'MBS: {comp_mbs}',
            ]),
            ('BANKOVNI PODACI', [
                f'{comp_bank}',
                f'IBAN: {comp_iban}',
            ]),
            ('KAPITAL I UPRAVA', [
                f'Temeljni kapital: {comp_capital}',
                f'Član uprave: {comp_director}',
            ]),
        ]
        y_start = line_y - 5*mm
        for ci, (title, lines) in enumerate(footer_cols):
            x = lm + ci * col_w
            canvas_obj.setFont(BF, 7)
            canvas_obj.setFillColor(NAVY)
            canvas_obj.drawString(x, y_start, title)
            canvas_obj.setFont(NF, 6.5)
            canvas_obj.setFillColor(colors.HexColor('#666666'))
            for li, line in enumerate(lines):
                canvas_obj.drawString(x, y_start - (li+1)*4*mm, line)
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    buf.seek(0)
    return buf



@app.route('/reports/nepotpuni-nalozi')
@require_perm('can_view_reports')
def report_nepotpuni():
    conn = get_db()
    orders = conn.execute('''
        SELECT to2.*, e.name as employee_name,
               v.name as vehicle_name,
               emp_appr.name as approved_by_name,
               emp_val.name as validator_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        LEFT JOIN vehicles v ON to2.vehicle_id = v.id
        LEFT JOIN employees emp_appr ON to2.approved_by_id = emp_appr.id
        LEFT JOIN employees emp_val ON to2.validator_id = emp_val.id
        WHERE (to2.is_deleted = 0 OR to2.is_deleted IS NULL)
        ORDER BY to2.id DESC
    ''').fetchall()
    conn.close()

    REQUIRED = {
        'employee_id':          'Zaposlenik',
        'destination':          'Destinacija',
        'purpose':              'Svrha putovanja',
        'client_info':          'Klijent',
        'trip_start_datetime':  'Datum polaska',
        'trip_end_datetime':    'Datum povratka',
        'vehicle_id':           'Vozilo',
        'start_km':             'Početna km',
        'end_km':               'Završna km',
        'report_text':          'Izvješće s puta',
        'place_of_report':      'Mjesto izvještaja',
        'validator_id':         'Validator',
        'approved_by_id':       'Odobrio',
    }

    incomplete = []
    for o in orders:
        missing = [label for field, label in REQUIRED.items() if not o[field]]
        if missing:
            row = dict(o)
            row['missing_fields'] = missing
            incomplete.append(row)

    return render_template('report_nepotpuni.html', orders=incomplete, required=REQUIRED, active='reports-incomplete')


@app.route('/reports/nepotpuni-nalozi/export')
@require_perm('can_view_reports')
def report_nepotpuni_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    orders = conn.execute('''
        SELECT to2.*, e.name as employee_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        WHERE (to2.is_deleted = 0 OR to2.is_deleted IS NULL)
        ORDER BY to2.id DESC
    ''').fetchall()
    conn.close()

    REQUIRED = {
        'employee_id':         'Zaposlenik',
        'destination':         'Destinacija',
        'purpose':             'Svrha putovanja',
        'client_info':         'Klijent',
        'trip_start_datetime': 'Datum polaska',
        'trip_end_datetime':   'Datum povratka',
        'vehicle_id':          'Vozilo',
        'start_km':            'Početna km',
        'end_km':              'Završna km',
        'report_text':         'Izvješće s puta',
        'place_of_report':     'Mjesto izvještaja',
        'validator_id':        'Validator',
        'approved_by_id':      'Odobrio',
    }

    incomplete = []
    for o in orders:
        missing = [label for field, label in REQUIRED.items() if not o[field]]
        if missing:
            row = dict(o)
            row['missing_fields'] = missing
            incomplete.append(row)

    NAVY = "1A3A5C"
    LIGHT = "E8F0F7"
    RED_LIGHT = "FFF0F0"
    BORDER_C = "AAC4DB"
    WHITE = "FFFFFF"
    thin = Side(style='thin', color=BORDER_C)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nepotpuni PN"

    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = "PUTNI NALOZI — NEPOTPUNI PODACI"
    t.font = Font(name='Arial', bold=True, size=13, color=NAVY)
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.fill = PatternFill('solid', start_color=LIGHT)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:D2')
    s = ws['A2']
    s.value = f"Generirano: {datetime.now().strftime('%d.%m.%Y. %H:%M')}  |  Ukupno nepotpunih: {len(incomplete)}"
    s.font = Font(name='Arial', size=9, color="666666")
    s.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16
    ws.append([])

    headers = ['Br. PN', 'Zaposlenik', 'Status', 'Nedostaje']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col)
        c.value = h
        c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 22

    STATUS_HR = {'draft':'Nacrt','submitted':'Predano','approved':'Odobreno',
                 'rejected':'Odbijeno','knjizeno':'Knjiženo'}
    for i, o in enumerate(incomplete):
        r = 5 + i
        bg = WHITE if i % 2 == 0 else LIGHT
        def dc(col, val, wrap=False, fg='000000'):
            c = ws.cell(row=r, column=col)
            c.value = val
            c.font = Font(name='Arial', size=9, color=fg)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(vertical='center', wrap_text=wrap)
            c.border = border
        dc(1, f"PN {o['auto_id']}", fg=NAVY)
        dc(2, o['employee_name'] or '—')
        dc(3, STATUS_HR.get(o['status'], o['status']))
        dc(4, ', '.join(o['missing_fields']), wrap=True)
        ws.row_dimensions[r].height = 20

    col_widths = [14, 22, 14, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='Nepotpuni_PN.xlsx', as_attachment=True)


# ─── VEHICLE LOG ──────────────────────────────────────────────────────────────

MONTH_NAMES = {1:'Siječanj',2:'Veljača',3:'Ožujak',4:'Travanj',5:'Svibanj',
               6:'Lipanj',7:'Srpanj',8:'Kolovoz',9:'Rujan',10:'Listopad',
               11:'Studeni',12:'Prosinac'}

@app.route('/vehicle-log')
@login_required
def vehicle_log_list():
    user = get_current_user()
    conn = get_db()
    if not user_has_vehicle_log_access(user, conn):
        conn.close()
        return render_template('403.html', user=user), 403
    audit('view', module='sluzbeni_automobil', entity='list')
    logs = conn.execute('''
        SELECT vl.*, v.name as vehicle_name, v.reg_plate
        FROM vehicle_log vl
        LEFT JOIN vehicles v ON vl.vehicle_id = v.id
        ORDER BY vl.year DESC, vl.month DESC
    ''').fetchall()
    vehicles = get_vehicles_for_user(conn, user)
    conn.close()
    log_list = []
    for log in logs:
        d = dict(log)
        d['month_name'] = MONTH_NAMES.get(d['month'], '')
        log_list.append(d)
    return render_template('vehicle_log.html', logs=log_list,
                          vehicles=vehicles,
                          month_names=MONTH_NAMES,
                          active='vehicle-log')

@app.route('/vehicle-log/new')
@login_required
def vehicle_log_new():
    conn = get_db()
    user = get_current_user()
    if not user_can_edit_vehicle_log(user, conn):
        conn.close()
        return render_template('403.html', user=user), 403
    vehicles = get_vehicles_for_user(conn, user)
    now = datetime.now()
    prev_month = now.month - 1 if now.month > 1 else 12
    prev_year = now.year if now.month > 1 else now.year - 1
    prev_log = conn.execute(
        "SELECT end_km FROM vehicle_log WHERE year=? AND month=? ORDER BY id DESC LIMIT 1",
        (prev_year, prev_month)
    ).fetchone()
    prev_end_km = prev_log['end_km'] if prev_log and prev_log['end_km'] else None
    conn.close()
    return render_template('vehicle_log_form.html',
                          vehicles=vehicles,
                          log=None, preview=None, pn_list=[],
                          month_names=MONTH_NAMES, active='vehicle-log',
                          now_month=datetime.now().month,
                          prev_end_km=prev_end_km,
                          can_edit=True)

@app.route('/vehicle-log/<int:log_id>/edit')
@login_required
def vehicle_log_edit(log_id):
    user = get_current_user()
    conn = get_db()
    if not user_has_vehicle_log_access(user, conn):
        conn.close()
        return render_template('403.html', user=user), 403
    log = conn.execute("SELECT * FROM vehicle_log WHERE id=?", (log_id,)).fetchone()
    if not log: 
        conn.close()
        return redirect(url_for('vehicle_log_list'))
    can_edit = user_can_edit_vehicle_log(user, conn, log_id)
    vehicles = get_vehicles_for_user(conn, user)
    pn_list = _get_pn_for_month(conn, log['year'], log['month'])
    director = conn.execute("SELECT signature_path FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    dir_sig = director['signature_path'] if director and director['signature_path'] else None
    # Dohvati ime vozila za readonly prikaz
    vehicle_name = None
    if log['vehicle_id']:
        veh = conn.execute("SELECT name, reg_plate FROM vehicles WHERE id=?", (log['vehicle_id'],)).fetchone()
        if veh:
            vehicle_name = f"{veh['name']} {veh['reg_plate'] or ''}".strip()
    conn.close()
    log_dict = dict(log)
    log_dict['vehicle_name'] = vehicle_name
    return render_template('vehicle_log_form.html',
                          vehicles=vehicles,
                          log=log_dict, preview=None, pn_list=pn_list,
                          month_names=MONTH_NAMES, active='vehicle-log',
                          now_month=datetime.now().month,
                          director_sig=dir_sig,
                          can_edit=can_edit)

def _get_pn_for_month(conn, year, month):
    """Find PN nalozi whose departure_date falls in given year/month."""
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{last_day:02d}"
    rows = conn.execute('''
        SELECT to2.auto_id, to2.departure_date, to2.start_km, to2.end_km,
               to2.destination, e.name as employee_name
        FROM travel_orders to2
        LEFT JOIN employees e ON to2.employee_id = e.id
        WHERE (to2.is_deleted=0 OR to2.is_deleted IS NULL)
          AND to2.departure_date >= ? AND to2.departure_date <= ?
        ORDER BY to2.departure_date
    ''', (date_from, date_to)).fetchall()
    return rows_to_dicts(rows)

@app.route('/api/vehicle-log/parse-csv', methods=['POST'])
@login_required
def parse_vehicle_csv():
    """Parse Toyota CSV and return preview data."""
    import io as _io
    from collections import defaultdict

    if 'csv_file' not in request.files:
        return jsonify({'error': 'No file'}), 400

    f = request.files['csv_file']
    content = f.read().decode('utf-8', errors='replace')
    lines = content.strip().split('\n')

    # Parsiramo svaku vožnju zasebno
    # Dohvati home adresu — prvo s vozila, fallback na globalne postavke
    vehicle_id = request.form.get('vehicle_id', '').strip()
    _conn_s = get_db()
    _s = {r['key']: r['value'] for r in _conn_s.execute("SELECT * FROM settings").fetchall()}
    home_addr = ''
    home_city = ''
    if vehicle_id:
        _veh = _conn_s.execute("SELECT home_address, home_city FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
        if _veh:
            home_addr = (_veh['home_address'] or '').strip()
            home_city = (_veh['home_city'] or '').strip()
    if not home_addr:
        home_addr = _s.get('vehicle_home_address', '').strip()
    if not home_city:
        home_city = _s.get('vehicle_home_city', '').strip()
    _conn_s.close()
    home_city = home_city or ''

    def is_home(addr):
        # Vraća True SAMO ako adresa sadrži točnu kućnu adresu
        # Matični grad se NE koristi za provjeru odredišta — npr. "31000 Osijek" je u svim osječkim adresama
        if home_addr and home_addr.lower() in addr.lower():
            return True
        return False

    def is_home_city(addr):
        # Koristi se samo za provjeru POLAZIŠTA (je li to lokalni dan)
        if home_addr and home_addr.lower() in addr.lower():
            return True
        return home_city.lower() in addr.lower()

    trips = []
    for line in lines[1:]:
        parts = line.strip().split(';')
        if len(parts) >= 5:
            try:
                dep_datetime = parts[1]
                dep_date = dep_datetime[:10]
                dep_time = dep_datetime[11:16]
                km = float(parts[4].replace(',', '.'))
                dep_addr = parts[0].strip()
                dest_addr = parts[2].strip()
                trips.append({
                    'date': dep_date,
                    'time': dep_time,
                    'km': round(km, 2),
                    'dep': dep_addr,
                    'dest': dest_addr,
                    'is_pn': False,  # klasificira se u drugom prolazu
                })
            except: pass

    if not trips:
        return jsonify({'error': 'No data found in CSV'}), 400

    # Sortiraj po datumu i vremenu PRIJE klasifikacije
    trips.sort(key=lambda r: (r['date'], r['time']))

    # Grupiraj po danu
    from collections import defaultdict
    daily_trips = defaultdict(list)
    for t in trips:
        daily_trips[t['date']].append(t)

    # Trip chain klasifikacija — kronološki po danu
    for date in sorted(daily_trips.keys()):
        day_trips = daily_trips[date]
        in_pn_chain = False
        for t in day_trips:
            dest_home = is_home(t['dest'])  # stroga provjera — samo točna kućna adresa
            basic_pn = t['km'] > 50 or (not is_home_city(t['dep']) and not is_home_city(t['dest']))
            if not in_pn_chain:
                if basic_pn:
                    in_pn_chain = True
                    t['is_pn'] = True
                else:
                    t['is_pn'] = False
            else:
                t['is_pn'] = True
                if dest_home:
                    in_pn_chain = False

    # Izračunaj dnevne zbirove
    daily_km = {d: round(sum(t['km'] for t in ts), 2) for d, ts in daily_trips.items()}
    daily_pn_km = {d: round(sum(t['km'] for t in ts if t['is_pn']), 2) for d, ts in daily_trips.items()}

    all_dates = sorted(daily_km.keys())
    first_date = all_dates[0]
    year = int(first_date[:4])
    month = int(first_date[5:7])

    total_km = round(sum(daily_km.values()), 2)
    official_km = round(sum(daily_pn_km.values()), 2)
    private_km = round(total_km - official_km, 2)

    import calendar
    last_day = calendar.monthrange(year, month)[1]
    daily_breakdown = []
    cumulative = 0
    for day in range(1, last_day + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        km_day = daily_km.get(date_str, 0)
        pn_day = daily_pn_km.get(date_str, 0)
        priv_day = round(km_day - pn_day, 2)
        has_mixed = pn_day > 0 and priv_day > 0  # dan ima i PN i privatno
        daily_breakdown.append({
            'date': date_str,
            'km': km_day,
            'is_pn': pn_day > 0 and priv_day == 0,
            'pn_km': pn_day,
            'private_km': priv_day,
            'is_mixed': has_mixed,
            'cumulative_km': round(cumulative, 2),
            'trips': daily_trips.get(date_str, []),
        })
        cumulative = round(cumulative + km_day, 2)

    return jsonify({
        'year': year,
        'month': month,
        'total_km': total_km,
        'official_km': official_km,
        'private_km': private_km,
        'daily_breakdown': daily_breakdown,
        'filename': f.filename if hasattr(f, 'filename') else ''
    })

@app.route('/api/vehicle-log', methods=['POST'])
@login_required
def save_vehicle_log():
    data = request.json
    conn = get_db()
    user = get_current_user()
    log_id = data.get('id')
    if not user_can_edit_vehicle_log(user, conn, int(log_id) if log_id else None):
        conn.close()
        return jsonify({'error': 'Nemate pravo uređivanja evidencije vozila'}), 403
    fields = {
        'vehicle_id': data.get('vehicle_id') or None,
        'year': int(data.get('year', 2026)),
        'month': int(data.get('month', 1)),
        'start_km': float(data.get('start_km', 0)),
        'end_km': float(data.get('end_km', 0)),
        'total_km': float(data.get('total_km', 0)),
        'official_km': float(data.get('official_km', 0)),
        'private_km': float(data.get('private_km', 0)),
        'notes': data.get('notes', ''),
        'updated_at': datetime.now().isoformat(),
    }
    if log_id:
        sets = ', '.join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE vehicle_log SET {sets} WHERE id=?", list(fields.values()) + [log_id])
    else:
        # Novi zapis — spremi potpis zaposlenika (trenutnog korisnika)
        user = get_current_user()
        if user and user.get('user_id'):
            emp = conn.execute(
                "SELECT signature_path FROM employees WHERE id=(SELECT employee_id FROM users WHERE id=?)",
                (user['user_id'],)
            ).fetchone()
            if emp and emp['signature_path']:
                fields['employee_signature_path'] = emp['signature_path']
        fields['created_at'] = datetime.now().isoformat()
        cols = ', '.join(fields.keys())
        placeholders = ', '.join('?' for _ in fields)
        c = conn.execute(f"INSERT INTO vehicle_log ({cols}) VALUES ({placeholders})", list(fields.values()))
        log_id = c.lastrowid
    # Spremi dnevne podatke
    daily_days = data.get('daily_days', [])
    if daily_days:
        conn.execute("DELETE FROM vehicle_log_days WHERE log_id=?", (log_id,))
        for d in daily_days:
            import json as _json
            trips_json = _json.dumps(d.get('trips', []), ensure_ascii=False) if d.get('trips') else None
            conn.execute("""INSERT OR REPLACE INTO vehicle_log_days
                (log_id, date, start_km, end_km, official_km, private_km, total_km, comment, is_pn, trips_json)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (log_id, d['date'],
                 float(d.get('start_km', 0)), float(d.get('end_km', 0)),
                 float(d.get('official_km', 0)), float(d.get('private_km', 0)),
                 float(d.get('total_km', 0)),
                 d.get('comment', ''), 1 if d.get('is_pn') else 0, trips_json))

    conn.commit()
    conn.close()
    audit('create' if not data.get('id') else 'edit', module='sluzbeni_automobil', entity='vehicle_log', entity_id=log_id)
    return jsonify({'success': True, 'id': log_id})

@app.route('/api/vehicle-log/<int:log_id>/days', methods=['GET'])
@require_perm('can_view_vehicle_log')
def get_vehicle_log_days(log_id):
    import json as _json
    conn = get_db()
    days = conn.execute(
        "SELECT * FROM vehicle_log_days WHERE log_id=? ORDER BY date", (log_id,)
    ).fetchall()
    conn.close()
    result = []
    for d in days:
        row = dict(d)
        if row.get('trips_json'):
            try:
                row['trips'] = _json.loads(row['trips_json'])
            except:
                row['trips'] = []
        else:
            row['trips'] = []
        result.append(row)
    return jsonify(result)

@app.route('/api/vehicle-log/<int:log_id>/approve', methods=['POST'])
@login_required
def approve_vehicle_log(log_id):
    """Dodaj potpis direktora kao odobrenje."""
    conn = get_db()
    user = get_current_user()
    if not user_can_edit_vehicle_log(user, conn, log_id):
        conn.close()
        return jsonify({'error': 'Nemate pravo odobravanja evidencije'}), 403
    director = conn.execute("SELECT * FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    if not director or not director['signature_path']:
        conn.close()
        return jsonify({'error': 'Direktor nema uploadanog potpisa'}), 400
    user = get_current_user()
    conn.execute(
        "UPDATE vehicle_log SET is_approved=1, approved_at=?, approved_by_id=? WHERE id=?",
        (datetime.now().isoformat(), user.get('user_id') if user else None, log_id)
    )
    conn.commit()
    conn.close()
    audit('approve', module='sluzbeni_automobil', entity='vehicle_log', entity_id=log_id)
    return jsonify({'success': True, 'signature_path': director['signature_path']})

@app.route('/api/vehicle-log/<int:log_id>', methods=['DELETE'])
@login_required
def delete_vehicle_log(log_id):
    conn = get_db()
    user = get_current_user()
    if not user_can_edit_vehicle_log(user, conn, log_id):
        conn.close()
        return jsonify({'error': 'Nemate pravo brisanja evidencije'}), 403
    audit('delete', module='sluzbeni_automobil', entity='vehicle_log', entity_id=log_id)
    conn = get_db()
    conn.execute("DELETE FROM vehicle_log WHERE id=?", (log_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/vehicle-log/<int:log_id>/pn')
@require_perm('can_view_vehicle_log')
def vehicle_log_pn(log_id):
    conn = get_db()
    log = conn.execute("SELECT year, month FROM vehicle_log WHERE id=?", (log_id,)).fetchone()
    if not log:
        conn.close()
        return jsonify([])
    pn_list = _get_pn_for_month(conn, log['year'], log['month'])
    conn.close()
    return jsonify(pn_list)

@app.route('/vehicle-log/<int:log_id>/excel')
@require_perm('can_view_vehicle_log')
def vehicle_log_excel(log_id):
    import calendar as _cal, io as _io, os

    conn = get_db()
    log = conn.execute("SELECT * FROM vehicle_log WHERE id=?", (log_id,)).fetchone()
    if not log: return "Not found", 404
    pn_list = _get_pn_for_month(conn, log['year'], log['month'])
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (log['vehicle_id'],)).fetchone() if log['vehicle_id'] else None
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    # Dohvati potpise
    director = conn.execute("SELECT * FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    conn.close()

    log = dict(log)
    year, month = log['year'], log['month']
    month_name = MONTH_NAMES.get(month, str(month))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()

    NAVY  = "1A3A5C"
    LIGHT = "E8F0F7"
    WHITE = "FFFFFF"
    BORDER_C = "AAC4DB"
    thin = Side(style='thin', color=BORDER_C)
    med  = Side(style='medium', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    med_border = Border(left=med, right=med, top=med, bottom=med)

    def hc(cell, val, bold=False, bg=None, align='center', size=10, color='000000', wrap=False):
        cell.value = val
        cell.font = Font(name='Arial', bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if bg: cell.fill = PatternFill('solid', start_color=bg)
        cell.border = border

    # ═══════════════════════════════════════════════════════
    # SHEET 1: Sažetak
    # ═══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Sažetak'

    # Column widths — A:spacer, B-E: 4 široke kolone, F+G: dvije uže koje zajedno = 1 široka
    # Tako svih 5 rekapitulacijskih kolona ima jednaku vizualnu širinu, a PN tablica pokriva B:G
    col_widths = [('A',2),('B',17),('C',17),('D',17),('E',17),('F',8.5),('G',8.5)]
    for col, w in col_widths:
        ws.column_dimensions[col].width = w

    # Row 2: Title — span B:G
    ws.merge_cells('B2:G2')
    ws['B2'].value = f"MJESEČNO IZVJEŠĆE O KORIŠTENJU SLUŽBENOG VOZILA — {month_name.upper()} {year}"
    ws['B2'].font = Font(name='Arial', bold=True, size=13, color=WHITE)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].fill = PatternFill('solid', start_color=NAVY)
    ws['B2'].border = border
    ws.row_dimensions[2].height = 34

    # Rows 4-7: Meta info
    meta = [
        ('Vozilo:', (vehicle['name'] + ' ' + vehicle['reg_plate']) if vehicle else '—'),
        ('Vlasnik:', settings.get('company_name', '')),
        ('Korisnik:', settings.get('company_director', '')),
        ('Mjesec:', f"{month_name} {year}"),
    ]
    for i, (label, val) in enumerate(meta):
        r = 4 + i
        lc = ws.cell(r, 2)
        lc.value = label
        lc.font = Font(name='Arial', bold=True, size=9, color=NAVY)
        lc.alignment = Alignment(vertical='center')
        vc = ws.cell(r, 3)
        vc.value = val
        vc.font = Font(name='Arial', size=9)
        vc.alignment = Alignment(vertical='center')
        ws.row_dimensions[r].height = 16

    # Row 9: REKAPITULACIJA header — B:G (puni raspon)
    ws.merge_cells('B9:G9')
    hc(ws['B9'], 'REKAPITULACIJA', bold=True, bg=NAVY, align='center', color=WHITE, size=11)
    ws.row_dimensions[9].height = 26

    # Row 10: Rekapitulacija headers
    # B,C,D,E = po 1 stupac (širina 17); Privatno km = merge F:G (8.5+8.5=17) → jednaka vizualna širina
    recap_headers = ['Početna km', 'Završna km', 'Ukupno km', 'Službeno km', 'Privatno km']
    for ci, h in enumerate(recap_headers):
        col = 2 + ci
        if ci == 4:
            ws.merge_cells(start_row=10, start_column=6, end_row=10, end_column=7)
        c = ws.cell(10, col)
        hc(c, h, bold=True, bg=LIGHT, align='center', size=9)
    ws.row_dimensions[10].height = 22

    # Row 11: Rekapitulacija values
    vals = [log.get('start_km'), log.get('end_km'), log.get('total_km'), log.get('official_km'), log.get('private_km')]
    colors_vals = [None, None, None, '27AE60', 'E67E22']
    for ci, (v, fc) in enumerate(zip(vals, colors_vals)):
        col = 2 + ci
        if ci == 4:
            ws.merge_cells(start_row=11, start_column=6, end_row=11, end_column=7)
        c = ws.cell(11, col)
        c.value = v
        c.number_format = '#,##0.00'
        c.font = Font(name='Arial', bold=True, size=12, color=fc or '000000')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[11].height = 28
    # Eksplicitni desni border na G10 i G11 (rub merged Privatno km ćelije)
    from openpyxl.styles import Border as _B, Side as _S
    _rb = _B(right=_S(style='thin', color=BORDER_C), top=_S(style='thin', color=BORDER_C), bottom=_S(style='thin', color=BORDER_C))
    ws.cell(10, 7).border = _rb
    ws.cell(10, 7).fill = PatternFill('solid', start_color=LIGHT)
    ws.cell(11, 7).border = _rb

    # Row 13: PN NALOZI header (ako postoje)
    pn_start_row = 13
    if pn_list:
        ws.merge_cells(f'B{pn_start_row}:G{pn_start_row}')
        hc(ws[f'B{pn_start_row}'], 'PUTNI NALOZI U OVOM MJESECU', bold=True, bg=NAVY, align='center', color=WHITE, size=10)
        ws.row_dimensions[pn_start_row].height = 24

        # 6 kolona u B:G: Br.PN | Datum | Destinacija | Zaposlenik | Poč.km | Zav.km
        pn_headers = ['Br. PN', 'Datum', 'Destinacija', 'Zaposlenik', 'Poč. km', 'Zav. km']
        for ci, h in enumerate(pn_headers):
            hc(ws.cell(pn_start_row+1, 2+ci), h, bold=True, bg=LIGHT, align='center', size=9)
        ws.row_dimensions[pn_start_row+1].height = 20

        for j, pn in enumerate(pn_list):
            r = pn_start_row + 2 + j
            bg = WHITE if j % 2 == 0 else 'F0F4F8'
            row_vals = [
                f"PN {pn['auto_id']}", pn.get('departure_date',''),
                pn.get('destination',''), pn.get('employee_name',''),
                pn.get('start_km',''), pn.get('end_km','')
            ]
            for ci, val in enumerate(row_vals):
                c = ws.cell(r, 2+ci)
                c.value = val
                c.font = Font(name='Arial', size=9)
                c.fill = PatternFill('solid', start_color=bg)
                c.border = border
                c.alignment = Alignment(vertical='center', horizontal='center')
            ws.row_dimensions[r].height = 18

    # Napomene
    note_row = pn_start_row + 2 + len(pn_list) + 2
    if log.get('notes'):
        ws.cell(note_row, 2).value = 'Napomene:'
        ws.cell(note_row, 2).font = Font(name='Arial', bold=True, size=9, color=NAVY)
        ws.merge_cells(f'B{note_row+1}:G{note_row+1}')
        ws.cell(note_row+1, 2).value = log['notes']
        ws.cell(note_row+1, 2).font = Font(name='Arial', size=9)
        ws.cell(note_row+1, 2).alignment = Alignment(wrap_text=True)
        note_row += 3

    # Potpisi (Sažetak sheet)
    sig_row = note_row + 2
    ws.row_dimensions[sig_row].height = 16
    import calendar as _cal_xl
    _last_day = _cal_xl.monthrange(year, month)[1]
    ws.cell(sig_row, 2).value = f"Datum: {_last_day:02d}.{month:02d}.{year}."
    ws.cell(sig_row, 2).font = Font(name='Arial', size=9)
    ws.cell(sig_row, 3).value = 'Potpis zaposlenika:'
    ws.cell(sig_row, 3).font = Font(name='Arial', size=9, bold=True, color=NAVY)
    ws.cell(sig_row, 5).value = 'Odobrio:'
    ws.cell(sig_row, 5).font = Font(name='Arial', size=9, bold=True, color=NAVY)

    # Slike potpisa
    app_dir = os.path.dirname(__file__)
    upload_dir = os.path.join(app_dir, 'uploads')

    if log.get('employee_signature_path'):
        sig_path = os.path.join(upload_dir, log['employee_signature_path'])
        if os.path.exists(sig_path):
            img = XLImage(sig_path)
            img.width, img.height = 200, 60
            ws.add_image(img, f'C{sig_row+1}')
            ws.row_dimensions[sig_row+1].height = 50

    if log.get('is_approved') and director and director['signature_path']:
        dir_sig_path = os.path.join(upload_dir, director['signature_path'])
        if os.path.exists(dir_sig_path):
            img2 = XLImage(dir_sig_path)
            img2.width, img2.height = 160, 50
            ws.add_image(img2, f'E{sig_row+1}')
            if (ws.row_dimensions[sig_row+1].height or 0) < 50:
                ws.row_dimensions[sig_row+1].height = 50

    # ═══════════════════════════════════════════════════════
    # SHEET 2: Evidencija (Toyota-style monthly log)
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title='Evidencija')
    last_day = _cal.monthrange(year, month)[1]

    from openpyxl.styles import Font as F2, Alignment as A2, Border as B2, Side as S2, PatternFill as PF2
    from openpyxl.utils import get_column_letter as gcl

    def th(size=9, bold=False, italic=False, color='000000'):
        return F2(name='Arial', size=size, bold=bold, italic=italic, color=color)
    def al(h='left', v='center', wrap=False):
        return A2(horizontal=h, vertical=v, wrap_text=wrap)
    s_thin = S2(style='thin', color='000000')
    s_med  = S2(style='medium', color='000000')
    def brd(l='thin',r='thin',t='thin',b='thin'):
        mk = lambda s: S2(style=s, color='000000') if s else None
        return B2(left=mk(l), right=mk(r), top=mk(t), bottom=mk(b))
    full_brd = brd()
    no_brd   = B2()
    grey_fill   = PF2('solid', start_color='CCCCCC')
    stripe_fill = PF2('solid', start_color='D9D9D9')
    navy_fill   = PF2('solid', start_color=NAVY)

    def sc2(row, col, val, font=None, align=None, border=None, fill=None, nfmt=None, merge_to=None):
        c = ws2.cell(row, col)
        c.value = val
        if font:   c.font = font
        if align:  c.alignment = align
        if border: c.border = border
        if fill:   c.fill = fill
        if nfmt:   c.number_format = nfmt
        if merge_to:
            ws2.merge_cells(start_row=row, start_column=col, end_row=merge_to[0], end_column=merge_to[1])
        return c

    col_w2 = [('A',1.5),('B',10.5),('C',9.5),('D',9.5),('E',9.5),('F',9.0),('G',13.0),('H',22),('I',14)]
    for col, w in col_w2:
        ws2.column_dimensions[col].width = w

    veh_name  = (vehicle['name']) if vehicle else 'Toyota C-HR'
    reg_plate = vehicle['reg_plate'] if vehicle else ''
    owner     = settings.get('company_name', '')
    user_name = settings.get('company_director', '')
    month_label = f"{month:02d}-{str(year)[2:]}"

    # Row 1 blank
    ws2.row_dimensions[1].height = 14

    # Rows 2-5: info desno
    info_rows = [
        (2, 'Mjesec:', month_label),
        (3, 'Reg. oznaka:', reg_plate),
        (4, 'Vlasnik:', owner),
        (5, 'Korisnik:', user_name),
    ]
    for r, lbl, val in info_rows:
        c_lbl = ws2.cell(r, 7)
        c_lbl.value = lbl
        c_lbl.font = th(9, bold=True)
        c_lbl.alignment = al('left')
        c_lbl.border = full_brd
        ws2.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        c_val = ws2.cell(r, 8)
        c_val.value = val
        c_val.font = th(9, bold=True)
        c_val.alignment = al('center')
        c_val.border = full_brd
        ws2.row_dimensions[r].height = 15

    # Borders za B-F (prazne ćelije) i popravak merged H:I border
    for r in range(2, 6):
        for c in range(2, 7):  # B-F
            cell = ws2.cell(r, c)
            cell.border = full_brd
            if not cell.fill or cell.fill.fill_type == 'none':
                cell.fill = PF2('solid', start_color='F5F8FA')
        # G label i H:I value imaju border iz petlje gore (G=7) — dodaj i za I (merged ali border treba na 9)
        ws2.cell(r, 9).border = full_brd  # desni rub merged H:I ćelije

    ws2.row_dimensions[6].height = 8

    # Rows 7-8: naslov
    sc2(7, 2, 'MJESEČNO IZVJEŠĆE', font=th(11, bold=True, italic=True), align=al('center'), merge_to=(7,9))
    ws2.row_dimensions[7].height = 16
    sc2(8, 2, 'O KORIŠTENJU SLUŽBENOG VOZILA', font=th(11, bold=True, italic=True), align=al('center'), merge_to=(8,9))
    ws2.row_dimensions[8].height = 16
    ws2.row_dimensions[9].height = 8

    # Row 10: tip vozila / namjena
    sc2(10, 2, 'Tip vozila:', font=th(9, bold=True))
    sc2(10, 3, 'Toyota C-HR', font=th(9))
    sc2(10, 7, 'Namjena vozila:', font=th(9, bold=True), align=al('right'))
    sc2(10, 8, 'putničko', font=th(9, bold=True), align=al('center'), border=full_brd, merge_to=(10,9))
    ws2.row_dimensions[10].height = 14
    # Row 11: Gorivo ispod Namjena vozila
    sc2(11, 7, 'Gorivo:', font=th(9, bold=True), align=al('right'))
    sc2(11, 8, 'benzin', font=th(9, bold=True), align=al('center'), border=full_brd, merge_to=(11,9))
    ws2.row_dimensions[11].height = 14
    ws2.row_dimensions[12].height = 8

    # Row 13 (bio 12): DNEVNE UDALJENOSTI — bez Gorivo
    sc2(13, 2, 'DNEVNE UDALJENOSTI', font=th(9, bold=True))
    ws2.row_dimensions[13].height = 16
    ws2.row_dimensions[14].height = 8

    # Row 15: sub-headers
    sc2(15, 2, '', border=full_brd)
    sc2(15, 3, 'Prevaljeni kilometri', font=th(9, bold=True), align=al('center'), border=full_brd, merge_to=(15,5))
    sc2(15, 6, 'Namjena', font=th(9, bold=True), align=al('center'), border=full_brd, merge_to=(15,7))
    sc2(15, 8, 'Relacija', font=th(9), align=al('center'), border=full_brd, merge_to=(15,9))
    ws2.row_dimensions[15].height = 14

    # Row 16: column labels
    labels16 = [(2,'Datum'),(3,'početno'),(4,'završno'),(5,'službeno'),(6,'privatno'),(7,'ukupno'),(8,'komentar')]
    for col, lbl in labels16:
        c = ws2.cell(16, col)
        c.value = lbl
        c.font = th(9, bold=True)
        c.alignment = al('center')
        c.border = full_brd
        c.fill = stripe_fill
    ws2.merge_cells(start_row=16, start_column=8, end_row=16, end_column=9)
    ws2.row_dimensions[16].height = 14
    ws2.row_dimensions[17].height = 6

    # PN mapa po datumu — pokriva sve dane višednevnih PN-ova
    from datetime import date as _date, timedelta as _td
    pn_by_day = {}
    for pn in pn_list:
        dep = pn.get('departure_date', '')
        if not dep:
            continue
        entry = {
            'label': f"PN {pn['auto_id']}",
            'start_km': float(pn.get('start_km') or 0),
            'end_km': float(pn.get('end_km') or 0),
        }
        trip_start = (pn.get('trip_start_datetime') or dep)[:10]
        trip_end   = (pn.get('trip_end_datetime')   or dep)[:10]
        try:
            d_start = _date.fromisoformat(trip_start)
            d_end   = _date.fromisoformat(trip_end)
            if d_end < d_start: d_end = d_start
        except:
            d_start = d_end = _date.fromisoformat(dep[:10])
        cur_d = d_start
        while cur_d <= d_end:
            pn_by_day[cur_d.isoformat()] = entry
            cur_d += _td(days=1)

    # Dohvati stvarne dnevne podatke iz baze
    conn_xl = get_db()
    saved_days_xl = conn_xl.execute(
        "SELECT * FROM vehicle_log_days WHERE log_id=? ORDER BY date", (log_id,)
    ).fetchall()
    conn_xl.close()
    saved_days_xl_dict = {r['date']: dict(r) for r in saved_days_xl}

    total_official = float(log.get('official_km') or 0)
    total_private  = float(log.get('private_km') or 0)
    start_km_val   = float(log.get('start_km') or 0)

    pn_dates = list(pn_by_day.keys())
    non_pn_count = last_day - len(set(pn_dates))
    private_per_day = round(total_private / max(non_pn_count, 1), 2) if non_pn_count > 0 else 0
    pn_official = {}
    if pn_dates and total_official > 0:
        per_pn = round(total_official / len(pn_dates), 2)
        for d in pn_dates:
            pn_official[d] = per_pn

    cur_km = start_km_val
    for day in range(1, last_day + 1):
        r = 16 + day + 1  # data počinje od row 18
        date_str = f"{year}-{month:02d}-{day:02d}"
        is_pn = date_str in pn_by_day

        if date_str in saved_days_xl_dict:
            sd = saved_days_xl_dict[date_str]
            start_km_d = float(sd.get('start_km') or cur_km)
            end_km_day = float(sd.get('end_km') or cur_km)
            official   = float(sd.get('official_km') or 0)
            private    = float(sd.get('private_km') or 0)
            total_day  = float(sd.get('total_km') or 0)
            raw_comment = sd.get('comment') or ''
            if is_pn:
                pn_label = pn_by_day[date_str]['label']
                if raw_comment in ('PN', ''):
                    comment = pn_label
                elif raw_comment == 'PN+privatno':
                    comment = f"{pn_label}+privatno"
                else:
                    comment = raw_comment
            elif official > 0 and raw_comment in ('PN', 'PN+privatno'):
                # Dan ima službene km ali nije u pn_by_day (trip_end_datetime nije postavljen)
                # Zadrži comment ali označi crvenom
                comment = raw_comment
            else:
                comment = raw_comment if raw_comment else ('privatno' if private > 0 else '')
            cur_km     = end_km_day
        else:
            official   = pn_official.get(date_str, 0) if is_pn else 0
            private    = 0 if is_pn else private_per_day
            total_day  = round(official + private, 2)
            end_km_day = round(cur_km + total_day, 2)
            pn_label   = pn_by_day[date_str]['label'] if is_pn else ''
            comment    = pn_label if is_pn else ('privatno' if private > 0 else '')
            cur_km     = end_km_day

        fill2 = stripe_fill if day % 2 == 0 else None

        sc2(r, 2, date_str, font=th(9), align=al('center'), border=full_brd, fill=fill2)
        sc2(r, 3, round(cur_km, 0), font=th(9), align=al('center'), border=full_brd, fill=fill2, nfmt='#,##0')
        sc2(r, 4, end_km_day, font=th(9), align=al('center'), border=full_brd, fill=fill2, nfmt='#,##0')
        sc2(r, 5, official if official else 0, font=th(9), align=al('center'), border=full_brd, fill=fill2, nfmt='#,##0')
        sc2(r, 6, private if private else 0, font=th(9), align=al('center'), border=full_brd, fill=fill2, nfmt='#,##0')
        sc2(r, 7, total_day, font=th(9), align=al('center'), border=full_brd, fill=fill2, nfmt='#,##0')
        sc2(r, 8, comment, font=th(9, color='C0392B' if (comment.startswith('PN ') or (official > 0 and comment in ('PN', 'PN+privatno'))) else '000000'), align=al('center'), border=full_brd, fill=fill2, merge_to=(r,9))
        ws2.row_dimensions[r].height = 14
        cur_km = end_km_day

    # Ukupno row
    tr = 16 + last_day + 2
    sc2(tr, 2, 'Ukupno :', font=th(9, bold=True), align=al('right'), border=full_brd, fill=stripe_fill)
    sc2(tr, 3, round(cur_km - start_km_val, 0), font=th(9, bold=True), align=al('center'), border=full_brd, fill=stripe_fill, nfmt='#,##0')
    sc2(tr, 4, '', border=full_brd, fill=stripe_fill)
    sc2(tr, 5, round(total_official, 0), font=th(9, bold=True), align=al('center'), border=full_brd, fill=stripe_fill, nfmt='#,##0')
    sc2(tr, 6, round(total_private, 0), font=th(9, bold=True), align=al('center'), border=full_brd, fill=stripe_fill, nfmt='#,##0')
    sc2(tr, 7, round(total_official+total_private, 0), font=th(9, bold=True), align=al('center'), border=full_brd, fill=stripe_fill, nfmt='#,##0')
    sc2(tr, 8, '', border=full_brd, fill=stripe_fill, merge_to=(tr,9))
    ws2.row_dimensions[tr].height = 18

    # Završno stanje
    sr = tr + 2
    last_date = f"{year}-{month:02d}-{last_day:02d}"
    sc2(sr, 2, last_date, font=th(9, bold=True), align=al('center'), border=full_brd, fill=stripe_fill)
    sc2(sr, 3, round(cur_km, 0), font=th(9, bold=True), align=al('center'), border=full_brd, fill=stripe_fill, nfmt='#,##0')
    sc2(sr, 4, '     ( Završno stanje KM  -  prijenos u slijedeći mjesec )', font=th(9, italic=True), align=al('center'), merge_to=(sr,9))
    ws2.row_dimensions[sr].height = 18

    # Potpisi row
    sigr2 = sr + 2
    ws2.row_dimensions[sigr2].height = 16
    ws2.row_dimensions[sigr2+1].height = 48
    sc2(sigr2, 2, 'Datum:', font=th(9))
    sc2(sigr2, 3, f"{last_day:02d}.{month:02d}.{year}", font=th(9))
    sc2(sigr2, 4, 'Potpis:', font=th(9))
    sc2(sigr2, 7, 'Odobrio:', font=th(9))

    # Slike potpisa u Evidencija sheet
    app_dir = os.path.dirname(__file__)
    upload_dir = os.path.join(app_dir, 'uploads')

    if log.get('employee_signature_path'):
        sig_path = os.path.join(upload_dir, log['employee_signature_path'])
        if os.path.exists(sig_path):
            img_e = XLImage(sig_path)
            img_e.width, img_e.height = 200, 55
            ws2.add_image(img_e, f'D{sigr2+1}')
            ws2.row_dimensions[sigr2+1].height = 48

    if log.get('is_approved') and director and director['signature_path']:
        dir_sig_path = os.path.join(upload_dir, director['signature_path'])
        if os.path.exists(dir_sig_path):
            img_d = XLImage(dir_sig_path)
            img_d.width, img_d.height = 100, 30
            ws2.add_image(img_d, f'G{sigr2+1}')

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Evidencija_{month:02d}{year}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)

@app.route('/vehicle-log/<int:log_id>/pdf')
@require_perm('can_view_vehicle_log')
def vehicle_log_pdf(log_id):
    """Generiraj PDF izvještaj evidencije vozila."""
    import io, os, calendar as _cal
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, HRFlowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    conn = get_db()
    log = conn.execute("SELECT * FROM vehicle_log WHERE id=?", (log_id,)).fetchone()
    if not log: return "Not found", 404
    pn_list = _get_pn_for_month(conn, log['year'], log['month'])
    vehicle  = conn.execute("SELECT * FROM vehicles WHERE id=?", (log['vehicle_id'],)).fetchone() if log['vehicle_id'] else None
    director = conn.execute("SELECT * FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()

    log = dict(log)
    year, month = log['year'], log['month']
    month_name = MONTH_NAMES.get(month, str(month))
    last_day = _cal.monthrange(year, month)[1]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=14*mm, bottomMargin=14*mm,
                            leftMargin=18*mm, rightMargin=18*mm)
    W = A4[0] - 36*mm
    styles = getSampleStyleSheet()

    # Fontovi
    try:
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        arial_path = '/Library/Fonts/Arial Unicode.ttf'
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont('ArialU', arial_path))
            NF, BF = 'ArialU', 'ArialU'
        else:
            NF, BF = 'Helvetica', 'Helvetica-Bold'
    except:
        NF, BF = 'Helvetica', 'Helvetica-Bold'

    BLUE  = colors.HexColor('#1a3a5c')
    LBLUE = colors.HexColor('#e8f0f7')
    GRAY  = colors.HexColor('#666666')
    GREEN = colors.HexColor('#27ae60')
    ORANGE= colors.HexColor('#e67e22')

    def ps(name='n', size=9, bold=False, align=TA_LEFT, color=colors.black):
        return ParagraphStyle(name, parent=styles['Normal'],
                              fontSize=size, fontName=BF if bold else NF,
                              alignment=align, textColor=color, leading=size*1.3)

    def P(txt, **kw): return Paragraph(str(txt) if txt else '', ps(**kw))

    app_dir = os.path.dirname(__file__)
    upload_dir = os.path.join(app_dir, 'uploads')

    story = []

    # ── HEADER: logo + naslov ──────────────────────────────────────────────
    conn_logo = get_db()
    _logo_row = conn_logo.execute("SELECT value FROM settings WHERE key='company_logo'").fetchone()
    conn_logo.close()
    _logo_file = _logo_row['value'] if _logo_row and _logo_row['value'] else 'logo.png'
    logo_path = os.path.join(app_dir, 'static', _logo_file)
    if not os.path.exists(logo_path):
        logo_path = os.path.join(app_dir, 'logo.png')
    title_block = [
        P('EVIDENCIJA KORIŠTENJA SLUŽBENOG VOZILA', name='t', size=13, bold=True, align=TA_RIGHT, color=BLUE),
        P(f"{month_name.upper()} {year}", name='s', size=10, align=TA_RIGHT, color=GRAY),
    ]
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=41.2*mm, height=12.7*mm)
        hdr = Table([[logo, title_block]], colWidths=[70*mm, W-70*mm])
    else:
        hdr = Table([['', title_block]], colWidths=[70*mm, W-70*mm])
    hdr.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT')]))
    story.append(hdr)
    story.append(HRFlowable(width=W, thickness=1.5, color=BLUE, spaceAfter=6))

    # ── META INFO ──────────────────────────────────────────────────────────
    veh_str = (vehicle['name'] + ' ' + vehicle['reg_plate']) if vehicle else '—'
    meta = Table([
        [P('Vozilo:', bold=True, color=BLUE), P(veh_str),
         P('Vlasnik:', bold=True, color=BLUE), P(settings.get('company_name',''))],
        [P('Korisnik:', bold=True, color=BLUE), P(settings.get('company_director','')),
         P('Razdoblje:', bold=True, color=BLUE), P(f"{month_name} {year}")],
    ], colWidths=[28*mm, 60*mm, 28*mm, 54*mm])
    meta.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    story.append(meta)
    story.append(Spacer(1, 6))

    # ── REKAPITULACIJA ─────────────────────────────────────────────────────
    story.append(Table([[P('REKAPITULACIJA', bold=True, align=TA_CENTER, color=colors.white)]],
                       colWidths=[W],
                       style=TableStyle([('BACKGROUND',(0,0),(0,0),BLUE),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)])))
    recap_data = [
        [P('Početna km', bold=True, align=TA_CENTER), P('Završna km', bold=True, align=TA_CENTER),
         P('Ukupno km', bold=True, align=TA_CENTER), P('Službeno km', bold=True, align=TA_CENTER),
         P('Privatno km', bold=True, align=TA_CENTER)],
        [P(f"{log.get('start_km',0):,.2f}", align=TA_CENTER, size=12),
         P(f"{log.get('end_km',0):,.2f}", align=TA_CENTER, size=12),
         P(f"{log.get('total_km',0):,.2f}", align=TA_CENTER, size=12),
         P(f"{log.get('official_km',0):,.2f}", align=TA_CENTER, size=12, bold=True, color=GREEN),
         P(f"{log.get('private_km',0):,.2f}", align=TA_CENTER, size=12, bold=True, color=ORANGE)],
    ]
    recap = Table(recap_data, colWidths=[W/5]*5)
    recap.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),LBLUE),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#aac4db')),
        ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(recap)
    story.append(Spacer(1, 10))

    # ── PUTNI NALOZI ───────────────────────────────────────────────────────
    if pn_list:
        story.append(Table([[P('PUTNI NALOZI U OVOM MJESECU', bold=True, align=TA_CENTER, color=colors.white)]],
                           colWidths=[W],
                           style=TableStyle([('BACKGROUND',(0,0),(0,0),BLUE),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)])))
        pn_data = [[P('Br. PN', bold=True, align=TA_CENTER), P('Datum', bold=True, align=TA_CENTER),
                    P('Destinacija', bold=True), P('Zaposlenik', bold=True),
                    P('Poč. km', bold=True, align=TA_CENTER), P('Zav. km', bold=True, align=TA_CENTER)]]
        for pn in pn_list:
            pn_data.append([
                P(f"PN {pn['auto_id']}", align=TA_CENTER),
                P(pn.get('departure_date',''), align=TA_CENTER),
                P(pn.get('destination','')),
                P(pn.get('employee_name','')),
                P(str(pn.get('start_km','')), align=TA_CENTER),
                P(str(pn.get('end_km','')), align=TA_CENTER),
            ])
        pn_tbl = Table(pn_data, colWidths=[W*0.13, W*0.15, W*0.26, W*0.27, W*0.095, W*0.095])
        pn_styles = [
            ('BACKGROUND',(0,0),(-1,0),LBLUE),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#aac4db')),
            ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]
        for i in range(1, len(pn_data)):
            if i % 2 == 0:
                pn_styles.append(('BACKGROUND',(0,i),(-1,i),colors.HexColor('#f0f4f8')))
        pn_tbl.setStyle(TableStyle(pn_styles))
        story.append(pn_tbl)
        story.append(Spacer(1, 10))

    # ── NAPOMENE ───────────────────────────────────────────────────────────
    if log.get('notes'):
        story.append(P('Napomene:', bold=True, color=BLUE))
        story.append(Spacer(1, 3))
        story.append(P(log['notes']))
        story.append(Spacer(1, 10))

    # ── STRANICA 2: EVIDENCIJA (dnevni pregled) ────────────────────────────
    from reportlab.platypus import PageBreak
    story.append(PageBreak())

    # Naslov strana 2
    story.append(Table([[P('DNEVNA EVIDENCIJA KORIŠTENJA VOZILA', bold=True, align=TA_CENTER, color=colors.white)]],
                       colWidths=[W],
                       style=TableStyle([('BACKGROUND',(0,0),(0,0),BLUE),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)])))
    story.append(Spacer(1, 6))

    # Dnevni podaci — iz vehicle_log_days ako postoje, inače paušal
    conn2 = get_db()
    saved_days = conn2.execute(
        "SELECT * FROM vehicle_log_days WHERE log_id=? ORDER BY date", (log_id,)
    ).fetchall()
    conn2.close()
    saved_days_dict = {r['date']: dict(r) for r in saved_days}

    # pn_by_day mapa — pokriva sve dane višednevnih PN-ova
    from datetime import date as _date2, timedelta as _td2
    pn_by_day = {}
    for pn in pn_list:
        dep = pn.get('departure_date', '')
        if not dep:
            continue
        label = f"PN {pn['auto_id']}"
        trip_start = (pn.get('trip_start_datetime') or dep)[:10]
        trip_end   = (pn.get('trip_end_datetime')   or dep)[:10]
        try:
            d_s = _date2.fromisoformat(trip_start)
            d_e = _date2.fromisoformat(trip_end)
            if d_e < d_s: d_e = d_s
        except:
            d_s = d_e = _date2.fromisoformat(dep[:10])
        cur_d = d_s
        while cur_d <= d_e:
            pn_by_day[cur_d.isoformat()] = label
            cur_d += _td2(days=1)
    total_official = float(log.get('official_km') or 0)
    total_private  = float(log.get('private_km') or 0)
    start_km_val   = float(log.get('start_km') or 0)
    pn_dates = list(pn_by_day.keys())
    non_pn_count = last_day - len(set(pn_dates))
    private_per_day = round(total_private / max(non_pn_count, 1), 2) if non_pn_count > 0 else 0
    pn_official = {}
    if pn_dates and total_official > 0:
        per_pn = round(total_official / len(pn_dates), 2)
        for d in pn_dates:
            pn_official[d] = per_pn

    daily_data = [[
        P('Datum', bold=True, align=TA_CENTER, size=8), P('Početno', bold=True, align=TA_CENTER, size=8),
        P('Završno', bold=True, align=TA_CENTER, size=8), P('Službeno', bold=True, align=TA_CENTER, size=8),
        P('Privatno', bold=True, align=TA_CENTER, size=8), P('Ukupno', bold=True, align=TA_CENTER, size=8),
        P('Komentar', bold=True, align=TA_CENTER, size=8),
    ]]

    cur_km = start_km_val
    for day in range(1, last_day + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        is_pn = date_str in pn_by_day

        if date_str in saved_days_dict:
            sd = saved_days_dict[date_str]
            start_km_d = float(sd.get('start_km') or cur_km)
            end_km_d   = float(sd.get('end_km') or cur_km)
            official   = float(sd.get('official_km') or 0)
            private    = float(sd.get('private_km') or 0)
            total_d    = float(sd.get('total_km') or 0)
            raw_comment = sd.get('comment') or ''
            if is_pn:
                pn_label = pn_by_day[date_str]
                if raw_comment in ('PN', ''):
                    comment = pn_label
                elif raw_comment == 'PN+privatno':
                    comment = f"{pn_label}+privatno"
                else:
                    comment = raw_comment
            elif official > 0 and raw_comment in ('PN', 'PN+privatno'):
                comment = raw_comment
            else:
                comment = raw_comment if raw_comment else ('privatno' if private > 0 else '')
            cur_km     = end_km_d
        else:
            start_km_d = cur_km
            official   = pn_official.get(date_str, 0) if is_pn else 0
            private    = 0 if is_pn else private_per_day
            total_d    = round(official + private, 2)
            end_km_d   = round(cur_km + total_d, 2)
            pn_label   = pn_by_day[date_str] if is_pn else ''
            comment    = pn_label if is_pn else ('privatno' if private > 0 else '')
            cur_km     = end_km_d

        has_official = official > 0
        comment_color = GREEN if (comment.startswith('PN ') or (has_official and comment in ('PN', 'PN+privatno'))) else colors.black
        daily_data.append([
            P(date_str, align=TA_CENTER, size=8),
            P(f"{cur_km:,.0f}", align=TA_CENTER, size=8),
            P(f"{end_km_d:,.0f}", align=TA_CENTER, size=8),
            P(f"{official:,.0f}", align=TA_CENTER, size=8, color=GREEN if has_official else colors.black),
            P(f"{private:,.0f}", align=TA_CENTER, size=8),
            P(f"{total_d:,.0f}", align=TA_CENTER, size=8),
            P(comment, size=8, color=comment_color),
        ])
        cur_km = end_km_d

    # Ukupno row
    daily_data.append([
        P('UKUPNO', bold=True, align=TA_RIGHT, size=8),
        P('', size=8), P('', size=8),
        P(f"{total_official:,.0f}", bold=True, align=TA_CENTER, size=8, color=GREEN),
        P(f"{total_private:,.0f}", bold=True, align=TA_CENTER, size=8),
        P(f"{total_official+total_private:,.0f}", bold=True, align=TA_CENTER, size=8),
        P('', size=8),
    ])

    day_cw = [22*mm, 18*mm, 18*mm, 18*mm, 18*mm, 18*mm, W-112*mm]
    day_tbl = Table(daily_data, colWidths=day_cw, repeatRows=1)
    day_styles = [
        ('BACKGROUND',(0,0),(-1,0),LBLUE),
        ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#aac4db')),
        ('TOPPADDING',(0,0),(-1,-1),2), ('BOTTOMPADDING',(0,0),(-1,-1),2),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        # Ukupno row
        ('BACKGROUND',(0,-1),(-1,-1),LBLUE),
        ('LINEABOVE',(0,-1),(-1,-1),1,BLUE),
    ]
    for i in range(1, len(daily_data)-1):
        if i % 2 == 0:
            day_styles.append(('BACKGROUND',(0,i),(-1,i),colors.HexColor('#f5f8fb')))
    day_tbl.setStyle(TableStyle(day_styles))
    story.append(day_tbl)
    story.append(Spacer(1, 8))

    # Završno stanje
    story.append(Table([
        [P(f"{year}-{month:02d}-{last_day:02d}", bold=True, align=TA_CENTER, size=8),
         P(f"{cur_km:,.0f}", bold=True, align=TA_CENTER, size=8),
         P('( Završno stanje KM — prijenos u sljedeći mjesec )', size=8, align=TA_CENTER)],
    ], colWidths=[22*mm, 18*mm, W-40*mm],
    style=TableStyle([('GRID',(0,0),(-1,-1),0.3,LBLUE),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)])))
    story.append(Spacer(1, 14))

    # ── POTPISI ────────────────────────────────────────────────────────────
    story.append(HRFlowable(width=W, thickness=0.5, color=LBLUE, spaceAfter=6))
    sig_left = [P(f"Datum: {last_day:02d}.{month:02d}.{year}.", size=9)]
    sig_mid  = [P('Potpis zaposlenika:', size=9)]
    sig_right= [P('Odobrio:', size=9)]

    if log.get('employee_signature_path'):
        sp = os.path.join(upload_dir, log['employee_signature_path'])
        if os.path.exists(sp):
            sig_mid.append(RLImage(sp, width=50*mm, height=15*mm))

    if log.get('is_approved') and director and director['signature_path']:
        dp = os.path.join(upload_dir, director['signature_path'])
        if os.path.exists(dp):
            sig_right.append(RLImage(dp, width=50*mm, height=15*mm))

    sig_tbl = Table([[sig_left, sig_mid, sig_right]], colWidths=[40*mm, 70*mm, 60*mm])
    sig_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP')]))
    story.append(sig_tbl)

    doc.build(story)
    buf.seek(0)
    fname = f"Evidencija_{month:02d}{year}.pdf"
    from flask import send_file
    return send_file(buf, mimetype='application/pdf', download_name=fname, as_attachment=False)


# ─── CAR LOGS ─────────────────────────────────────────────────────────────────

MONTHS_HR = {1:'Siječanj',2:'Veljača',3:'Ožujak',4:'Travanj',5:'Svibanj',
             6:'Lipanj',7:'Srpanj',8:'Kolovoz',9:'Rujan',10:'Listopad',
             11:'Studeni',12:'Prosinac'}

@app.route('/car-logs')
@require_perm('can_view_vehicle_log')
def car_logs_list():
    conn = get_db()
    logs = conn.execute(
        "SELECT * FROM car_logs ORDER BY year DESC, month DESC"
    ).fetchall()
    conn.close()
    return render_template('car_logs.html', logs=rows_to_dicts(logs), months=MONTHS_HR)

@app.route('/api/car-logs/parse-csv', methods=['POST'])
@require_perm('can_edit_vehicle_log')
def parse_car_csv():
    """Parse Toyota CSV and return computed monthly summary + linked PN nalozi"""
    import csv, io
    from collections import defaultdict

    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file'}), 400

    start_km_input = request.form.get('start_km', '')

    # Read CSV
    content_bytes = file.read().decode('utf-8-sig', errors='replace')
    lines = content_bytes.strip().split('\n')

    rows = []
    for line in lines[1:]:
        parts = line.strip().split(';')
        if len(parts) >= 5:
            try:
                rows.append({
                    'date': parts[1][:10],
                    'km': float(parts[4].replace(',', '.')),
                    'dep': parts[0],
                    'dest': parts[2],
                })
            except: pass

    if not rows:
        return jsonify({'error': 'Nema podataka u CSV-u'}), 400

    # Sort by date
    rows.sort(key=lambda x: x['date'])
    dates = [r['date'] for r in rows]
    year = int(dates[0][:4])
    month = int(dates[0][5:7])

    # Daily totals
    daily_km = defaultdict(float)
    for r in rows:
        daily_km[r['date']] += r['km']

    # Identify PN days (trips outside Osijek / long distance)
    pn_days = set()
    for r in rows:
        if ('Zagreb' in r['dep'] or 'Zagreb' in r['dest'] or
            ('Osijek' not in r['dep'] and 'Višnjevac' not in r['dep'] and
             'Tvrđavica' not in r['dep'] and
             'Osijek' not in r['dest'] and 'Višnjevac' not in r['dest'] and
             r['km'] > 20)):
            pn_days.add(r['date'])

    total_km = round(sum(daily_km.values()), 2)
    official_km = round(sum(v for d, v in daily_km.items() if d in pn_days), 2)
    private_km = round(total_km - official_km, 2)

    # Start km
    start_km = float(start_km_input) if start_km_input else None
    end_km = round(start_km + total_km, 2) if start_km is not None else None

    # Find matching PN nalozi by date range
    conn = get_db()
    date_from = min(dates)
    date_to = max(dates)
    pn_rows = conn.execute("""
        SELECT to2.auto_id, to2.departure_date, to2.destination,
               to2.trip_start_datetime, to2.trip_end_datetime
        FROM travel_orders to2
        WHERE (to2.is_deleted=0 OR to2.is_deleted IS NULL)
          AND (
            (to2.departure_date >= ? AND to2.departure_date <= ?) OR
            (substr(to2.trip_start_datetime,1,10) >= ? AND substr(to2.trip_start_datetime,1,10) <= ?)
          )
        ORDER BY to2.departure_date
    """, (date_from, date_to, date_from, date_to)).fetchall()
    conn.close()

    pn_list = [r['auto_id'] for r in pn_rows]

    # Daily breakdown for display
    all_dates = sorted(daily_km.keys())

    return jsonify({
        'year': year,
        'month': month,
        'month_name': MONTHS_HR.get(month, str(month)),
        'date_from': date_from,
        'date_to': date_to,
        'total_km': total_km,
        'official_km': official_km,
        'private_km': private_km,
        'start_km': start_km,
        'end_km': end_km,
        'pn_list': pn_list,
        'pn_days': list(pn_days),
        'daily': [{'date': d, 'km': round(daily_km[d],2), 'is_pn': d in pn_days} for d in all_dates],
        'row_count': len(rows),
    })


@app.route('/api/car-logs', methods=['POST'])
@require_perm('can_edit_vehicle_log')
def save_car_log():
    data = request.json
    conn = get_db()
    # Check for duplicate
    existing = conn.execute(
        "SELECT id FROM car_logs WHERE month=? AND year=?",
        (data['month'], data['year'])
    ).fetchone()
    if existing:
        conn.execute("""UPDATE car_logs SET start_km=?,end_km=?,total_km=?,
                        official_km=?,private_km=?,pn_list=?,notes=? WHERE id=?""",
                     (data.get('start_km'), data.get('end_km'), data.get('total_km'),
                      data.get('official_km'), data.get('private_km'),
                      ','.join(data.get('pn_list',[])), data.get('notes',''),
                      existing['id']))
        log_id = existing['id']
    else:
        c = conn.execute("""INSERT INTO car_logs
            (month,year,start_km,end_km,total_km,official_km,private_km,pn_list,notes)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (data['month'], data['year'], data.get('start_km'), data.get('end_km'),
             data.get('total_km'), data.get('official_km'), data.get('private_km'),
             ','.join(data.get('pn_list',[])), data.get('notes','')))
        log_id = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': log_id})


@app.route('/api/car-logs/<int:log_id>', methods=['DELETE'])
@require_perm('can_edit_vehicle_log')
def delete_car_log(log_id):
    audit('delete', module='sluzbeni_automobil', entity='car_log', entity_id=log_id)
    conn = get_db()
    conn.execute("DELETE FROM car_logs WHERE id=?", (log_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/car-logs/<int:log_id>/export')
@require_perm('can_view_vehicle_log')
def export_car_log(log_id):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import copy, io

    conn = get_db()
    log = conn.execute("SELECT * FROM car_logs WHERE id=?", (log_id,)).fetchone()
    conn.close()
    if not log:
        return "Not found", 404

    log = dict(log)
    month = log['month']
    year = log['year']

    # Use 022026 sheet as template
    try:
        template_wb = load_workbook('/mnt/user-data/uploads/Evidencija_koris_tenja_auta_2026.xlsx')
        template_ws = template_wb['022026']
    except:
        return "Template not found", 500

    import datetime
    from collections import defaultdict

    wb2 = load_workbook('/mnt/user-data/uploads/Evidencija_koris_tenja_auta_2026.xlsx')
    ws_new = wb2.copy_worksheet(wb2['022026'])
    ws_new.title = f'{month:02d}{year}'

    ws_new['H2'] = datetime.datetime(year, month, 1)
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    ws_new['C50'] = f'{last_day:02d}.{month:02d}.{year}'

    # We need daily data - stored as JSON in notes or re-parse
    # For export, build from stored totals with approximation
    # Better: store daily data in notes as JSON
    daily_json = log.get('notes', '')
    daily_data = []
    try:
        import json
        daily_data = json.loads(daily_json) if daily_json and daily_json.startswith('[') else []
    except: pass

    start_km = float(log['start_km'] or 0)
    pn_list = [p.strip() for p in (log['pn_list'] or '').split(',') if p.strip()]

    # If we have daily data, use it; otherwise build skeleton
    if daily_data:
        cumulative = start_km
        day_rows = []
        for entry in daily_data:
            d = entry['date']
            km = float(entry['km'])
            day_num = int(d[8:10])
            is_pn = entry.get('is_pn', False)
            # Find which PN
            pn_label = ''
            if is_pn:
                # Match by date to pn_list
                conn2 = get_db()
                pn_match = conn2.execute("""
                    SELECT auto_id FROM travel_orders
                    WHERE (departure_date=? OR substr(trip_start_datetime,1,10)=?)
                    AND (is_deleted=0 OR is_deleted IS NULL)
                    LIMIT 1
                """, (d, d)).fetchone()
                conn2.close()
                pn_label = pn_match['auto_id'] if pn_match else 'PN'
            day_rows.append((day_num, cumulative, cumulative+km,
                             km if is_pn else 0, km if not is_pn else 0, pn_label or ('privatno' if km > 0 else '')))
            cumulative += km
    else:
        day_rows = []

    # Write to sheet
    base_row = 18
    for row in ws_new.iter_rows(min_row=base_row, max_row=base_row+35):
        for cell in row: cell.value = None

    if day_rows:
        for i, (day_num, s, e, off, priv, comment) in enumerate(day_rows):
            r = base_row + i
            ws_new.cell(r, 2).value = f'=$H$2+{day_num-1}'
            ws_new.cell(r, 3).value = round(s, 2)
            ws_new.cell(r, 4).value = round(e, 2)
            ws_new.cell(r, 5).value = round(off, 2)
            ws_new.cell(r, 6).value = round(priv, 2)
            ws_new.cell(r, 7).value = f'=E{r}+F{r}'
            ws_new.cell(r, 8).value = comment

        total_row = base_row + len(day_rows)
        ws_new.cell(total_row, 2).value = 'Ukupno :'
        ws_new.cell(total_row, 3).value = f'=D{total_row-1}-C{base_row}'
        ws_new.cell(total_row, 4).value = ' '
        ws_new.cell(total_row, 5).value = f'=SUM(E{base_row}:E{total_row-1})'
        ws_new.cell(total_row, 6).value = f'=SUM(F{base_row}:F{total_row-1})'
        ws_new.cell(total_row, 7).value = f'=SUM(E{total_row}:F{total_row})'

        end_row = total_row + 2
        ws_new.cell(end_row, 2).value = f'=B{total_row-1}'
        ws_new.cell(end_row, 3).value = f'=D{total_row-1}'
        ws_new.cell(end_row, 4).value = '     ( Završno stanje KM  -  prijenos u slijedeći mjesec )'

    buf = io.BytesIO()
    wb2.save(buf)
    buf.seek(0)
    fname = f"Evidencija_{MONTHS_HR.get(month,str(month))}_{year}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)

# ─── PROFILES ─────────────────────────────────────────────────────────────────

PERM_LABELS = {
    'can_view_orders': 'Putni nalozi — pregled',
    'can_edit_orders': 'Putni nalozi — kreiranje/uređivanje',
    'can_delete_orders': 'Putni nalozi — brisanje',
    'can_approve_orders': 'Putni nalozi — odobravanje',
    'can_view_quotes': 'Ponude — pregled',
    'can_edit_quotes': 'Ponude — kreiranje/uređivanje',
    'can_delete_quotes': 'Ponude — brisanje',
    'can_view_reports': 'Izvještaji — pregled',
    'can_view_vehicle_log': 'Službeni automobil — pregled',
    'can_view_pool_vehicles': 'Evidencija pool automobila',
    'can_view_worktime': 'Radno vrijeme — pregled',
    'can_edit_worktime': 'Radno vrijeme — kreiranje/uređivanje',
    'can_confirm_worktime': 'Radno vrijeme — potvrđivanje',
    'can_reopen_worktime': 'Radno vrijeme — vraćanje u prethodni status',
    'can_copy_worktime': 'Radno vrijeme — kopiranje',
    'can_view_invoices': 'Ulazni računi — pregled',
    'can_edit_invoices': 'Ulazni računi — unos/uređivanje',
    'can_liquidate_invoices': 'Ulazni računi — likvidacija',
    'can_edit_invoices_liquidated': 'Ulazni računi — uređivanje nakon likvidacije',
    'can_view_loans': 'Pozajmice — pregled',
    'can_edit_loans': 'Pozajmice — kreiranje/uređivanje',
    'can_lock_loans': 'Pozajmice — zaključavanje plana otplate',
}

@app.route('/api/profiles', methods=['GET'])
@admin_required
def list_profiles():
    conn = get_db()
    profiles = conn.execute("SELECT * FROM profiles ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(p) for p in profiles])

@app.route('/api/profiles', methods=['POST'])
@admin_required
def create_profile():
    audit('create', module='postavke', entity='profile')
    data = request.json
    conn = get_db()
    perms = {k: 1 if data.get(k) else 0 for k in PERM_LABELS}
    cols = ', '.join(perms.keys())
    placeholders = ', '.join('?' for _ in perms)
    conn.execute(
        f"INSERT INTO profiles (name, description, {cols}) VALUES (?, ?, {placeholders})",
        [data.get('name',''), data.get('description','')] + list(perms.values())
    )
    conn.commit()
    p = conn.execute("SELECT * FROM profiles ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return jsonify(dict(p))

@app.route('/api/profiles/<int:profile_id>', methods=['PUT'])
@admin_required
def update_profile(profile_id):
    data = request.json
    conn = get_db()
    perms = {k: 1 if data.get(k) else 0 for k in PERM_LABELS}
    sets = 'name=?, description=?, ' + ', '.join(f"{k}=?" for k in perms)
    vals = [data.get('name',''), data.get('description','')] + list(perms.values()) + [profile_id]
    conn.execute(f"UPDATE profiles SET {sets} WHERE id=?", vals)
    conn.commit()
    p = conn.execute("SELECT * FROM profiles WHERE id=?", (profile_id,)).fetchone()
    conn.close()
    return jsonify(dict(p))

@app.route('/api/profiles/<int:profile_id>', methods=['DELETE'])
@admin_required
def delete_profile(profile_id):
    audit('delete', module='postavke', entity='profile', entity_id=profile_id)
    conn = get_db()
    conn.execute("UPDATE users SET profile_id=NULL WHERE profile_id=?", (profile_id,))
    conn.execute("DELETE FROM profiles WHERE id=?", (profile_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ─── USER MANAGEMENT ──────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    conn = get_db()
    users = conn.execute('''SELECT u.*, e.name as employee_name, p.name as profile_name
                           FROM users u
                           LEFT JOIN employees e ON u.employee_id = e.id
                           LEFT JOIN profiles p ON u.profile_id = p.id
                           ORDER BY u.username''').fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    audit('create', module='postavke', entity='user', detail=f'Novi korisnik: {request.json.get("username","") if request.is_json else ""}')
    data = request.json
    conn = get_db()
    try:
        conn.execute('''INSERT INTO users (username, password_hash, is_admin, is_active,
                        employee_id, display_name, email, auth_provider, profile_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 'local', ?)''',
                    (data['username'],
                     generate_password_hash(data.get('password', 'changeme123'), method='pbkdf2:sha256'),
                     1 if data.get('is_admin') else 0,
                     1 if data.get('is_active', True) else 0,
                     data.get('employee_id') or None,
                     data.get('display_name', data['username']),
                     data.get('email', ''),
                     data.get('profile_id') or None))
        conn.commit()
        user = conn.execute("SELECT * FROM users WHERE username=?", (data['username'],)).fetchone()
        conn.close()
        return jsonify(dict(user))
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(user_id):
    data = request.json
    conn = get_db()
    fields = {
        'is_admin': 1 if data.get('is_admin') else 0,
        'is_active': 1 if data.get('is_active', True) else 0,
        'display_name': data.get('display_name', ''),
        'email': data.get('email', ''),
        'employee_id': data.get('employee_id') or None,
        'profile_id': data.get('profile_id') or None,
    }
    if data.get('password'):
        fields['password_hash'] = generate_password_hash(data['password'], method='pbkdf2:sha256')
    sets = ', '.join(f"{k}=?" for k in fields)
    conn.execute(f"UPDATE users SET {sets} WHERE id=?", list(fields.values()) + [user_id])
    conn.commit()
    user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    conn.close()
    return jsonify(dict(user))

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    audit('delete', module='postavke', entity='user', entity_id=user_id)
    current = get_current_user()
    if current and current['user_id'] == user_id:
        return jsonify({'error': 'Ne možete obrisati vlastiti račun'}), 400
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/users/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.json
    current = get_current_user()
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (current['user_id'],)).fetchone()
    if not check_password_hash(user['password_hash'], data.get('current_password', '')):
        conn.close()
        return jsonify({'error': 'Trenutna lozinka nije ispravna'}), 400
    conn.execute("UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?",
                (generate_password_hash(data['new_password'], method='pbkdf2:sha256'), current['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_password(user_id):
    data = request.json
    temp_password = data.get('temp_password', '').strip()
    if not temp_password or len(temp_password) < 4:
        return jsonify({'error': 'Lozinka mora imati najmanje 4 znaka'}), 400
    conn = get_db()
    conn.execute("UPDATE users SET password_hash=?, must_change_password=1 WHERE id=?",
                (generate_password_hash(temp_password, method='pbkdf2:sha256'), user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/settings')
@admin_required
def settings_page():
    audit('view', module='postavke', entity='list')
    conn = get_db()
    settings = {row['key']: row['value'] for row in conn.execute("SELECT * FROM settings").fetchall()}
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    vehicles = conn.execute("SELECT * FROM vehicles ORDER BY name").fetchall()
    clients = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
    destinations = conn.execute("SELECT * FROM destinations ORDER BY name").fetchall()
    categories = conn.execute("SELECT * FROM expense_categories ORDER BY name").fetchall()
    templates = conn.execute("SELECT * FROM report_templates ORDER BY name").fetchall()
    conn.close()
    return render_template('settings.html', settings=settings,
                           employees=rows_to_dicts(employees),
                           vehicles=rows_to_dicts(vehicles),
                           clients=rows_to_dicts(clients),
                           destinations=rows_to_dicts(destinations),
                           categories=rows_to_dicts(categories),
                           templates=rows_to_dicts(templates))

@app.route('/api/settings', methods=['POST'])
@admin_required
def save_settings():
    audit('edit', module='postavke', entity='settings', detail='Opće postavke ažurirane')
    data = request.json
    conn = get_db()
    for key, value in data.items():
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()
    _invalidate_settings_cache()  # Osvježi cache nakon promjene postavki
    return jsonify({'success': True})

# Generic codebook CRUD
def make_codebook_routes(table, required_field='name'):
    def list_items():
        conn = get_db()
        items = conn.execute(f"SELECT * FROM {table} ORDER BY {required_field}").fetchall()
        conn.close()
        return jsonify([dict(i) for i in items])

    def create_item():
        data = request.json
        conn = get_db()
        # Enforce single default: if new item is default, clear others
        if data.get('is_default'):
            conn.execute(f"UPDATE {table} SET is_default=0")
        cols = ', '.join(data.keys())
        placeholders = ', '.join('?' for _ in data)
        c = conn.execute(f"INSERT INTO {table} ({cols}) VALUES ({placeholders})", list(data.values()))
        conn.commit()
        item = conn.execute(f"SELECT * FROM {table} WHERE id=?", (c.lastrowid,)).fetchone()
        conn.close()
        return jsonify(dict(item))

    def update_item(item_id):
        data = request.json
        conn = get_db()
        # Enforce single default: if updating to default, clear others first
        if data.get('is_default'):
            conn.execute(f"UPDATE {table} SET is_default=0 WHERE id!=?", (item_id,))
        sets = ', '.join(f"{k}=?" for k in data)
        conn.execute(f"UPDATE {table} SET {sets} WHERE id=?", list(data.values()) + [item_id])
        conn.commit()
        item = conn.execute(f"SELECT * FROM {table} WHERE id=?", (item_id,)).fetchone()
        conn.close()
        return jsonify(dict(item))

    def delete_item(item_id):
        conn = get_db()
        conn.execute(f"DELETE FROM {table} WHERE id=?", (item_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})

    list_items.__name__ = f'list_{table}'
    create_item.__name__ = f'create_{table}'
    update_item.__name__ = f'update_{table}'
    delete_item.__name__ = f'delete_{table}'

    app.route(f'/api/{table}', methods=['GET'])(list_items)
    app.route(f'/api/{table}', methods=['POST'])(create_item)
    app.route(f'/api/{table}/<int:item_id>', methods=['PUT'])(update_item)
    app.route(f'/api/{table}/<int:item_id>', methods=['DELETE'])(delete_item)

for t in ['vehicles', 'destinations', 'expense_categories', 'report_templates']:  # clients ima vlastite rute
    make_codebook_routes(t)

@app.route('/api/employees', methods=['GET'])
@admin_required
def list_employees():
    conn = get_db()
    items = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

@app.route('/api/employees', methods=['POST'])
@admin_required
def create_employee():
    data = request.json
    conn = get_db()
    if data.get('is_default'):
        conn.execute("UPDATE employees SET is_default=0")
    cols = ', '.join(k for k in data if k != 'signature_file')
    placeholders = ', '.join('?' for k in data if k != 'signature_file')
    vals = [v for k, v in data.items() if k != 'signature_file']
    c = conn.execute(f"INSERT INTO employees ({cols}) VALUES ({placeholders})", vals)
    conn.commit()
    item = conn.execute("SELECT * FROM employees WHERE id=?", (c.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(item))

@app.route('/api/employees/<int:item_id>', methods=['PUT'])
@admin_required
def update_employee(item_id):
    data = request.json
    conn = get_db()
    if data.get('is_default'):
        conn.execute("UPDATE employees SET is_default=0 WHERE id!=?", (item_id,))
    sets = ', '.join(f"{k}=?" for k in data)
    conn.execute(f"UPDATE employees SET {sets} WHERE id=?", list(data.values()) + [item_id])
    conn.commit()
    item = conn.execute("SELECT * FROM employees WHERE id=?", (item_id,)).fetchone()
    conn.close()
    return jsonify(dict(item))

@app.route('/api/employees/<int:item_id>', methods=['DELETE'])
@admin_required
def delete_employee(item_id):
    conn = get_db()
    conn.execute("DELETE FROM employees WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/employees/<int:item_id>/signature', methods=['POST'])
@admin_required
def upload_signature(item_id):
    if 'signature' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['signature']
    filename = f"sig_{item_id}.png"
    path = os.path.join(UPLOAD_FOLDER, filename)
    img = PILImage.open(f)
    img.save(path, 'PNG')
    conn = get_db()
    conn.execute("UPDATE employees SET signature_path=? WHERE id=?", (filename, item_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'path': filename})

# ─── WORKTIME MODULE ────────────────────────────────────────────────────────

# Croatian public holidays 2026
HOLIDAYS_2026 = {
    '2026-01-01': 'Nova godina',
    '2026-01-06': 'Bogojavljenje / Sveta tri kralja',
    '2026-04-05': 'Uskrs',
    '2026-04-06': 'Uskrsni ponedjeljak',
    '2026-05-01': 'Praznik rada',
    '2026-05-30': 'Dan državnosti',
    '2026-06-04': 'Tijelovo',
    '2026-06-22': 'Dan antifašističke borbe',
    '2026-08-05': 'Dan pobjede i domovinske zahvalnosti',
    '2026-08-15': 'Velika Gospa',
    '2026-11-01': 'Dan svih svetih',
    '2026-11-18': 'Dan sjećanja na žrtve Domovinskog rata',
    '2026-12-25': 'Božić',
    '2026-12-26': 'Sveti Stjepan',
}

# Croatian public holidays 2027
HOLIDAYS_2027 = {
    '2027-01-01': 'Nova godina',
    '2027-01-06': 'Bogojavljenje / Sveta tri kralja',
    '2027-03-28': 'Uskrs',
    '2027-03-29': 'Uskrsni ponedjeljak',
    '2027-05-01': 'Praznik rada',
    '2027-05-30': 'Dan državnosti',
    '2027-05-27': 'Tijelovo',
    '2027-06-22': 'Dan antifašističke borbe',
    '2027-08-05': 'Dan pobjede i domovinske zahvalnosti i Dan hrvatskih branitelja',
    '2027-08-15': 'Velika Gospa',
    '2027-11-01': 'Svi sveti',
    '2027-11-18': 'Dan sjećanja na žrtve Domovinskog rata',
    '2027-12-25': 'Božić',
    '2027-12-26': 'Sveti Stjepan',
}

# Kombinirani rječnik blagdana
HOLIDAYS_2028 = {
    '2028-01-01': 'Nova godina',
    '2028-01-06': 'Bogojavljenje / Sveta tri kralja',
    '2028-04-16': 'Uskrs',
    '2028-04-17': 'Uskrsni ponedjeljak',
    '2028-05-01': 'Praznik rada',
    '2028-05-30': 'Dan državnosti',
    '2028-06-15': 'Tijelovo',
    '2028-06-22': 'Dan antifašističke borbe',
    '2028-08-05': 'Dan pobjede i domovinske zahvalnosti i Dan hrvatskih branitelja',
    '2028-08-15': 'Velika Gospa',
    '2028-11-01': 'Svi sveti',
    '2028-11-18': 'Dan sjećanja na žrtve Domovinskog rata',
    '2028-12-25': 'Božić',
    '2028-12-26': 'Sveti Stjepan',
}

ALL_HOLIDAYS = {**HOLIDAYS_2026, **HOLIDAYS_2027, **HOLIDAYS_2028}

# Fund of hours 2026 per month (from official HR labour law data)
WORK_FUND_2026 = {
    1:  {'fond': 176, 'radni': 160, 'neradni': 16, 'obracunskih': 22, 'radnih_dana': 20, 'neradnih_dana': 2},
    2:  {'fond': 160, 'radni': 160, 'neradni': 0,  'obracunskih': 20, 'radnih_dana': 20, 'neradnih_dana': 0},
    3:  {'fond': 176, 'radni': 176, 'neradni': 0,  'obracunskih': 22, 'radnih_dana': 22, 'neradnih_dana': 0},
    4:  {'fond': 176, 'radni': 168, 'neradni': 8,  'obracunskih': 22, 'radnih_dana': 21, 'neradnih_dana': 1},
    5:  {'fond': 168, 'radni': 160, 'neradni': 8,  'obracunskih': 21, 'radnih_dana': 20, 'neradnih_dana': 1},
    6:  {'fond': 176, 'radni': 160, 'neradni': 16, 'obracunskih': 22, 'radnih_dana': 20, 'neradnih_dana': 2},
    7:  {'fond': 184, 'radni': 184, 'neradni': 0,  'obracunskih': 23, 'radnih_dana': 23, 'neradnih_dana': 0},
    8:  {'fond': 168, 'radni': 160, 'neradni': 8,  'obracunskih': 21, 'radnih_dana': 20, 'neradnih_dana': 1},
    9:  {'fond': 176, 'radni': 176, 'neradni': 0,  'obracunskih': 22, 'radnih_dana': 22, 'neradnih_dana': 0},
    10: {'fond': 176, 'radni': 176, 'neradni': 0,  'obracunskih': 22, 'radnih_dana': 22, 'neradnih_dana': 0},
    11: {'fond': 168, 'radni': 160, 'neradni': 8,  'obracunskih': 21, 'radnih_dana': 20, 'neradnih_dana': 1},
    12: {'fond': 184, 'radni': 176, 'neradni': 8,  'obracunskih': 23, 'radnih_dana': 22, 'neradnih_dana': 1},
}

WORKTIME_ROWS = [
    (1,  'Početak rada',                                              False),
    (2,  'Završetak rada',                                            False),
    (3,  'Sati zastoja bez krivice radnika',                          True),
    (4,  'UKUPNO DNEVNO RADNO VRIJEME',                              False),
    (5,  'Redovni sati rada danju prema rasporedu',                   True),
    (6,  'Prekovremeni sati rada danju',                              True),
    (7,  'Sati rada danju u dane blagdana ili neradne dane',          True),
    (8,  'Redovni sati rada noću prema rasporedu',                    True),
    (9,  'Prekovremeni sati rada noću',                               True),
    (10, 'Sati rada noću u dane blagdana ili neradne dane',           True),
    (11, 'Plaćeni sati praznika i blagdana',                          True),
    (12, 'Vrijeme i sati zastoja za koje radnik nije kriv',           True),
    (13, 'Sati izostajanja s posla u toku dana bez odobrenja',        True),
    (14, 'Sati provedeni u štrajku',                                  True),
    (15, 'Sati isključenja s rada (lockout)',                          True),
    (16, 'Sati korištenja godišnjeg odmora',                          True),
    (17, 'Sati privr.nesposobnosti - bolovanja',                      True),
    (18, 'Rodiljni, roditeljski dopust ili korištenje drugih prava',  True),
    (19, 'Sati plaćenog dopusta',                                     True),
    (20, 'Preraspodijeljeno radno vrijeme',                           True),
    (21, 'Sati pripravnosti, te sati rada po pozivu',                 True),
    (22, 'Odbio raditi zbog neprovedenih mjera zaštite',              True),
]

MONTHS_HR = {1:'Siječanj',2:'Veljača',3:'Ožujak',4:'Travanj',5:'Svibanj',
             6:'Lipanj',7:'Srpanj',8:'Kolovoz',9:'Rujan',10:'Listopad',
             11:'Studeni',12:'Prosinac'}

def get_day_type(year, month, day):
    import calendar
    date_str = f"{year}-{month:02d}-{day:02d}"
    weekday = calendar.weekday(year, month, day)
    if date_str in ALL_HOLIDAYS:
        return 'holiday', ALL_HOLIDAYS[date_str]
    if weekday == 6:
        return 'sunday', ''
    if weekday == 5:
        return 'saturday', ''
    return 'workday', ''


def get_holidays_by_year():
    """Grupiraj ALL_HOLIDAYS po godini, dinamički."""
    import json as _json
    by_year = {}
    for date_str, name in ALL_HOLIDAYS.items():
        year = int(date_str[:4])
        if year not in by_year:
            by_year[year] = []
        # Formatiraj datum kao DD.MM.YYYY. i dan u tjednu
        from datetime import date as _date
        d = _date(int(date_str[:4]), int(date_str[5:7]), int(date_str[8:10]))
        DAYS_HR = {0:'Pon',1:'Uto',2:'Sri',3:'Čet',4:'Pet',5:'Sub',6:'Ned'}
        day_hr = DAYS_HR[d.weekday()]
        formatted = f"{date_str[8:10]}.{date_str[5:7]}.{date_str[:4]}."
        by_year[year].append([formatted, day_hr, name, date_str])
    # Sortiraj po datumu unutar svake godine
    for year in by_year:
        by_year[year].sort(key=lambda x: x[3])
        # Makni interni date_str
        by_year[year] = [row[:3] for row in by_year[year]]
    return by_year

@app.route('/worktime')
@require_perm('can_view_worktime')
def worktime_list():
    audit('view', module='radno_vrijeme', entity='list')
    current_user = get_current_user() or {}
    conn = get_db()
    reports = conn.execute('''
        SELECT wr.*, e.name as employee_name
        FROM worktime_reports wr
        LEFT JOIN employees e ON wr.employee_id = e.id
        ORDER BY wr.year DESC, wr.month DESC, e.name
    ''').fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    import json as _json
    holidays_by_year = get_holidays_by_year()
    available_years = sorted(holidays_by_year.keys())
    conn2 = get_db()
    fund_years = [r['year'] for r in conn2.execute("SELECT DISTINCT year FROM work_fund ORDER BY year").fetchall()]
    conn2.close()
    all_years = sorted(set(available_years) | set(fund_years))
    return render_template('worktime_list.html',
                           reports=rows_to_dicts(reports),
                           employees=rows_to_dicts(employees),
                           months=MONTHS_HR,
                           active='worktime',
                           settings=settings,
                           holidays_json=_json.dumps(holidays_by_year),
                           available_years=all_years,
                           is_admin=current_user.get('is_admin', False))


@app.route('/worktime/new')
@require_perm('can_edit_worktime')
def worktime_new():
    conn = get_db()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    return render_template('worktime_form.html',
                           report=None, entries={}, employees=rows_to_dicts(employees),
                           rows=WORKTIME_ROWS, months=MONTHS_HR,
                           holidays=ALL_HOLIDAYS,
                           holidays_by_year={2026: HOLIDAYS_2026, 2027: HOLIDAYS_2027, 2028: HOLIDAYS_2028},
                           work_fund=WORK_FUND_2026,
                           active='worktime', settings=settings)


@app.route('/worktime/<int:report_id>/edit')
@require_perm('can_view_worktime')
def worktime_edit(report_id):
    import calendar
    conn = get_db()
    report = conn.execute("SELECT wr.*, e.name as employee_name FROM worktime_reports wr LEFT JOIN employees e ON wr.employee_id = e.id WHERE wr.id=?", (report_id,)).fetchone()
    if not report:
        return "Not found", 404
    raw_entries = conn.execute("SELECT * FROM worktime_entries WHERE report_id=?", (report_id,)).fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    # Build entries dict: {row_num: {day: hours}}
    entries = {}
    for e in raw_entries:
        if e['row_num'] not in entries:
            entries[e['row_num']] = {}
        entries[e['row_num']][e['day']] = e['hours']
    return render_template('worktime_form.html',
                           report=row_to_dict(report), entries=entries,
                           employees=rows_to_dicts(employees),
                           rows=WORKTIME_ROWS, months=MONTHS_HR,
                           holidays=ALL_HOLIDAYS, work_fund=WORK_FUND_2026,
                           active='worktime', settings=settings)


@app.route('/api/worktime', methods=['POST'])
@require_perm('can_edit_worktime')
def save_worktime():
    _wt_data = request.json or {}
    _wt_id = _wt_data.get('id')
    _wt_action = 'edit' if _wt_id else 'create'
    audit(_wt_action, module='radno_vrijeme', entity='worktime_report',
          entity_id=int(_wt_id) if _wt_id else None,
          detail=f'Status: {_wt_data.get("status","")}')
    import calendar
    data = request.json
    report_id = data.get('id')
    employee_id = data.get('employee_id')
    year = int(data.get('year', 2026))
    month = int(data.get('month', 1))
    status = data.get('status', 'draft')
    notes = data.get('notes', '')
    entries = data.get('entries', {})  # {row_num: {day: hours}}

    # Check confirm permission
    if status == 'confirmed':
        user = get_current_user()
        if not user.get('is_admin') and not user.get('can_confirm_worktime'):
            return jsonify({'error': 'Nemate pravo potvrđivanja izvješća o radnom vremenu'}), 403

    conn = get_db()
    c = conn.cursor()
    now = datetime.now().isoformat()

    if report_id:
        existing = conn.execute("SELECT status FROM worktime_reports WHERE id=?", (report_id,)).fetchone()
        if existing and existing['status'] == 'confirmed':
            user = get_current_user()
            if not user.get('is_admin') and not user.get('can_reopen_worktime'):
                conn.close()
                return jsonify({'error': 'Nemate pravo vraćanja potvrđenog izvješća u prethodni status'}), 403
            # Allow reopen only to submitted or draft
            if status not in ('submitted', 'draft'):
                conn.close()
                return jsonify({'error': 'Potvrđeno izvješće može se vratiti samo u status Predano ili Nacrt'}), 400
        # submitted can only be changed to confirmed or back to draft
        if existing and existing['status'] == 'submitted' and status not in ('confirmed', 'draft'):
            conn.close()
            return jsonify({'error': 'Predano izvješće može se samo potvrditi ili vratiti u nacrt'}), 400
        confirmed_at = now if status == 'confirmed' else None
        confirmed_by = get_current_user()['user_id'] if status == 'confirmed' else None
        submitted_at_val = now if status == 'submitted' else (existing['submitted_at'] if existing and 'submitted_at' in existing.keys() else None)
        submitted_by_val = get_current_user()['user_id'] if status == 'submitted' else (existing['submitted_by'] if existing and 'submitted_by' in existing.keys() else None)
        # Clear submitted info if returning to draft
        if status == 'draft':
            submitted_at_val = None
            submitted_by_val = None
        c.execute("""UPDATE worktime_reports SET employee_id=?, year=?, month=?, status=?,
                     notes=?, updated_at=?, submitted_at=?, submitted_by=?, confirmed_at=?, confirmed_by=? WHERE id=?""",
                  (employee_id, year, month, status, notes, now, submitted_at_val, submitted_by_val, confirmed_at, confirmed_by, report_id))
    else:
        # Check for duplicate before INSERT
        existing_check = conn.execute(
            "SELECT wr.id, e.name as emp_name FROM worktime_reports wr LEFT JOIN employees e ON wr.employee_id = e.id WHERE wr.employee_id=? AND wr.year=? AND wr.month=?",
            (employee_id, year, month)
        ).fetchone()
        if existing_check:
            conn.close()
            month_name = MONTHS_HR.get(month, str(month))
            emp_name = existing_check['emp_name'] or 'odabranog zaposlenika'
            return jsonify({'error': f'Evidencija radnog vremena za {emp_name} za {month_name} {year}. već postoji (ID: {existing_check["id"]}).'}), 409
        submitted_at_new = now if status == 'submitted' else None
        submitted_by_new = get_current_user()['user_id'] if status == 'submitted' else None
        confirmed_at = now if status == 'confirmed' else None
        confirmed_by = get_current_user()['user_id'] if status == 'confirmed' else None
        c.execute("""INSERT INTO worktime_reports (employee_id, year, month, status, notes, created_at, updated_at, submitted_at, submitted_by, confirmed_at, confirmed_by)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                  (employee_id, year, month, status, notes, now, now, submitted_at_new, submitted_by_new, confirmed_at, confirmed_by))
        report_id = c.lastrowid

    # Save entries
    c.execute("DELETE FROM worktime_entries WHERE report_id=?", (report_id,))
    for row_num_str, days in entries.items():
        row_num = int(row_num_str)
        for day_str, hours in days.items():
            if hours and float(hours) != 0:
                c.execute("INSERT INTO worktime_entries (report_id, day, row_num, hours) VALUES (?,?,?,?)",
                          (report_id, int(day_str), row_num, float(hours)))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': report_id})


@app.route('/api/worktime/<int:report_id>', methods=['DELETE'])
@require_perm('can_edit_worktime')
def delete_worktime(report_id):
    audit('delete', module='radno_vrijeme', entity='worktime_report', entity_id=report_id)
    conn = get_db()
    conn.execute("DELETE FROM worktime_reports WHERE id=?", (report_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/worktime/<int:report_id>/copy')
@require_perm('can_copy_worktime')
def copy_worktime(report_id):
    audit('copy', module='radno_vrijeme', entity='worktime_report', entity_id=report_id)
    conn = get_db()
    report = conn.execute("SELECT * FROM worktime_reports WHERE id=?", (report_id,)).fetchone()
    entries = conn.execute("SELECT * FROM worktime_entries WHERE report_id=?", (report_id,)).fetchall()
    employees = conn.execute("SELECT * FROM employees ORDER BY name").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    entry_dict = {}
    for e in entries:
        if e['row_num'] not in entry_dict:
            entry_dict[e['row_num']] = {}
        entry_dict[e['row_num']][e['day']] = e['hours']
    return render_template('worktime_form.html',
                           report=None, entries=entry_dict,
                           copy_from=row_to_dict(report),
                           employees=rows_to_dicts(employees),
                           rows=WORKTIME_ROWS, months=MONTHS_HR,
                           holidays=ALL_HOLIDAYS, work_fund=WORK_FUND_2026,
                           active='worktime', settings=settings)


@app.route('/worktime/<int:report_id>/pdf')
@require_perm('can_view_worktime')
def worktime_pdf(report_id):
    audit('export_pdf', module='radno_vrijeme', entity='worktime_report', entity_id=report_id)
    import calendar
    conn = get_db()
    report = conn.execute("SELECT wr.*, e.name as employee_name FROM worktime_reports wr LEFT JOIN employees e ON wr.employee_id = e.id WHERE wr.id=?", (report_id,)).fetchone()
    if not report:
        return "Not found", 404
    entries_raw = conn.execute("SELECT * FROM worktime_entries WHERE report_id=?", (report_id,)).fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    # Get employee signature
    employee_data = conn.execute("SELECT * FROM employees WHERE id=?", (report['employee_id'],)).fetchone()
    # Get director signature
    director_data = conn.execute("SELECT * FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    conn.close()

    entries = {}
    for e in entries_raw:
        if e['row_num'] not in entries:
            entries[e['row_num']] = {}
        entries[e['row_num']][e['day']] = e['hours']

    report = row_to_dict(report)
    report_dict = row_to_dict(report) if not isinstance(report, dict) else report
    report_status = report_dict.get('status', '')
    # Employee signature: show when submitted or confirmed
    if report_status in ('submitted', 'confirmed') and employee_data:
        report_dict['employee_signature'] = row_to_dict(employee_data).get('signature_path')
    else:
        report_dict['employee_signature'] = None
    # Director signature: show only when confirmed
    if report_status == 'confirmed' and director_data:
        report_dict['director_signature'] = row_to_dict(director_data).get('signature_path')
    else:
        report_dict['director_signature'] = None
    report_dict['director_name'] = row_to_dict(director_data).get('name', '') if director_data else ''
    report['director_name'] = row_to_dict(director_data).get('name', '') if director_data else ''
    year, month = report['year'], report['month']
    days_in_month = calendar.monthrange(year, month)[1]

    buf = create_worktime_pdf(report, entries, days_in_month, settings)
    filename = f"RadnoVrijeme_{report['employee_name'].replace(' ','_')}_{year}_{month:02d}.pdf"
    return send_file(buf, mimetype='application/pdf', download_name=filename, as_attachment=False)


def create_worktime_pdf(report, entries, days_in_month, settings):
    import io, calendar
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=6*mm, rightMargin=6*mm,
                            topMargin=6*mm, bottomMargin=6*mm)

    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        pdfmetrics.registerFont(TTFont('Arial', '/Library/Fonts/Arial Unicode.ttf'))
        font = 'Arial'
    except:
        font = 'Helvetica'

    styles = getSampleStyleSheet()
    small = ParagraphStyle('small', fontName=font, fontSize=7, leading=8, alignment=TA_CENTER)
    day_style = ParagraphStyle('day', fontName=font, fontSize=5.5, leading=6, alignment=TA_CENTER)
    tiny = ParagraphStyle('tiny', fontName=font, fontSize=6.5, leading=7.5, alignment=TA_LEFT)
    bold_small = ParagraphStyle('bold_small', fontName=font, fontSize=7.5, leading=8.5, alignment=TA_CENTER, fontWeight='bold')

    year, month = report['year'], report['month']
    month_name = MONTHS_HR.get(month, str(month)).upper()
    company = settings.get('company_name', '')
    employee = report.get('employee_name', '')
    director_name = report.get('director_name', '')
    employee_sig_path = report.get('employee_signature')
    director_sig_path = report.get('director_signature')

    story = []

    # Title — larger and bold
    title_style = ParagraphStyle('title', fontName=font, fontSize=14, leading=17, alignment=TA_CENTER, fontWeight='bold')
    sub_style = ParagraphStyle('sub', fontName=font, fontSize=9, leading=11, alignment=TA_CENTER)
    story.append(Paragraph("EVIDENCIJA O RADNOM VREMENU", title_style))
    story.append(Paragraph(f"MJESEC: {month_name} {year}  |  Djelatnik: {employee}  |  {company}", sub_style))
    story.append(Spacer(1, 2*mm))

    # Build table
    # Columns: R.br. | Opis | day1..dayN | Ostvareno
    header_row1 = [Paragraph('R.br.', small), Paragraph('O P I S', small)]
    header_row2 = ['', '']
    header_row3 = ['', '']

    col_widths = [8*mm, 52*mm]

    for d in range(1, days_in_month + 1):
        date_str = f"{year}-{month:02d}-{d:02d}"
        wd = calendar.weekday(year, month, d)
        day_names = ['PON','UTO','SRI','ČET','PET','SUB','NED']
        is_holiday = date_str in ALL_HOLIDAYS
        is_weekend = wd >= 5
        header_row1.append(Paragraph(str(d), small))
        header_row2.append(Paragraph(day_names[wd], day_style))
        header_row3.append('')
        col_widths.append(6*mm)

    header_row1.extend([Paragraph('OSTVARENO', small), Paragraph('PLAN', small)])
    header_row2.extend(['', ''])
    header_row3.extend(['', ''])
    col_widths.extend([14*mm, 14*mm])

    data = [header_row1, header_row2]

    # Data rows
    for row_num, desc, is_numeric in WORKTIME_ROWS:
        row_data = [
            Paragraph(str(row_num), small),
            Paragraph(desc, tiny),
        ]
        row_total = 0
        for d in range(1, days_in_month + 1):
            val = entries.get(row_num, {}).get(d, '')
            if val and float(val) != 0:
                row_data.append(Paragraph(str(int(val) if float(val) == int(float(val)) else val), small))
                row_total += float(val)
            else:
                row_data.append('')
        # Ostvareno
        row_data.append(Paragraph(str(int(row_total) if row_total == int(row_total) else row_total), small) if row_total else '')
        row_data.append('')  # Plan
        data.append(row_data)

    # Grand total row
    grand_total = sum(
        sum(float(v) for v in day_vals.values())
        for day_vals in entries.values()
    )
    total_row = [Paragraph('', small), Paragraph('UKUPNO SATI:', bold_small)]
    for d in range(1, days_in_month + 1):
        day_total = sum(entries.get(rn, {}).get(d, 0) for rn, _, is_num in WORKTIME_ROWS if is_num)
        total_row.append(Paragraph(str(int(day_total)) if day_total else '', bold_small))
    total_row.append(Paragraph(str(int(grand_total) if grand_total == int(grand_total) else grand_total), bold_small))
    total_row.append('')
    data.append(total_row)

    t = Table(data, colWidths=col_widths, repeatRows=2)

    # Styling
    style_cmds = [
        ('FONTNAME', (0,0), (-1,-1), font),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,1), colors.HexColor('#4A7AAC')),
        ('TEXTCOLOR', (0,0), (-1,1), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,2), (-1,-3), [colors.white, colors.HexColor('#EEF4FA')]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#4A7AAC')),
        ('TEXTCOLOR', (0,-1), (-1,-1), colors.white),
        ('FONTSIZE', (0,-1), (-1,-1), 7.5),
        ('FONTNAME', (0,-1), (-1,-1), font),
        ('ROWHEIGHT', (0,0), (-1,0), 12),
        ('ROWHEIGHT', (0,1), (-1,1), 9),
        ('FONTSIZE', (0,2), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]

    # Highlight weekends and holidays in header
    for d in range(1, days_in_month + 1):
        col = d + 1
        date_str = f"{year}-{month:02d}-{d:02d}"
        wd = calendar.weekday(year, month, d)
        if date_str in ALL_HOLIDAYS:
            style_cmds.append(('BACKGROUND', (col,0), (col,1), colors.HexColor('#E74C3C')))
            style_cmds.append(('BACKGROUND', (col,2), (col,-2), colors.HexColor('#FDECEA')))
        elif wd >= 5:
            style_cmds.append(('BACKGROUND', (col,0), (col,1), colors.HexColor('#5B8DB8')))
            style_cmds.append(('BACKGROUND', (col,2), (col,-2), colors.HexColor('#E8F0F7')))

    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    # Signature block below table
    story.append(Spacer(1, 4*mm))
    from reportlab.platypus import Image as RLImage

    sig_style = ParagraphStyle('sig', fontName=font, fontSize=8, leading=10, alignment=TA_CENTER)
    sig_label = ParagraphStyle('sig_label', fontName=font, fontSize=7, leading=9, alignment=TA_CENTER, textColor=colors.HexColor('#666666'))
    sig_name = ParagraphStyle('sig_name', fontName=font, fontSize=9, leading=11, alignment=TA_CENTER, fontWeight='bold')

    def make_sig_cell(sig_path, label, name):
        """Build signature cell — image if available, else blank line."""
        items = []
        if sig_path:
            full_path = os.path.join(UPLOAD_FOLDER, sig_path)
            if os.path.exists(full_path):
                try:
                    img = RLImage(full_path, width=35*mm, height=12*mm)
                    img.hAlign = 'CENTER'
                    items.append(img)
                except:
                    items.append(Paragraph('_' * 30, sig_style))
            else:
                items.append(Paragraph('_' * 30, sig_style))
        else:
            items.append(Paragraph('_' * 30, sig_style))
        items.append(Paragraph(label, sig_label))
        items.append(Paragraph(name, sig_name))
        return items

    emp_cell = make_sig_cell(employee_sig_path, 'Djelatnik', employee)
    dir_cell = make_sig_cell(director_sig_path, 'Odgovorna osoba', director_name or company)

    from reportlab.platypus import KeepInFrame
    sig_data = [[emp_cell, dir_cell]]
    sig_table = Table(sig_data, colWidths=[doc.width/2, doc.width/2])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEBEFORE', (1,0), (1,-1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    story.append(sig_table)

    doc.build(story)
    buf.seek(0)
    return buf



# ─── ULAZNI RAČUNI MODULE ────────────────────────────────────────────────────

INVOICE_STORAGE_BASE = os.environ.get("INVOICE_STORAGE_PATH", os.path.join(os.path.dirname(__file__), "uploads", "racuni"))

def get_invoice_folder(year, month):
    """Returns folder path like /path/02 2026, creates it if missing."""
    import os
    folder = os.path.join(INVOICE_STORAGE_BASE, f"{month:02d} {year}")
    os.makedirs(folder, exist_ok=True)
    return folder

def ocr_file(file_path, mime_type):
    """Run OCR on PDF or image, return raw text."""
    import os
    try:
        import pytesseract
        from PIL import Image

        # Set tesseract path explicitly for macOS Homebrew
        pytesseract.pytesseract.tesseract_cmd = os.environ.get('TESSERACT_PATH', '/opt/homebrew/bin/tesseract')

        if mime_type == 'application/pdf' or file_path.lower().endswith('.pdf'):
            # First try direct text extraction (digital PDF)
            try:
                import fitz
                doc = fitz.open(file_path)
                direct_text = ""
                for page in doc:
                    direct_text += page.get_text()
                doc.close()
                if len(direct_text.strip()) > 50:
                    return direct_text  # Digital PDF — use directly
            except:
                pass

            # Scanned PDF — convert to images and OCR
            from pdf2image import convert_from_path
            pages = convert_from_path(file_path, dpi=300,
                                      poppler_path=os.environ.get('POPPLER_PATH', '/opt/homebrew/bin'))
            text = ""
            for page_img in pages:
                text += pytesseract.image_to_string(page_img, lang='hrv+eng') + "\n"
            return text
        else:
            img = Image.open(file_path)
            text = pytesseract.image_to_string(img, lang='hrv+eng')
            return text
    except Exception as e:
        return f"OCR_ERROR: {e}"

def parse_invoice_data(text):
    """Extract invoice fields from OCR text using regex."""
    import re

    result = {
        'invoice_number': '',
        'partner_name': '',
        'partner_oib': '',
        'invoice_date': '',
        'due_date': '',
        'amount_total': '',
    }

    lines = text.split('\n')
    text_clean = re.sub(r'[ \t]+', ' ', text)
    # Fix common OCR letter substitutions
    text_clean = re.sub(r'\bDIB\b', 'OIB', text_clean)
    text_clean = re.sub(r'\bFUR\b', 'EUR', text_clean)

    # ── Invoice number ────────────────────────────────────────────
    for pat in [
        r'(?:FISK[A-Z]*\s+)?[Rr]a[čc]un\s*br[.:]?\s*([\d]{4,}(?:[\s\-/][\d]+)*)',
        r'[Rr]a[čc]un\s*br[.:]?\s*([A-Z0-9][\w\-/\s]{2,20})',
        r'[Bb]roj\s+ra[čc]una[:\s]*([A-Z0-9][\-/A-Z0-9]+)',
    ]:
        m = re.search(pat, text_clean, re.IGNORECASE)
        if m:
            num = re.sub(r'\s+', '', m.group(1)).strip().rstrip('.')
            if len(num) >= 3:
                result['invoice_number'] = num
                break

    # ── OIB ───────────────────────────────────────────────────────
    # Look for OIB/DIB label followed by digits (allow ? and spaces as OCR noise)
    for pat in [
        r'(?:OIB|DIB|OlB)\s*(?:Kupca|kupca|partnera)?[:\s]*([0-9?Oo ]{10,14})',
    ]:
        m = re.search(pat, text_clean, re.IGNORECASE)
        if m:
            raw = m.group(1)
            # Replace OCR noise: ? → 0, O/o → 0, space → ''
            cleaned = re.sub(r'[?Oo]', '0', raw)
            cleaned = re.sub(r'[^0-9]', '', cleaned)
            if len(cleaned) >= 10:
                result['partner_oib'] = cleaned[:11]
                break

    # Fallback: standalone 11-digit sequence (not part of longer number)
    if not result['partner_oib']:
        for m in re.finditer(r'(?<![0-9])(\d{11})(?![0-9])', text_clean):
            candidate = m.group(1)
            # Skip if it looks like a timestamp or other non-OIB number
            context = text_clean[max(0,m.start()-20):m.start()]
            if not re.search(r'(?:AID|ARPC|JIR|kod|ISA)', context, re.I):
                result['partner_oib'] = candidate
                break

    # ── Dates ────────────────────────────────────────────────────
    # Find invoice date — look near "Datum" or "datum" or isporuka
    # but exclude kartična transakcija section dates
    # Split text at KARTIČNA TRANSAKCIJA to avoid picking up card dates first
    main_text = text_clean.split('KARTIČNA')[0] if 'KARTIČNA' in text_clean else text_clean
    
    date_pat = r'\b(\d{1,2}\.\d{1,2}\.\d{4})\b'
    all_dates = re.findall(date_pat, main_text)
    if all_dates:
        result['invoice_date'] = all_dates[0]
    
    # Fallback: search full text if no date found in main section
    if not result['invoice_date']:
        all_dates_full = re.findall(r'\b(\d{1,2}\.\d{1,2}\.\d{4})\b', text_clean)
        if all_dates_full:
            result['invoice_date'] = all_dates_full[0]

    # ── Amount ───────────────────────────────────────────────────
    # Priority order: explicit IZNOS line, then ZA PLATITI + next line, then Način plaćanja
    amount_found = False

    # 1. IZNOS: 34,87 EUR (most explicit)
    m = re.search(r'IZNOS[:\s]+([\d]+[.,][\d]{2})\s*(?:EUR|€|FUR|HRK)?', text_clean, re.IGNORECASE)
    if m:
        raw = m.group(1).replace(',', '.')
        try:
            result['amount_total'] = float(raw)
            amount_found = True
        except: pass

    # 2. ZA PLATITI — value on same or next line, before big numbers
    if not amount_found:
        # Find ZA PLATITI line index
        for i, line in enumerate(lines):
            if re.search(r'ZA\s*PLATITI', line, re.I):
                # Check same line and next 2 lines for amount
                check_lines = lines[i:i+3]
                for cl in check_lines:
                    m = re.search(r'([\d]+[.,][\d]{2})\s*(?:EUR|€|FUR)?', cl)
                    if m:
                        raw = m.group(1).replace(',', '.')
                        try:
                            val = float(raw)
                            if 0 < val < 100000:  # sanity check
                                result['amount_total'] = val
                                amount_found = True
                                break
                        except: pass
                if amount_found:
                    break

    # 3. Način plaćanja: kartica XX,XX
    if not amount_found:
        m = re.search(r'[Nn]a[čc]in\s*pla[ćc]anja[^\n]*?([\d]+[.,][\d]{2})', text_clean)
        if m:
            raw = m.group(1).replace(',', '.')
            try:
                val = float(raw)
                if 0 < val < 100000:
                    result['amount_total'] = val
                    amount_found = True
            except: pass

    # 4. Vrijed. bez PDV + PDV amount
    if not amount_found:
        m = re.search(r'Vrijed\.?\s*bez\s*PDV[:\s]*([\d.,]+)', text_clean, re.I)
        if m:
            bez = m.group(1).replace(',', '.')
            # Look for PDV amount nearby
            pdv = re.search(r'(?:PDV|porez)[:\s]*([\d.,]+)', text_clean[text_clean.find('bez PDV'):text_clean.find('bez PDV')+200], re.I)
            try:
                val = float(bez)
                if pdv:
                    val += float(pdv.group(1).replace(',','.'))
                if 0 < val < 100000:
                    result['amount_total'] = round(val, 2)
                    amount_found = True
            except: pass

    # ── Partner name ────────────────────────────────────────────
    # First non-empty line that looks like a company name
    company_suffixes = ['d.o.o', 'd.d', 'j.d.o.o', 'obrt', 'ltd', 'gmbh', ' ad', 's.p.', 's.r.o']
    known_companies = ['petrol', 'konzum', 'dm', 'lidl', 'kaufland', 'spar', 'tommy', 'ina', 'hep',
                       'hrvatska', 'telekom', 't-com', 'optima', 'a1 ', 'vip', 'iskon']
    conn_skip = get_db()
    _company_skip = conn_skip.execute("SELECT value FROM settings WHERE key='company_name'").fetchone()
    conn_skip.close()
    _company_name = (_company_skip['value'].lower() if _company_skip and _company_skip['value'] else '')
    skip_words = ['oib', 'iban', 'račun', 'invoice', 'fisk', 'datum', 'ukupno', 'plaćanje',
                  'kartica', 'pdv'] + ([_company_name] if _company_name else [])

    for line in lines[:15]:
        l = line.strip()
        if len(l) < 4: continue
        if re.search(r'^[\d\s.,]+$', l): continue
        if any(sw in l.lower() for sw in skip_words): continue

        if any(sfx in l.lower() for sfx in company_suffixes) or \
           any(kw in l.lower() for kw in known_companies):
            result['partner_name'] = re.sub(r'[^\w\s.,&\-]', '', l).strip()[:80]
            break
    
    # Fallback: first meaningful line
    if not result['partner_name']:
        for line in lines[:8]:
            l = line.strip()
            if len(l) < 4: continue
            if re.search(r'^[\d\s.,]+$', l): continue
            if any(sw in l.lower() for sw in skip_words): continue
            if not re.search(r'\d{2}\.\d{2}\.\d{4}', l):
                result['partner_name'] = re.sub(r'[:\-]+$', '', l).strip()[:80]
                break

    return result


def make_stored_filename(partner_name, invoice_date):
    """Generate filename like 'Naziv partnera DD-MM-YYYY.pdf'"""
    import re
    # Parse date to DD-MM-YYYY format
    date_str = invoice_date or ''
    # Try to normalize
    date_clean = re.sub(r'[.\-/]', '-', date_str).strip()
    parts = date_clean.split('-')
    if len(parts) == 3:
        if len(parts[0]) == 4:  # YYYY-MM-DD
            date_clean = f"{parts[2]}-{parts[1]}-{parts[0]}"
        else:  # DD-MM-YYYY already
            date_clean = f"{parts[0]}-{parts[1]}-{parts[2]}"

    safe_partner = re.sub(r'[^\w\s\-]', '', partner_name or 'Nepoznat')
    safe_partner = safe_partner.strip()[:50]
    return f"{safe_partner} {date_clean}.pdf" if date_clean else f"{safe_partner}.pdf"


@app.route('/invoices')
@require_perm('can_view_invoices')
def invoice_list():
    audit('view', module='ulazni_racuni', entity='list')
    conn = get_db()
    invoices = conn.execute('''
        SELECT i.*, u.username as created_by_username
        FROM invoices i
        LEFT JOIN users u ON u.id = i.created_by
        WHERE i.is_deleted=0 OR i.is_deleted IS NULL
        ORDER BY i.invoice_date DESC, i.id DESC
    ''').fetchall()
    cards = conn.execute("SELECT * FROM bank_cards WHERE is_active=1").fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    suppliers = conn.execute("SELECT * FROM suppliers ORDER BY name").fetchall()
    conn.close()
    return render_template('invoice_list.html',
                           invoices=rows_to_dicts(invoices),
                           cards=rows_to_dicts(cards),
                           suppliers=rows_to_dicts(suppliers),
                           active='invoices',
                           settings=settings)

@app.route('/invoices/deleted')
@require_perm('can_view_invoices')
def invoice_deleted_list():
    conn = get_db()
    invoices = conn.execute('''
        SELECT i.*, u.username as created_by_username
        FROM invoices i
        LEFT JOIN users u ON u.id = i.created_by
        WHERE i.is_deleted=1
        ORDER BY i.deleted_at DESC, i.id DESC
    ''').fetchall()
    settings = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings").fetchall()}
    conn.close()
    return render_template('invoice_deleted.html',
                           invoices=rows_to_dicts(invoices),
                           active='invoices',
                           settings=settings)


@app.route('/invoices/upload', methods=['POST'])
@require_perm('can_edit_invoices')
def invoice_upload():
    """Upload file, run OCR, return extracted data for review form."""
    import os, tempfile
    if 'file' not in request.files:
        return jsonify({'error': 'Nema datoteke'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Nema naziva datoteke'}), 400

    # Save to temp
    ext = os.path.splitext(f.filename)[1].lower()
    allowed = {'.pdf', '.jpg', '.jpeg', '.png', '.tiff', '.tif'}
    if ext not in allowed:
        return jsonify({'error': f'Format {ext} nije podržan. Koristite PDF, JPG ili PNG.'}), 400

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    f.save(tmp.name)
    tmp.close()

    # OCR
    mime = 'application/pdf' if ext == '.pdf' else f'image/{ext.lstrip(".")}'
    ocr_text = ocr_file(tmp.name, mime)
    os.unlink(tmp.name)

    # Parse
    parsed = parse_invoice_data(ocr_text)
    parsed['ocr_raw'] = ocr_text[:2000]
    parsed['original_filename'] = f.filename

    return jsonify({'success': True, 'data': parsed})


@app.route('/invoices/save', methods=['POST'])
@require_perm('can_edit_invoices')
def invoice_save():
    """Save invoice record and copy file to configured storage folder."""
    import os, shutil, tempfile
    from datetime import datetime as dt

    data = request.form
    file = request.files.get('file')

    invoice_id = data.get('id')
    partner_name = data.get('partner_name', '').strip()
    invoice_date = data.get('invoice_date', '').strip()
    due_date = data.get('due_date', '').strip()
    invoice_number = data.get('invoice_number', '').strip()
    partner_oib = data.get('partner_oib', '').strip()
    amount_total = data.get('amount_total', '').strip()
    notes = data.get('notes', '').strip()
    ocr_raw = data.get('ocr_raw', '')

    try:
        amount_f = float(amount_total.replace(',', '.')) if amount_total else None
    except:
        amount_f = None

    # Determine year/month from invoice_date
    import re
    year, month = dt.now().year, dt.now().month
    if invoice_date:
        parts = re.split(r'[.\-/]', invoice_date)
        try:
            if len(parts) == 3:
                if len(parts[0]) == 4:
                    year, month = int(parts[0]), int(parts[1])
                else:
                    year, month = int(parts[2]), int(parts[1])
        except:
            pass

    now = dt.now().isoformat()
    conn = get_db()

    if invoice_id:
        conn.execute("""UPDATE invoices SET partner_name=?, partner_oib=?, invoice_number=?,
                     invoice_date=?, due_date=?, amount_total=?, notes=?, updated_at=? WHERE id=?""",
                     (partner_name, partner_oib, invoice_number, invoice_date, due_date,
                      amount_f, notes, now, invoice_id))
        conn.commit()
        conn.close()
        audit('edit', module='ulazni_racuni', entity='invoice', entity_id=int(invoice_id),
              detail=f'{partner_name} · {invoice_number}')
        return jsonify({'success': True, 'id': int(invoice_id)})

    # New invoice — copy file to storage
    stored_filename = None
    stored_path = None
    original_filename = data.get('original_filename', '')

    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        fn = make_stored_filename(partner_name, invoice_date)
        if not fn.endswith('.pdf') and ext != '.pdf':
            fn = fn.replace('.pdf', ext)

        try:
            folder = get_invoice_folder(year, month)
            dest = os.path.join(folder, fn)
            # Avoid overwrite
            base, ex = os.path.splitext(dest)
            counter = 1
            while os.path.exists(dest):
                dest = f"{base}_{counter}{ex}"
                counter += 1
            file.save(dest)
            stored_filename = os.path.basename(dest)
            stored_path = dest
        except Exception as e:
            conn.close()
            return jsonify({'error': f'Greška pri spremanju datoteke: {e}'}), 500

    user = get_current_user()
    created_by = user.get('user_id') if user else None
    conn.execute("""INSERT INTO invoices (partner_name, partner_oib, invoice_number,
                 invoice_date, due_date, amount_total, original_filename, stored_filename,
                 stored_path, notes, ocr_raw, created_by, created_at, updated_at)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (partner_name, partner_oib, invoice_number, invoice_date, due_date,
                  amount_f, original_filename, stored_filename, stored_path,
                  notes, ocr_raw, created_by, now, now))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    audit('create', module='ulazni_racuni', entity='invoice', entity_id=new_id,
          detail=f'{partner_name} · {invoice_number}')
    return jsonify({'success': True, 'id': new_id})


@app.route('/api/invoices/<int:inv_id>/note', methods=['POST'])
@require_perm('can_edit_invoices')
def invoice_note(inv_id):
    from datetime import datetime as dt
    data = request.json
    conn = get_db()
    conn.execute("UPDATE invoices SET notes=?, updated_at=? WHERE id=?",
                 (data.get('notes', ''), dt.now().isoformat(), inv_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/invoices/<int:inv_id>', methods=['GET'])
@require_perm('can_view_invoices')
def invoice_get(inv_id):
    conn = get_db()
    inv = conn.execute('''
        SELECT i.*, u.username as created_by_username
        FROM invoices i LEFT JOIN users u ON u.id = i.created_by
        WHERE i.id=?
    ''', (inv_id,)).fetchone()
    conn.close()
    if not inv:
        return jsonify({'error': 'Nije pronađen'}), 404
    return jsonify(row_to_dict(inv))

@app.route('/api/invoices/<int:inv_id>/payment', methods=['POST'])
@require_perm('can_edit_invoices')
def invoice_payment(inv_id):
    _pay_data = request.json or {}
    _pay_action = 'payment' if _pay_data.get('is_paid') else 'payment_cancel'
    audit(_pay_action, module='ulazni_racuni', entity='invoice', entity_id=inv_id,
          detail=f'Plaćanje: {_pay_data.get("paid_card_last4","")} · {_pay_data.get("paid_at","")}')
    data = request.json
    is_paid = 1 if data.get('is_paid') else 0
    paid_at = data.get('paid_at', '')
    paid_card = data.get('paid_card_last4', '')
    from datetime import datetime as dt
    conn = get_db()
    conn.execute("UPDATE invoices SET is_paid=?, paid_at=?, paid_card_last4=?, updated_at=? WHERE id=?",
                 (is_paid, paid_at, paid_card, dt.now().isoformat(), inv_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/director-signature')
@login_required
def director_signature_info():
    """Vrati URL direktorovog potpisa za preview u liquidate modalu."""
    conn = get_db()
    director = conn.execute("SELECT * FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    conn.close()
    if not director or not director['signature_path']:
        return jsonify({'url': None})
    return jsonify({'url': f"/uploads/{director['signature_path']}"})

@app.route('/uploads/<path:filename>')
@login_required
def serve_upload(filename):
    """Serviraj uploadane datoteke (potpisi, itd.)"""
    import os
    from flask import send_from_directory
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/invoices/<int:inv_id>/original-pdf')
@require_perm('can_view_invoices')
def invoice_original_pdf(inv_id):
    """Serve the ORIGINAL (pre-liquidation) PDF for preview."""
    import os
    conn = get_db()
    inv = conn.execute("SELECT stored_path FROM invoices WHERE id=?", (inv_id,)).fetchone()
    conn.close()
    if not inv or not inv['stored_path'] or not os.path.exists(inv['stored_path']):
        return "Datoteka nije pronađena", 404
    path = inv['stored_path']
    ext = os.path.splitext(path)[1].lower()
    # Ako je slika, konvertiraj u PDF za prikaz
    if ext in ('.jpg', '.jpeg', '.png', '.tiff', '.tif'):
        try:
            import fitz
            img_doc = fitz.open(path)
            pdf_bytes = img_doc.convert_to_pdf()
            img_doc.close()
            from flask import Response
            return Response(pdf_bytes, mimetype='application/pdf',
                          headers={'Content-Disposition': 'inline'})
        except Exception as e:
            return f"Greška: {e}", 500
    return send_file(path, mimetype='application/pdf',
                    download_name=os.path.basename(path), as_attachment=False)

@app.route('/api/invoices/<int:inv_id>/liquidate', methods=['POST'])
@require_perm('can_liquidate_invoices')
def invoice_liquidate(inv_id):
    """Stamp LIKVIDIRANO + director signature on the PDF with custom positions."""
    audit('liquidate', module='ulazni_racuni', entity='invoice', entity_id=inv_id,
          detail='Likvidacija — žig i potpis na PDF')
    import os
    from datetime import datetime as dt
    import fitz

    conn = get_db()
    inv = conn.execute("SELECT * FROM invoices WHERE id=?", (inv_id,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Račun nije pronađen'}), 404
    if inv['is_liquidated']:
        conn.close()
        return jsonify({'error': 'Račun je već likvidiran'}), 400

    inv = row_to_dict(inv)
    stored_path = inv.get('stored_path')
    if not stored_path or not os.path.exists(stored_path):
        conn.close()
        return jsonify({'error': 'Originalna datoteka nije pronađena na disku'}), 404

    # Pozicije iz request body (relativne 0-1, od lijevog gornjeg kuta)
    data = request.json or {}
    # stamp: {x, y, w, h} — sve relativno na dimenzije stranice
    stamp_pos  = data.get('stamp',  {'x': 0.65, 'y': 0.15, 'w': 0.30, 'h': 0.10})
    sig_pos    = data.get('sig',    {'x': 0.65, 'y': 0.25, 'w': 0.25, 'h': 0.08})
    page_idx   = int(data.get('page', -1))  # -1 = zadnja stranica

    director = conn.execute("SELECT * FROM employees WHERE is_direktor=1 LIMIT 1").fetchone()
    dir_sig_path = None
    if director and director['signature_path']:
        sig_full = os.path.join(UPLOAD_FOLDER, director['signature_path'])
        if os.path.exists(sig_full):
            dir_sig_path = sig_full

    now_str = dt.now().strftime('%d.%m.%Y. %H:%M')
    user = get_current_user()

    base, ext = os.path.splitext(stored_path)
    liq_path = base + '_LIKVIDIRANO.pdf'

    try:
        if ext.lower() in ('.jpg', '.jpeg', '.png', '.tiff', '.tif'):
            img_doc = fitz.open(stored_path)
            pdf_bytes = img_doc.convert_to_pdf()
            doc = fitz.open("pdf", pdf_bytes)
        else:
            doc = fitz.open(stored_path)

        # Odaberi stranicu (page_idx, ili zadnja ako -1)
        total_pages = len(doc)
        pidx = page_idx if page_idx >= 0 else total_pages - 1
        pidx = max(0, min(pidx, total_pages - 1))
        page = doc[pidx]
        pw, ph = page.rect.width, page.rect.height

        # Žig LIKVIDIRANO
        sx = stamp_pos['x'] * pw
        sy = stamp_pos['y'] * ph
        sw = stamp_pos['w'] * pw
        sh = stamp_pos['h'] * ph

        text = "LIKVIDIRANO"
        char_count = len(text)

        # Koristi fontSize koji je frontend izračunao i poslao (u PDF pt)
        if stamp_pos.get('fontSize'):
            fontsize = float(stamp_pos['fontSize'])
        else:
            # Fallback auto-izračun
            target_w = sw * 0.70
            fontsize = target_w / (char_count * 0.6)
            fontsize = min(fontsize, sh * 0.75)
            fontsize = max(8, fontsize)

        # Vertikalno centrirano (insert_text y je baseline)
        text_y = sy + sh * 0.55 + fontsize * 0.35
        # Horizontalno centrirano
        approx_text_w = char_count * 0.6 * fontsize
        text_x = sx + (sw - approx_text_w) / 2

        page.insert_text(
            fitz.Point(text_x, text_y),
            text,
            fontsize=fontsize, color=(0.8, 0.1, 0.1), fontname="helv"
        )

        # Potpis direktora
        if dir_sig_path:
            try:
                sigx = sig_pos['x'] * pw
                sigy = sig_pos['y'] * ph
                sigw = sig_pos['w'] * pw
                sigh = sig_pos['h'] * ph
                sig_rect = fitz.Rect(sigx, sigy, sigx + sigw, sigy + sigh)
                page.insert_image(sig_rect, filename=dir_sig_path)
            except:
                pass

        doc.save(liq_path)
        doc.close()

        conn.execute("""UPDATE invoices SET is_liquidated=1, liquidated_at=?,
                     liquidated_by=?, liquidated_pdf_path=?, updated_at=? WHERE id=?""",
                     (dt.now().isoformat(), user['user_id'] if user else None,
                      liq_path, dt.now().isoformat(), inv_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})

    except Exception as e:
        conn.close()
        return jsonify({'error': f'Greška pri likvidaciji: {e}'}), 500


@app.route('/invoices/<int:inv_id>/pdf')
@require_perm('can_view_invoices')
def invoice_pdf(inv_id):
    """Serve the liquidated PDF or original."""
    import os
    conn = get_db()
    inv = conn.execute("SELECT * FROM invoices WHERE id=?", (inv_id,)).fetchone()
    conn.close()
    if not inv:
        return "Not found", 404
    inv = row_to_dict(inv)
    # Prefer liquidated PDF
    path = inv.get('liquidated_pdf_path') if inv.get('is_liquidated') else inv.get('stored_path')
    if not path or not os.path.exists(path):
        return "Datoteka nije pronađena", 404
    fn = os.path.basename(path)
    return send_file(path, mimetype='application/pdf', download_name=fn, as_attachment=False)


@app.route('/api/invoices/<int:inv_id>', methods=['DELETE'])
@require_perm('can_edit_invoices')
def invoice_delete(inv_id):
    audit('delete', module='ulazni_racuni', entity='invoice', entity_id=inv_id)
    import os
    from datetime import datetime as dt
    conn = get_db()
    inv = conn.execute("SELECT * FROM invoices WHERE id=?", (inv_id,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Račun nije pronađen'}), 404
    inv = dict(inv)
    stored_path = inv.get('stored_path')
    new_stored_path = stored_path
    if stored_path and os.path.exists(stored_path):
        dir_part = os.path.dirname(stored_path)
        base, ext = os.path.splitext(os.path.basename(stored_path))
        new_filename = f"{base} BRISANO{ext}"
        new_stored_path = os.path.join(dir_part, new_filename)
        try:
            os.rename(stored_path, new_stored_path)
        except Exception as e:
            conn.close()
            return jsonify({'error': f'Greška pri preimenovanju datoteke: {e}'}), 500
    now = dt.now().isoformat()
    new_fname = os.path.basename(new_stored_path) if new_stored_path else inv.get('stored_filename')
    conn.execute(
        """UPDATE invoices SET is_deleted=1, deleted_at=?, stored_path=?,
           stored_filename=?, updated_at=? WHERE id=?""",
        (now, new_stored_path, new_fname, now, inv_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# Clients API
@app.route('/api/clients', methods=['GET'])
@login_required
def api_clients_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))

@app.route('/api/clients', methods=['POST'])
@login_required
def api_client_create():
    data = request.json
    if not data.get('name'):
        return jsonify({'error': 'Naziv je obavezan'}), 400
    conn = get_db()
    conn.execute("INSERT INTO clients (name, address, oib, is_client) VALUES (?,?,?,0)",
                 (data['name'].strip(), data.get('address','').strip(), data.get('oib','').strip()))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({'success': True, 'id': new_id})

@app.route('/api/clients/<int:cid>', methods=['PUT'])
@login_required
def api_client_update(cid):
    data = request.json
    conn = get_db()
    # Dozvoli ažuriranje svih polja uključujući is_client i is_default
    allowed = ['name', 'address', 'oib', 'is_client', 'is_default']
    updates = {k: v for k, v in data.items() if k in allowed}
    if updates:
        sets = ', '.join(f"{k}=?" for k in updates)
        conn.execute(f"UPDATE clients SET {sets} WHERE id=?", list(updates.values()) + [cid])
    conn.commit()
    item = conn.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    conn.close()
    return jsonify(dict(item) if item else {'success': True})

@app.route('/api/clients/<int:cid>', methods=['DELETE'])
@login_required
def api_client_delete(cid):
    conn = get_db()
    conn.execute("DELETE FROM clients WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ─── LOANS MODULE ─────────────────────────────────────────────────────────────

@app.route('/loans')
@require_perm('can_view_loans')
def loans_list():
    conn = get_db()
    loans = conn.execute("SELECT * FROM loans ORDER BY total_amount DESC").fetchall()
    loans_data = []
    import json as _json
    for loan in loans:
        loan_d = row_to_dict(loan)
        payments = conn.execute("SELECT * FROM loan_payments WHERE loan_id=?", (loan['id'],)).fetchall()
        loan_d['payments'] = rows_to_dicts(payments)
        # Ako postoji schedule, koristi označene stavke
        schedule = []
        try: schedule = _json.loads(loan_d.get('schedule_json') or '[]')
        except: pass
        today = datetime.now().strftime('%Y-%m-%d')
        if schedule:
            # Koristi status iz rasporeda (checkbox)
            repaid = sum(float(s.get('amount') or 0) for s in schedule
                        if s.get('paid') and s.get('type') != 'conversion')
            converted = sum(float(s.get('amount') or 0) for s in schedule
                           if s.get('paid') and s.get('type') == 'conversion')
        else:
            # Bez rasporeda: jednokratne i mjesečne do danas
            repaid = _calc_loan_repaid(loan_d, payments)
            # Konverzije čiji datum <= danas
            converted = sum(float(p['amount'] or 0) for p in rows_to_dicts(payments)
                           if p.get('payment_type') == 'conversion'
                           and p.get('payment_date','') <= today)
        loan_d['total_repaid'] = round(repaid + converted, 2)
        loan_d['total_converted'] = round(converted, 2)
        loan_d['remaining'] = round(loan['total_amount'] - loan_d['total_repaid'], 2)
        loans_data.append(loan_d)
    conn.close()
    return render_template('loans_list.html', loans=loans_data, active='loans')

@app.route('/loans/new')
@require_perm('can_edit_loans')
def loan_new():
    current_user = get_current_user() or {}
    can_lock = current_user.get('is_admin') or current_user.get('can_lock_loans', 0)
    return render_template('loan_form.html', loan=None, payments=[], active='loans', can_lock=can_lock)

@app.route('/loans/<int:loan_id>/edit')
@require_perm('can_view_loans')
def loan_edit(loan_id):
    current_user = get_current_user() or {}
    conn = get_db()
    loan = conn.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone()
    if not loan: return redirect(url_for('loans_list'))
    payments = conn.execute("SELECT * FROM loan_payments WHERE loan_id=? ORDER BY payment_date, recurring_start", (loan_id,)).fetchall()
    conn.close()
    import json as _json
    loan_d = row_to_dict(loan)
    try: loan_d['schedule'] = _json.loads(loan_d.get('schedule_json') or '[]')
    except: loan_d['schedule'] = []
    return render_template('loan_form.html', loan=loan_d,
                           payments=rows_to_dicts(payments), active='loans',
                           can_lock=current_user.get('is_admin') or current_user.get('can_lock_loans', 0))

@app.route('/api/loans/<int:loan_id>/lock', methods=['POST'])
@login_required
def loan_lock(loan_id):
    user = get_current_user()
    if not user or (not user.get('is_admin') and not user.get('can_lock_loans')):
        return jsonify({'error': 'Nemate pravo zaključavanja'}), 403
    data = request.json or {}
    locked = 1 if data.get('lock') else 0
    conn = get_db()
    conn.execute("UPDATE loans SET is_locked=?, updated_at=? WHERE id=?",
                 (locked, datetime.now().isoformat(), loan_id))
    conn.commit()
    conn.close()
    action = 'lock' if locked else 'unlock'
    audit(action, module='pozajmice', entity='loan', entity_id=loan_id)
    return jsonify({'success': True, 'is_locked': locked})

@app.route('/api/loans', methods=['POST'])
@require_perm('can_edit_loans')
def loan_save():
    data = request.json
    conn = get_db()
    loan_id = data.get('id')
    fields = {
        'name': data.get('name','').strip(),
        'total_amount': float(data.get('total_amount') or 0),
        'loan_date': data.get('loan_date',''),
        'interest_rate': float(data.get('interest_rate') or 0),
        'repayment_start': data.get('repayment_start',''),
        'repayment_end': data.get('repayment_end',''),
        'notes': data.get('notes',''),
        'updated_at': datetime.now().isoformat(),
    }
    import json as _json
    fields['schedule_json'] = _json.dumps(data.get('schedule', []))

    if loan_id:
        sets = ', '.join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE loans SET {sets} WHERE id=?", list(fields.values()) + [int(loan_id)])
    else:
        fields['created_at'] = datetime.now().isoformat()
        cols = ', '.join(fields.keys())
        ph = ', '.join('?' for _ in fields)
        c = conn.execute(f"INSERT INTO loans ({cols}) VALUES ({ph})", list(fields.values()))
        loan_id = c.lastrowid

    # Spremi uplate
    payments = data.get('payments', [])
    conn.execute("DELETE FROM loan_payments WHERE loan_id=?", (loan_id,))
    for p in payments:
        conn.execute("""INSERT INTO loan_payments
            (loan_id, payment_type, amount, payment_date, recurring_day,
             recurring_start, recurring_end, description, category)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (loan_id, p.get('payment_type'), float(p.get('amount') or 0),
             p.get('payment_date',''), p.get('recurring_day'),
             p.get('recurring_start',''), p.get('recurring_end',''),
             p.get('description',''), p.get('category','glavnica')))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': loan_id})

@app.route('/api/loans/<int:loan_id>', methods=['DELETE'])
@require_perm('can_edit_loans')
def loan_delete(loan_id):
    conn = get_db()
    loan = conn.execute("SELECT is_locked FROM loans WHERE id=?", (loan_id,)).fetchone()
    if loan and loan['is_locked']:
        conn.close()
        return jsonify({'error': 'Pozajmica je zaključana. Otključajte plan prije brisanja.'}), 400
    conn.execute("DELETE FROM loans WHERE id=?", (loan_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Work fund API
@app.route('/api/work-fund/<int:year>', methods=['GET'])
@login_required
def work_fund_get(year):
    conn = get_db()
    rows = conn.execute("SELECT * FROM work_fund WHERE year=? ORDER BY month", (year,)).fetchall()
    conn.close()
    return jsonify({'rows': rows_to_dicts(rows)})

@app.route('/api/work-fund', methods=['POST'])
@admin_required
def work_fund_save():
    data = request.json
    year = int(data.get('year'))
    rows = data.get('rows', [])
    conn = get_db()
    for r in rows:
        conn.execute("""INSERT INTO work_fund (year,month,fond,radni,neradni,obracunskih,radnih_dana,neradnih_dana)
            VALUES (?,?,?,?,?,?,?,?)
            ON CONFLICT(year,month) DO UPDATE SET
            fond=excluded.fond, radni=excluded.radni, neradni=excluded.neradni,
            obracunskih=excluded.obracunskih, radnih_dana=excluded.radnih_dana,
            neradnih_dana=excluded.neradnih_dana""",
            (year, r['month'], r['fond'], r['radni'], r['neradni'],
             r['obracunskih'], r['radnih_dana'], r['neradnih_dana']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Storage settings
@app.route('/api/storage/get-path')
@login_required
def storage_get_path():
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key='invoice_storage_path'").fetchone()
    conn.close()
    return jsonify({'path': row['value'] if row else ''})

@app.route('/api/storage/test-path', methods=['POST'])
@login_required
def storage_test_path():
    import os
    data = request.json or {}
    path = data.get('path', '').strip()
    if not path:
        return jsonify({'exists': False, 'error': 'Putanja je prazna'})
    exists = os.path.isdir(path)
    parent_exists = os.path.isdir(os.path.dirname(path)) if not exists else False
    return jsonify({'exists': exists, 'parent_exists': parent_exists})

# Suppliers CRUD
@app.route('/api/suppliers', methods=['GET'])
@require_perm('can_view_invoices')
def suppliers_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM suppliers ORDER BY name").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))

@app.route('/api/suppliers', methods=['POST'])
@admin_required
def supplier_create():
    data = request.json
    if not data.get('name'):
        return jsonify({'error': 'Naziv je obavezan'}), 400
    conn = get_db()
    conn.execute("INSERT INTO suppliers (name, oib, address) VALUES (?,?,?)",
                 (data['name'].strip(), data.get('oib','').strip(), data.get('address','').strip()))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({'success': True, 'id': new_id})

@app.route('/api/suppliers/<int:sid>', methods=['PUT'])
@admin_required
def supplier_update(sid):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE suppliers SET name=?, oib=?, address=? WHERE id=?",
                 (data.get('name','').strip(), data.get('oib','').strip(), data.get('address','').strip(), sid))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/suppliers/<int:sid>', methods=['DELETE'])
@admin_required
def supplier_delete(sid):
    conn = get_db()
    conn.execute("DELETE FROM suppliers WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Bank cards CRUD
@app.route('/api/bank_cards', methods=['GET'])
@require_perm('can_view_invoices')
def bank_cards_list():
    conn = get_db()
    cards = conn.execute("SELECT * FROM bank_cards ORDER BY card_name").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(cards))

@app.route('/api/bank_cards', methods=['POST'])
@admin_required
def bank_card_create():
    data = request.json
    conn = get_db()
    conn.execute("INSERT INTO bank_cards (card_name, last4, card_type, is_active, notes) VALUES (?,?,?,?,?)",
                 (data.get('card_name'), data.get('last4'), data.get('card_type',''), 1, data.get('notes','')))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/bank_cards/<int:card_id>', methods=['PUT'])
@admin_required
def bank_card_update(card_id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE bank_cards SET card_name=?, last4=?, card_type=?, is_active=?, notes=? WHERE id=?",
                 (data.get('card_name'), data.get('last4'), data.get('card_type',''),
                  data.get('is_active', 1), data.get('notes',''), card_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/bank_cards/<int:card_id>', methods=['DELETE'])
@admin_required
def bank_card_delete(card_id):
    conn = get_db()
    conn.execute("DELETE FROM bank_cards WHERE id=?", (card_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})



@app.route('/audit-log/export')
@admin_required
def audit_log_export():
    import csv, io
    user_filter = request.args.get('user', '')
    module_filter = request.args.get('module', '')
    action_filter = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = "SELECT * FROM audit_log WHERE 1=1"
    params = []
    if user_filter:
        query += " AND username=?"; params.append(user_filter)
    if module_filter:
        query += " AND module=?"; params.append(module_filter)
    if action_filter:
        query += " AND action=?"; params.append(action_filter)
    if date_from:
        query += " AND created_at >= ?"; params.append(date_from)
    if date_to:
        query += " AND created_at <= ?"; params.append(date_to + ' 23:59:59')
    query += " ORDER BY created_at DESC"

    conn = get_db()
    logs = conn.execute(query, params).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Datum i vrijeme', 'Korisnik', 'User ID', 'Akcija', 'Modul', 'Entitet', 'Entitet ID', 'Detalj', 'IP adresa'])
    for log in logs:
        writer.writerow([
            log['created_at'], log['username'], log['user_id'],
            log['action'], log['module'], log['entity'],
            log['entity_id'], log['detail'], log['ip_address']
        ])

    output.seek(0)
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=audit_log_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'}
    )

@app.route('/audit-log')
@admin_required
def audit_log_page():
    conn = get_db()
    user_filter = request.args.get('user', '')
    module_filter = request.args.get('module', '')
    action_filter = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = "SELECT * FROM audit_log WHERE 1=1"
    params = []
    if user_filter:
        query += " AND username=?"; params.append(user_filter)
    if module_filter:
        query += " AND module=?"; params.append(module_filter)
    if action_filter:
        query += " AND action=?"; params.append(action_filter)
    if date_from:
        query += " AND created_at >= ?"; params.append(date_from)
    if date_to:
        query += " AND created_at <= ?"; params.append(date_to + ' 23:59:59')
    query += " ORDER BY created_at DESC LIMIT 1000"

    logs = conn.execute(query, params).fetchall()
    users = conn.execute("SELECT DISTINCT username FROM audit_log ORDER BY username").fetchall()
    modules = conn.execute("SELECT DISTINCT module FROM audit_log WHERE module IS NOT NULL ORDER BY module").fetchall()
    actions = conn.execute("SELECT DISTINCT action FROM audit_log ORDER BY action").fetchall()
    conn.close()
    return render_template('audit_log.html',
                           logs=rows_to_dicts(logs),
                           users=[r['username'] for r in users],
                           modules=[r['module'] for r in modules],
                           actions=[r['action'] for r in actions],
                           filters={'user': user_filter, 'module': module_filter,
                                    'action': action_filter, 'date_from': date_from, 'date_to': date_to},
                           active='audit')

if __name__ == '__main__':
    init_db()
    print("\n✅ Putni Nalog App pokrenuta!")
    print("🌐 Otvori: http://localhost:5050\n")
    app.run(debug=False, port=5050)

